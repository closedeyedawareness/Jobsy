"""Add IT, Procurement, Facilities functions + Supply Chain / Design / Admin
roles and synonyms to jobsy_reference_library.xlsx. In-place openpyxl."""
from openpyxl import load_workbook

WB = "jobsy_reference_library.xlsx"; TODAY = "2026-07-02"; SRC = "Seed v1 (functions)"
wb = load_workbook(WB)
def H(ws): return {c.value: i+1 for i, c in enumerate(ws[1]) if c.value is not None}
def add(ws, d):
    h = H(ws); r = ws.max_row + 1
    for k, v in d.items():
        if k in h: ws.cell(r, h[k], v)
def gov(owner): return {"Source": SRC, "Owner": owner, "Status": "Active",
                        "EffectiveFrom": TODAY, "UpdatedAt": TODAY}

# ── new salary bands (function, level) -> (grade, min, p25, p50, p75, max) ──
BANDS = {
 ("IT","Medior"):(5,42000,46000,50000,55000,60000), ("IT","Senior"):(8,65000,72000,80000,89000,98000),
 ("IT","Lead"):(11,90000,101000,115000,130000,145000),
 ("Procurement","Medior"):(5,44000,48000,52000,57000,62000), ("Procurement","Senior"):(8,62000,70000,78000,86000,95000),
 ("Procurement","Lead"):(12,95000,107000,120000,135000,150000),
 ("Facilities","Junior"):(2,28000,30000,33000,36000,40000), ("Facilities","Medior"):(4,38000,41000,45000,49000,54000),
 ("Facilities","Senior"):(7,55000,61000,68000,75000,82000),
}
for (fn, lvl), (gr, mn, p25, p50, p75, mx) in BANDS.items():
    add(wb["SalaryBands"], {"Function":fn,"Level":lvl,"Grade":gr,"Min":mn,"P25":p25,"P50":p50,
                            "P75":p75,"Max":mx,"Currency":"EUR", **gov("Reward")})

# ── new categories ──
for cat, fn, desc in [("IT & Infrastructure","IT","IT support, systems, networks and security."),
                      ("Procurement & Supply Chain","Procurement","Buying, sourcing and supplier management."),
                      ("Facilities & Office","Facilities","Office, reception and facilities management.")]:
    add(wb["Categories"], {"Category":cat,"Function":fn,"Description":desc, **gov("Job Architecture")})

# ── role definitions ──
# (jobid, title, function, level, grade, category, isco, iscotitle, esco, skills_src, next, next_title)
ROLES = [
 ("J-IT-01","IT Support Specialist","IT","Medior",4,"IT & Infrastructure","3512","Information and communications technology user support technicians","ICT help desk agent","J-CS-01","J-IT-02","System Administrator"),
 ("J-IT-02","System Administrator","IT","Medior",6,"IT & Infrastructure","2522","Systems administrators","system administrator","J-ENG-05","J-IT-03","Network Engineer"),
 ("J-IT-03","Network Engineer","IT","Senior",8,"IT & Infrastructure","2523","Computer network professionals","ICT network engineer","J-ENG-05","J-IT-05","IT Manager"),
 ("J-IT-04","Security Engineer","IT","Senior",9,"IT & Infrastructure","2529","Database and network professionals not elsewhere classified","ICT security technician","J-ENG-05","J-IT-05","IT Manager"),
 ("J-IT-05","IT Manager","IT","Lead",11,"IT & Infrastructure","1330","Information and communications technology service managers","ICT operations manager","J-ENG-04","J-EXEC-10","Chief Information Officer"),
 ("J-PRO-01","Buyer","Procurement","Medior",5,"Procurement & Supply Chain","3323","Buyers","purchaser","J-FIN-02","J-PRO-02","Procurement Specialist"),
 ("J-PRO-02","Procurement Specialist","Procurement","Senior",7,"Procurement & Supply Chain","3323","Buyers","procurement specialist","J-FIN-03","J-PRO-04","Procurement Manager"),
 ("J-PRO-03","Category Manager","Procurement","Senior",8,"Procurement & Supply Chain","1219","Business services and administration managers not elsewhere classified","category manager","J-SAL-04","J-PRO-04","Procurement Manager"),
 ("J-PRO-04","Procurement Manager","Procurement","Senior",9,"Procurement & Supply Chain","1219","Business services and administration managers not elsewhere classified","procurement manager","J-OPS-03","J-PRO-05","Head of Procurement"),
 ("J-PRO-05","Head of Procurement","Procurement","Lead",12,"Procurement & Supply Chain","1219","Business services and administration managers not elsewhere classified","procurement manager","J-OPS-05","J-EXEC-01","Chief Executive Officer"),
 ("J-FAC-01","Receptionist","Facilities","Junior",2,"Facilities & Office","4226","Receptionists (general)","receptionist","J-HR-01","J-FAC-02","Facilities Coordinator"),
 ("J-FAC-02","Facilities Coordinator","Facilities","Medior",4,"Facilities & Office","3341","Office supervisors","facilities manager","J-OPS-01","J-FAC-04","Facility Manager"),
 ("J-FAC-03","Office Manager","Facilities","Medior",5,"Facilities & Office","3341","Office supervisors","office manager","J-OPS-01","J-FAC-04","Facility Manager"),
 ("J-FAC-04","Facility Manager","Facilities","Senior",7,"Facilities & Office","1219","Business services and administration managers not elsewhere classified","facilities manager","J-OPS-03","J-OPS-05","Head of Operations"),
 ("J-OPS-06","Warehouse Manager","Operations","Senior",8,"Operations & Projects","1324","Supply, distribution and related managers","warehouse manager","J-OPS-03","J-OPS-05","Head of Operations"),
 ("J-OPS-07","Supply Chain Manager","Operations","Senior",9,"Operations & Projects","1324","Supply, distribution and related managers","supply chain manager","J-OPS-03","J-OPS-05","Head of Operations"),
 ("J-PRD-06","Senior Product Designer","Product","Senior",9,"Product & Design","2166","Graphic and multimedia designers","user experience designer","J-PRD-05","J-PRD-04","Head of Product"),
 ("J-PRD-07","UX Researcher","Product","Medior",6,"Product & Design","2166","Graphic and multimedia designers","user experience analyst","J-PRD-05","J-PRD-06","Senior Product Designer"),
]
PROFILES = {
 "J-IT-01":("First-line IT support for staff — hardware, software, accounts and access — resolving or escalating incidents.","Resolve 1st/2nd line IT tickets; Manage accounts and access; Set up hardware and software; Maintain the asset register; Escalate complex incidents","IT support; Troubleshooting; Ticketing; Access management; Hardware setup","End-user support; Onboarding IT","Jira Service Management; Microsoft 365; Intune; Active Directory; TeamViewer"),
 "J-IT-02":("Keeps servers, systems and cloud infrastructure running securely and reliably; automates administration.","Administer servers and cloud infrastructure; Manage backups and patching; Automate routine tasks; Monitor systems and uptime; Support security controls","System administration; Cloud infrastructure; Scripting; Backup & recovery; Monitoring","Windows/Linux administration; Microsoft 365 admin","Azure/AWS; PowerShell/Bash; Active Directory; Intune; Datadog"),
 "J-IT-03":("Designs, builds and maintains the corporate network and connectivity; ensures performance and security.","Design and maintain network infrastructure; Manage firewalls and VPNs; Monitor performance; Troubleshoot connectivity; Document topology","Networking; Firewalls; VPN; Routing & switching; Monitoring","Enterprise networking; Cloud networking","Cisco/Fortinet; Meraki; Wireshark; Azure networking"),
 "J-IT-04":("Protects the organisation's systems and data; runs security controls, monitoring and incident response.","Run security monitoring and response; Manage vulnerabilities; Implement security controls; Support compliance; Raise security awareness","Information security; Vulnerability management; Incident response; Identity & access; Compliance","Cloud security; SOC operations","SIEM; Okta; CrowdStrike; Nessus; Azure Security"),
 "J-IT-05":("Leads the IT function — support, systems, infrastructure and security — and the IT team, budget and vendors.","Lead the IT team and roadmap; Own IT budget and vendors; Ensure security and uptime; Set IT policy; Partner with the business","IT leadership; Vendor management; IT strategy; Security governance; Budgeting","Enterprise IT; ITIL","ServiceNow; Microsoft 365; Azure; Okta; Power BI"),
 "J-PRO-01":("Buys goods and services against requirements; places orders and manages suppliers day to day.","Place and track purchase orders; Source and compare suppliers; Negotiate basic terms; Maintain supplier data; Resolve delivery issues","Purchasing; Supplier management; Negotiation; ERP/procurement systems; Cost awareness","Operational buying; Indirect procurement","SAP Ariba; Coupa; Excel; ERP"),
 "J-PRO-02":("Runs sourcing and procurement for a category or area; drives value, quality and supplier performance.","Run sourcing and tenders; Negotiate contracts; Manage supplier performance; Deliver savings; Ensure compliance","Strategic sourcing; Negotiation; Contract management; Category analysis; Supplier performance","Strategic procurement; Tendering","SAP Ariba; Coupa; Excel; Power BI"),
 "J-PRO-03":("Owns a spend category end to end — strategy, sourcing and supplier relationships — to maximise value.","Own category strategy; Run sourcing events; Manage key suppliers; Deliver category savings; Partner with stakeholders","Category management; Sourcing strategy; Negotiation; Stakeholder management; Cost modelling","Category strategy; Supplier management","SAP Ariba; Coupa; Power BI; Excel"),
 "J-PRO-04":("Leads procurement operations and the buying team; owns supplier strategy, savings and compliance.","Lead the procurement team; Own supplier and savings strategy; Manage contracts and risk; Ensure compliance; Report to leadership","Procurement leadership; Supplier strategy; Contract & risk; Team leadership; Savings delivery","Procurement operations; Supplier risk","SAP Ariba; Coupa; Power BI; ERP"),
 "J-PRO-05":("Owns procurement strategy across the company; accountable for spend, savings, risk and the function.","Set procurement strategy; Own total spend and savings; Manage supplier risk; Build the procurement team; Advise the board","Procurement strategy; Spend management; Supplier risk; Leadership; Governance","Indirect & direct procurement; Global sourcing","SAP Ariba; Coupa; Power BI; Board portal"),
 "J-FAC-01":("First point of contact at reception — welcomes visitors, handles calls and post, and supports the office.","Welcome visitors and manage reception; Handle phone and post; Book meeting rooms; Support office logistics; Keep front-of-house tidy","Reception; Communication; Organisation; Office software; Hospitality","Front office; Visitor management","Microsoft 365; Outlook; Booking systems"),
 "J-FAC-02":("Coordinates day-to-day facilities — office services, suppliers and requests — keeping the workplace running.","Coordinate facilities requests; Manage office suppliers; Support health & safety; Handle access and keys; Assist with events","Facilities coordination; Supplier liaison; Organisation; Health & safety; Office software","Office services; Workplace support","Microsoft 365; Facilities/ticketing tools"),
 "J-FAC-03":("Runs the office end to end — services, suppliers, events and administration — and supports leadership.","Run office operations and services; Manage suppliers and budget; Organise events; Support leadership admin; Own health & safety","Office management; Vendor management; Organisation; Budgeting; Administration","Office management; Executive support","Microsoft 365; Excel; Booking/facilities tools"),
 "J-FAC-04":("Leads facilities and workplace services across sites — property, services, safety, budget and vendors.","Lead facilities and workplace services; Manage property and contracts; Own facilities budget; Ensure safety and compliance; Manage vendors","Facilities management; Property & contracts; Health & safety; Budgeting; Vendor management","Multi-site facilities; Workplace strategy","Planon; Microsoft 365; Excel; Power BI"),
 "J-OPS-06":("Runs warehouse operations — inbound, storage, picking and dispatch — with a focus on safety and efficiency.","Manage warehouse operations and team; Optimise inbound/outbound flow; Own stock accuracy; Ensure safety; Improve efficiency","Warehouse management; Inventory; Logistics; Team leadership; Safety","Warehousing; Distribution","WMS; ERP; Excel; Power BI"),
 "J-OPS-07":("Owns end-to-end supply chain — planning, procurement flow, logistics and inventory — balancing cost and service.","Own supply and demand planning; Optimise logistics and inventory; Manage S&OP; Balance cost and service; Lead improvement","Supply chain planning; S&OP; Logistics; Inventory; Analytics","End-to-end supply chain; S&OP","SAP; Kinaxis; Excel; Power BI"),
 "J-PRD-06":("Leads design on complex product areas — from research to polished UI — and lifts design quality and the system.","Lead product design end to end; Drive research and validation; Own the design system; Mentor designers; Raise design quality","Product design; UX research; Design systems; Prototyping; Mentoring","SaaS product design; Design systems","Figma; FigJam; Maze; Miro; Notion"),
 "J-PRD-07":("Plans and runs user research — qualitative and quantitative — turning insight into product decisions.","Plan and run user research; Recruit and interview users; Synthesise insight; Share findings; Shape product decisions","User research; Interviewing; Usability testing; Synthesis; Communication","UX research; Mixed methods","Maze; Dovetail; Figma; UserTesting"),
}
MGMT = {"Junior":"Individual Contributor","Medior":"Individual Contributor","Senior":"Individual Contributor"}
jobs_ws, band_ws, jp_ws, rsm_ws, cp_ws = wb["Jobs"], wb["SalaryBands"], wb["JobProfiles"], wb["RoleSkillMap"], wb["CareerPaths"]
rh = H(rsm_ws)
skills_by_job = {}
for r in range(2, rsm_ws.max_row+1):
    jid = rsm_ws.cell(r, rh["JobID"]).value
    if jid: skills_by_job.setdefault(str(jid).strip(), []).append(
        (rsm_ws.cell(r,rh["SkillID"]).value, rsm_ws.cell(r,rh["RequiredLevel"]).value, rsm_ws.cell(r,rh["SkillType"]).value))

for (jid,title,fn,lvl,grade,cat,isco,iscot,esco,src,nxt,nxt_title) in ROLES:
    mgmt = "People Manager" if lvl == "Lead" or "Manager" in title or "Head" in title else MGMT.get(lvl,"Individual Contributor")
    add(jobs_ws, {"JobID":jid,"StandardTitle":title,"Function":fn,"Level":lvl,"Category":cat,"Grade":grade,
                  "IscoGroup":isco,"IscoTitle":iscot,"EscoLabel":esco, **gov("Job Architecture")})
    desc,resp,req,spec,tools = PROFILES[jid]
    add(jp_ws, {"JobID":jid,"Description":desc,"KeyResponsibilities":resp,"RequiredSkills":req,
                "Specialisms":spec,"ManagementLevel":mgmt,"TypicalTools":tools, **gov("Job Architecture")})
    for sid,rl,stype in skills_by_job.get(src, []):
        add(rsm_ws, {"JobID":jid,"SkillID":sid,"RequiredLevel":rl,"SkillType":stype, **gov("Talent & Capability")})
    add(cp_ws, {"JobID":jid,"NextJobID":nxt,"NextRole":nxt_title, **gov("Job Architecture")})

# ── synonyms (new + remaps) ──
tm = wb["TitleMapping"]; th = H(tm)
existing = {str(tm.cell(r,th["ExistingTitle"]).value).strip(): r for r in range(2,tm.max_row+1) if tm.cell(r,th["ExistingTitle"]).value}
SYN = {
 # IT
 "Helpdesk Medewerker":"J-IT-01","Service Desk Analyst":"J-IT-01","IT Support Engineer":"J-IT-01","Helpdesk Analyst":"J-IT-01","1st Line Support":"J-IT-01","2nd Line Support":"J-IT-01","Servicedesk Medewerker":"J-IT-01",
 "System Administrator":"J-IT-02","Sysadmin":"J-IT-02","Systems Administrator":"J-IT-02","Systeembeheerder":"J-IT-02","Systeem Beheerder":"J-IT-02",
 "Network Engineer":"J-IT-03","Netwerkbeheerder":"J-IT-03","Network Administrator":"J-IT-03","Infrastructure Engineer":"J-IT-03",
 "Security Engineer":"J-IT-04","Security Specialist":"J-IT-04","Cyber Security Engineer":"J-IT-04","Information Security Officer":"J-IT-04","Security Analyst":"J-IT-04",
 "IT Manager":"J-IT-05","IT Director":"J-IT-05","Head of IT":"J-IT-05","ICT Manager":"J-IT-05","Hoofd ICT":"J-IT-05","IT Operations Manager":"J-IT-05","Hoofd IT":"J-IT-05",
 # Procurement
 "Buyer":"J-PRO-01","Inkoper":"J-PRO-01","Junior Buyer":"J-PRO-01","Operational Buyer":"J-PRO-01","Inkoopmedewerker":"J-PRO-01",
 "Procurement Specialist":"J-PRO-02","Senior Buyer":"J-PRO-02","Strategic Buyer":"J-PRO-02","Tactical Buyer":"J-PRO-02","Senior Inkoper":"J-PRO-02","Inkoop Specialist":"J-PRO-02",
 "Category Manager":"J-PRO-03","Category Lead":"J-PRO-03",
 "Procurement Manager":"J-PRO-04","Purchasing Manager":"J-PRO-04","Inkoop Manager":"J-PRO-04","Inkoopmanager":"J-PRO-04","Sourcing Manager":"J-PRO-04",
 "Head of Procurement":"J-PRO-05","Procurement Director":"J-PRO-05","Head of Purchasing":"J-PRO-05","Inkoopdirecteur":"J-PRO-05","Director of Procurement":"J-PRO-05",
 # Facilities
 "Receptioniste":"J-FAC-01","Front Office Medewerker":"J-FAC-01","Front Desk":"J-FAC-01","Receptie":"J-FAC-01",
 "Facilitair Medewerker":"J-FAC-02","Facilities Officer":"J-FAC-02","Facilitair Coordinator":"J-FAC-02",
 "Kantoormanager":"J-FAC-03","Office Coordinator":"J-FAC-03","Management Assistant":"J-FAC-03","Directiesecretaresse":"J-FAC-03","Directie Assistent":"J-FAC-03","Managementassistent":"J-FAC-03",
 "Facilitair Manager":"J-FAC-04","Head of Facilities":"J-FAC-04","Hoofd Facilitair":"J-FAC-04","Facilities Manager":"J-FAC-04",
 # Supply chain
 "Warehouse Manager":"J-OPS-06","Magazijnmanager":"J-OPS-06","Warehouse Supervisor":"J-OPS-06","Logistiek Manager":"J-OPS-06","Logistics Manager":"J-OPS-06",
 "Supply Chain Manager":"J-OPS-07","Supply Chain Lead":"J-OPS-07","S&OP Manager":"J-OPS-07","Head of Supply Chain":"J-OPS-07",
 "Logistiek Medewerker":"J-OPS-01","Planner":"J-OPS-01","Logistics Coordinator":"J-OPS-01",
 "Administratief Medewerker":"J-OPS-01",
 # Design
 "Senior Product Designer":"J-PRD-06","Senior UX Designer":"J-PRD-06","Senior UI Designer":"J-PRD-06","Lead Designer":"J-PRD-06","Design Lead":"J-PRD-06",
 "UX Researcher":"J-PRD-07","User Researcher":"J-PRD-07","UX Research":"J-PRD-07","Design Researcher":"J-PRD-07",
 "Graphic Designer":"J-PRD-05","Content Designer":"J-PRD-05","Service Designer":"J-PRD-05",
}
added = remapped = 0
for et,jid in SYN.items():
    if et in existing:
        tm.cell(existing[et], th["JobID"], jid); tm.cell(existing[et], th["UpdatedAt"], TODAY); remapped += 1
    else:
        add(tm, {"ExistingTitle":et,"JobID":jid, **gov("Data Quality")}); added += 1

wb.save(WB)
print(f"roles added: {len(ROLES)} | bands added: {len(BANDS)} | synonyms added: {added} remapped: {remapped}")
