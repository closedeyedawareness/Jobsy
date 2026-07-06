"""
Add a BenefitGroup taxonomy column to BenefitsCatalog, and a new
"Holiday Purchase/Sale (Verlofkoop)" category — genuinely missing until
now, referenced in the source material (Dow Chemical Netherlands: 25
additional purchasable days) but never modeled.

Groups (5): Perquisite, Insurance/Protection, Retirement, Leave, Pay Premium.
"Perquisite" covers the non-cash/quasi-cash extras a user specifically
called out: Flexible Benefits, EAP, Holiday Purchase/Sale, plus the other
perk-like categories (discounts, bicycle, company car, meal, mobility,
childcare, representation, education).
"""
import random

from openpyxl import load_workbook

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"

random.seed(20260709)

wb = load_workbook(WB)


def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


def add_column(ws, name, default=""):
    h = headers(ws)
    if name in h:
        return h[name]
    col = ws.max_column + 1
    from openpyxl.styles import Alignment, Font, PatternFill
    c = ws.cell(1, col, name)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="0E7C66")
    c.alignment = Alignment(horizontal="left", vertical="center")
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col).value is None:
            ws.cell(r, col, default)
    return col


GROUP = {
    "Pension": "Retirement",
    "Health Insurance": "Insurance/Protection",
    "Extra Leave": "Leave",
    "Wellness & Vitality": "Perquisite",
    "Mobility & Commute": "Perquisite",
    "Meal & Lunch": "Perquisite",
    "Disability Top-up": "Insurance/Protection",
    "Parental Leave Top-up": "Leave",
    "Education & Development": "Perquisite",
    "Flexible Benefits / Cafetariaregeling": "Perquisite",
    "Generation Pact / Vitality Scheme": "Leave",
    "Employee Discounts & Perks": "Perquisite",
    "Bicycle Plan": "Perquisite",
    "Company Car / Lease": "Perquisite",
    "Life & Risk Insurance": "Insurance/Protection",
    "Short-Term Disability (Sick Pay Top-up)": "Insurance/Protection",
    "Childcare Support": "Perquisite",
    "EAP / Psychosocial Support": "Perquisite",
    "Representation Allowance": "Perquisite",
    "Shift Allowance (Ploegentoeslag)": "Pay Premium",
    "Holiday Purchase/Sale (Verlofkoop)": "Perquisite",
}

# ── 1. New category: Holiday Purchase/Sale ─────────────────────────────────
CATEGORY = "Holiday Purchase/Sale (Verlofkoop)"
cat_ws = wb["BenefitsCatalog"]
ch = headers(cat_ws)
existing_cats = {cat_ws.cell(r, ch["Category"]).value for r in range(2, cat_ws.max_row + 1)}
existing_ids = [str(cat_ws.cell(r, ch["BenefitID"]).value) for r in range(2, cat_ws.max_row + 1)
                if cat_ws.cell(r, ch["BenefitID"]).value]
next_bid = max((int(x.split("-")[1]) for x in existing_ids if x.startswith("BEN-")), default=0) + 1

if CATEGORY not in existing_cats:
    cat_ws.append([
        f"BEN-{next_bid:02d}", CATEGORY,
        "Additional leave days purchasable (koop) beyond the base CAO/statutory entitlement", "days",
        "0-25 additional days purchasable, cost = daily wage rate", "No (CAO/company scheme)",
        "N/A (unpaid leave via salary deduction, not itself taxable income)",
        "Flexible scheme allowing employees to purchase (or sell back) additional leave days "
        "beyond their base entitlement, typically funded via gross salary deduction/addition.",
        "Seed v2 (benefits, taxonomy expansion)", "Reward", "Active", TODAY, TODAY,
    ])

# ── 2. BenefitGroup column on BenefitsCatalog ──────────────────────────────
add_column(cat_ws, "BenefitGroup")
ch = headers(cat_ws)
for r in range(2, cat_ws.max_row + 1):
    cat = cat_ws.cell(r, ch["Category"]).value
    if cat in GROUP:
        cat_ws.cell(r, ch["BenefitGroup"], GROUP[cat])

# ── 3. Synthetic baseline + real observation for Holiday Purchase/Sale ────
INDUSTRIES = ["IND-TECH", "IND-FIN", "IND-HLTH", "IND-MFG", "IND-RET", "IND-PUB", "IND-PSV", "IND-LOG"]
INDUSTRY_RICHNESS = {
    "IND-TECH": 1.25, "IND-FIN": 1.30, "IND-HLTH": 0.95, "IND-MFG": 0.95,
    "IND-RET": 0.75, "IND-PUB": 0.90, "IND-PSV": 1.15, "IND-LOG": 0.80,
}
BASELINE_MEAN = 10.0  # plausible NL-market days purchasable, before industry richness
JITTER = 0.55
N_OBS_PER_GROUP = 14

obs_ws = wb["BenefitsObservations"]
oh = headers(obs_ws)
existing_obs_ids = [str(obs_ws.cell(r, oh["ObsID"]).value) for r in range(2, obs_ws.max_row + 1)
                    if obs_ws.cell(r, oh["ObsID"]).value]
next_obs_id = max((int(x.split("-")[1]) for x in existing_obs_ids if x.startswith("BO-")), default=0) + 1
n_cols = obs_ws.max_column


def append_obs_row(industry_id, value, unit, source, company, notes, status="Active"):
    global next_obs_id
    row = [None] * n_cols
    row[oh["ObsID"] - 1] = f"BO-{next_obs_id:05d}"
    row[oh["IndustryID"] - 1] = industry_id
    row[oh["Category"] - 1] = CATEGORY
    row[oh["Value"] - 1] = value
    row[oh["Unit"] - 1] = unit
    row[oh["Currency"] - 1] = ""
    row[oh["Source"] - 1] = source
    row[oh["Owner"] - 1] = "Reward"
    row[oh["Status"] - 1] = status
    row[oh["EffectiveFrom"] - 1] = TODAY
    row[oh["UpdatedAt"] - 1] = TODAY
    if "CompanyName" in oh:
        row[oh["CompanyName"] - 1] = company
    if "Notes" in oh:
        row[oh["Notes"] - 1] = notes
    obs_ws.append(row)
    next_obs_id += 1


n_synthetic = 0
for industry in INDUSTRIES:
    group_mean = BASELINE_MEAN * INDUSTRY_RICHNESS[industry]
    for _ in range(N_OBS_PER_GROUP):
        noise = random.gauss(1.0, JITTER)
        value = max(0.0, group_mean * noise)
        append_obs_row(industry, round(value, 1), "days",
                       "Seed v2 (benefits, taxonomy expansion)", "", "", "Active")
        n_synthetic += 1

append_obs_row(
    "IND-MFG", 25.0, "days",
    "NL chemicals/pharma CAO comparison (user-supplied CSV, Dec 2025)", "Dow Chemical Netherlands",
    "CAO: base 28 days + up to 25 additional purchasable days ('28+25koop=53' max with full buy-up). "
    "25 (the purchasable ceiling) used here as the Value.", "Active",
)

# ── 4. LevelBenefitsFactors: flat ──────────────────────────────────────────
lf_ws = wb["LevelBenefitsFactors"]
lh = headers(lf_ws)
existing_pairs = {(lf_ws.cell(r, lh["Level"]).value, lf_ws.cell(r, lh["Category"]).value)
                   for r in range(2, lf_ws.max_row + 1)}
n_factors = 0
for level in ["Junior", "Medior", "Senior", "Lead"]:
    if (level, CATEGORY) in existing_pairs:
        continue
    lf_ws.append([level, CATEGORY, 1.0, "Seed v2 (benefits, taxonomy expansion)",
                 "Reward", "Active", TODAY, TODAY])
    n_factors += 1

wb.save(WB)
print(f"BenefitsCatalog: +1 category ({CATEGORY}), BenefitGroup tagged on {len(GROUP)} categories")
print(f"BenefitsObservations: +{n_synthetic} synthetic, +1 real")
print(f"LevelBenefitsFactors: +{n_factors} rows")
print("Saved", WB)
