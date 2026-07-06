"""
Blend real, cited company-level benefits data into BenefitsObservations,
alongside the existing self-built synthetic (Seed v1) distribution.

Source material: user-supplied CAO summaries (Dec 2025) for Shell Nederland
Raffinaderij, Zeeland Refinery, AkzoNobel, and Ecolab Production Netherlands —
all IND-MFG (chemicals/refining, CAO Metalektro-adjacent).

Confidence tiers (Status column):
  Active — Shell & Zeeland Refinery: sourced from a detailed markdown table
           quoting the CAO directly (loon/toeslagen/pensioen/verlof).
  Draft  — AkzoNobel & Ecolab: sourced from a summary-only pass (secondary,
           not the primary CAO text) — flagged for verification before
           being treated as fully confirmed.

Only Pension, Disability Top-up, and Parental Leave Top-up have real
publicly-documented figures from this material. Health Insurance, Extra
Leave, Wellness, Mobility, Meal, and Education remain undocumented for
these companies and are deliberately NOT fabricated here.

Adds a CompanyName + Notes column to BenefitsObservations (additive,
backward compatible with the existing synthetic rows, which get blank
values for both) and new BenefitsSources citations.
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"
HDR_FILL = PatternFill("solid", fgColor="0E7C66")
HDR_FONT = Font(bold=True, color="FFFFFF")


def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


def add_column(ws, name, default=""):
    h = headers(ws)
    if name in h:
        return h[name]
    col = ws.max_column + 1
    c = ws.cell(1, col, name)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col).value is None:
            ws.cell(r, col, default)
    return col


wb = load_workbook(WB)

# ── 1. BenefitsSources: cite the CAOs used ─────────────────────────────────
src_ws = wb["BenefitsSources"]
sh = headers(src_ws)
existing_sources = {src_ws.cell(r, sh["Source"]).value for r in range(2, src_ws.max_row + 1)}

NEW_SOURCES = [
    ("Shell Nederland Raffinaderij CAO 2022-2025", "FNV / CNV / De Unie", "Company CAO (user-supplied compilation)",
     "", TODAY, "Shell NL Raffinaderij pension (SSPF) and disability (AOP) figures"),
    ("Zeeland Refinery CAO 2024-2025", "StAP Pensioen / CAO-partijen", "Company CAO (user-supplied compilation)",
     "", TODAY, "Zeeland Refinery pension (StAP) and parental-leave figures"),
    ("AkzoNobel CAO 2024", "AkzoNobel / trade unions", "Company CAO (user-supplied summary, unverified against primary text)",
     "", TODAY, "AkzoNobel parental-leave top-up figure — DRAFT, needs primary-source verification"),
    ("Ecolab Production Netherlands CAO 2023 + Pensioenreglement 2023", "Stichting Pensioenfonds Ecolab",
     "Company CAO (user-supplied summary, unverified against primary text)",
     "", TODAY, "Ecolab NL pension premium split — DRAFT, needs primary-source verification"),
]
for row in NEW_SOURCES:
    if row[0] not in existing_sources:
        src_ws.append(list(row))

# ── 2. BenefitsObservations: add CompanyName + Notes columns ───────────────
obs_ws = wb["BenefitsObservations"]
add_column(obs_ws, "CompanyName")
add_column(obs_ws, "Notes")
oh = headers(obs_ws)

existing_ids = [str(obs_ws.cell(r, oh["ObsID"]).value) for r in range(2, obs_ws.max_row + 1)
                if obs_ws.cell(r, oh["ObsID"]).value]
next_id = max((int(x.split("-")[1]) for x in existing_ids if x.startswith("BO-")), default=0) + 1


def append_observation(industry_id, category, value, unit, currency, source, company, notes, status="Active"):
    global next_id
    row = [None] * obs_ws.max_column
    row[oh["ObsID"] - 1] = f"BO-{next_id:05d}"
    row[oh["IndustryID"] - 1] = industry_id
    row[oh["Category"] - 1] = category
    row[oh["Value"] - 1] = value
    row[oh["Unit"] - 1] = unit
    row[oh["Currency"] - 1] = currency
    row[oh["Source"] - 1] = source
    row[oh["Owner"] - 1] = "Reward"
    row[oh["Status"] - 1] = status
    row[oh["EffectiveFrom"] - 1] = TODAY
    row[oh["UpdatedAt"] - 1] = TODAY
    row[oh["CompanyName"] - 1] = company
    row[oh["Notes"] - 1] = notes
    obs_ws.append(row)
    next_id += 1


REAL_OBSERVATIONS = [
    # -- Shell Nederland Raffinaderij (CAO 2022-2025) — Active: from detailed CAO table --
    dict(industry_id="IND-MFG", category="Pension", value=15.8, unit="%", currency="",
         source="Shell Nederland Raffinaderij CAO 2022-2025", company="Shell Nederland Raffinaderij",
         notes="CAO states employer pays >70% of total SSPF premium (~20-25% of pensionable base, middelloon). "
               "15.8% is a derived point estimate (70% x 22.5% midpoint) — the CAO does not state employer "
               "share directly as %-of-base, only as %-of-total-premium.", status="Active"),
    dict(industry_id="IND-MFG", category="Disability Top-up", value=70.0, unit="%", currency="",
         source="Shell Nederland Raffinaderij CAO 2022-2025", company="Shell Nederland Raffinaderij",
         notes="AOP via SSPF targets ~90% salary replacement up to the WIA wage cap, 70% above the cap. "
               "70% (the above-cap portion) is used here as the closest match to this category's 'insured "
               "above the WIA cap' basis. Employer-funded; premium-free for the employee during disability.",
         status="Active"),
    # -- Zeeland Refinery (CAO 2024-2025) — Active: from detailed CAO table --
    dict(industry_id="IND-MFG", category="Pension", value=17.5, unit="%", currency="",
         source="Zeeland Refinery CAO 2024-2025", company="Zeeland Refinery",
         notes="CAO/UVO: premium split employer ~17-18%, employee ~10-11% (indicatief), via StAP Pensioen "
               "(pensioenkring TotalEnergies Nederland). 17.5% = midpoint of employer range.", status="Active"),
    dict(industry_id="IND-MFG", category="Parental Leave Top-up", value=0.0, unit="%", currency="",
         source="Zeeland Refinery CAO 2024-2025", company="Zeeland Refinery",
         notes="CAO text: statutory (wettelijke) parental-leave arrangements apply; 'geen expliciete "
               "bedrijfsopvang/top-up in CAO' — explicitly no company-specific top-up beyond statutory. "
               "Recorded as a verified 0, not an absence of data.", status="Active"),
    # -- AkzoNobel (CAO ~2024) — Draft: summary-only, not primary CAO text --
    dict(industry_id="IND-MFG", category="Parental Leave Top-up", value=9.0, unit="%", currency="",
         source="AkzoNobel CAO 2024", company="AkzoNobel",
         notes="Reported as 'supplementary parental leave up to 9% salary for maximum 100 weeks' in a "
               "secondary summary (not the primary CAO text). The 100-week duration is unusually long for "
               "NL parental leave norms — VERIFY against the actual AkzoNobel CAO before treating as "
               "confirmed. Also reported: 80/90/100 control rules tied to disability coverage (unclear "
               "mechanism from this source — not imported as a figure).", status="Draft"),
    # -- Ecolab Production Netherlands (CAO 2023) — Draft: summary-only --
    dict(industry_id="IND-MFG", category="Pension", value=39.4, unit="%", currency="",
         source="Ecolab Production Netherlands CAO 2023 + Pensioenreglement 2023", company="Ecolab Production Netherlands",
         notes="Reported premium split 39.4% employer / 5.6% employee, with a 1.875%/year middelloon accrual "
               "rate, in a secondary summary (not the detailed CAO table available for Shell/Zeeland). "
               "39.4% is unusually high vs. the other IND-MFG data points here — VERIFY against the primary "
               "CAO/pensioenreglement before treating as confirmed.", status="Draft"),
]

for obs in REAL_OBSERVATIONS:
    append_observation(obs["industry_id"], obs["category"], obs["value"], obs["unit"], obs["currency"],
                        obs["source"], obs["company"], obs["notes"], obs["status"])

wb.save(WB)
print(f"Added {len(NEW_SOURCES)} BenefitsSources rows (skipped duplicates).")
print(f"Added {len(REAL_OBSERVATIONS)} real BenefitsObservations rows.")
print("Saved", WB)
