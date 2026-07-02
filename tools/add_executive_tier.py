"""
Add a C-suite executive tier to jobsy_reference_library.xlsx and remap the
C-suite synonyms onto it. In-place openpyxl edit; preserves everything else.
"""
from openpyxl import load_workbook

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-02"
SRC = "Seed v1 (exec tier)"
wb = load_workbook(WB)

def hdr(ws): return {c.value: i+1 for i, c in enumerate(ws[1]) if c.value is not None}

def append_row(ws, d):
    h = hdr(ws); r = ws.max_row + 1
    for k, v in d.items():
        if k in h:
            ws.cell(r, h[k], v)
    # governance defaults for any gov columns not explicitly set
    for gk, gv in {"Status": "Active", "EffectiveFrom": TODAY, "UpdatedAt": TODAY}.items():
        if gk in h and (d.get(gk) is None):
            ws.cell(r, h[gk], gv)

def gov(owner):
    return {"Source": SRC, "Owner": owner, "Status": "Active",
            "EffectiveFrom": TODAY, "UpdatedAt": TODAY}

# ── new executive roles: id -> (title, esco label, skills-source Head role) ──
EXECS = [
 ("J-EXEC-02", "Chief Technology Officer", "chief technology officer", "J-ENG-04"),
 ("J-EXEC-03", "Chief Operating Officer",  "chief operating officer",  "J-OPS-05"),
 ("J-EXEC-04", "Chief Financial Officer",  "chief financial officer",  "J-FIN-06"),
 ("J-EXEC-05", "Chief Human Resources Officer", "chief human resources officer", "J-HR-05"),
 ("J-EXEC-06", "Chief Marketing Officer",  "chief marketing officer",  "J-MKT-05"),
 ("J-EXEC-07", "Chief Commercial Officer",  "chief commercial officer", "J-SAL-05"),
 ("J-EXEC-08", "Chief Product Officer",    "chief product officer",    "J-PRD-04"),
 ("J-EXEC-09", "Chief Data Officer",       "chief data officer",       "J-DAT-05"),
 ("J-EXEC-10", "Chief Information Officer", "chief information officer", "J-ENG-04"),
 ("J-EXEC-11", "Chief Legal Officer",      "chief legal officer",      "J-LEG-03"),
]

PROFILES = {
 "J-EXEC-02": ("Owns the company's technology vision, architecture and engineering organisation; sets technical strategy and scales the platform and teams.",
   "Set technology strategy and architecture; Build and lead the engineering organisation; Own platform scalability, security and reliability; Advise the CEO and board on technology investment; Drive technical hiring and engineering culture",
   "Technology strategy; Engineering leadership; System architecture; Scaling teams; Security and reliability",
   "SaaS platforms; Scale-up engineering; Cloud architecture", "GitHub; AWS/Azure/GCP; Jira; Datadog; Power BI"),
 "J-EXEC-03": ("Owns company-wide operational strategy and execution; partners with the CEO to scale the business efficiently and reliably across all functions.",
   "Set operational strategy and company-wide execution; Drive efficiency and scalability; Own cross-functional delivery and performance; Partner with the CEO on business continuity; Build and lead the operations leadership team",
   "Operational strategy; Company-wide execution; Scaling operations; Cross-functional leadership; Performance management",
   "Company operations; Scale-up leadership; Process excellence", "SAP or Oracle; Salesforce; Power BI; Jira; Workday"),
 "J-EXEC-04": ("Owns financial strategy, controlling, treasury and investor reporting; accountable for the financial health and integrity of the company.",
   "Set financial strategy and capital allocation; Own controlling, treasury and reporting; Lead budgeting and forecasting; Manage investor and board reporting; Ensure financial controls and compliance",
   "Financial strategy; Controlling; Treasury; Investor relations; Financial controls",
   "PE/VC-backed finance; Scale-up finance; M&A", "SAP or Oracle; Excel; Power BI; NetSuite; Board portal"),
 "J-EXEC-05": ("Owns people strategy, organisation design, culture and reward; accountable for building the organisation and its leadership capability.",
   "Set people and talent strategy; Own organisation design and culture; Lead reward, performance and succession; Advise the CEO and board on people risk; Build the HR leadership team",
   "People strategy; Organisation design; Culture; Reward; Talent and succession",
   "Scale-up people leadership; Change management; Total reward", "Workday; AFAS; Power BI; Culture Amp; Board portal"),
 "J-EXEC-06": ("Owns marketing, brand and growth strategy; accountable for demand, positioning and the marketing organisation.",
   "Set marketing and brand strategy; Own demand generation and growth; Lead positioning and messaging; Manage marketing budget and ROI; Build the marketing leadership team",
   "Marketing strategy; Brand; Growth; Demand generation; Positioning",
   "B2B growth; Brand building; Digital marketing", "HubSpot; Google Analytics; Power BI; Salesforce; LinkedIn"),
 "J-EXEC-07": ("Owns commercial and revenue strategy across sales and go-to-market; accountable for revenue growth and the commercial organisation.",
   "Set commercial and revenue strategy; Own sales and go-to-market; Lead pricing and pipeline; Drive revenue growth and forecasting; Build the commercial leadership team",
   "Commercial strategy; Revenue growth; Go-to-market; Pricing; Sales leadership",
   "B2B sales leadership; Revenue operations; Scale-up commercial", "Salesforce; HubSpot; Power BI; Gong; LinkedIn"),
 "J-EXEC-08": ("Owns product vision, strategy and roadmap; accountable for product-market fit and the product organisation.",
   "Set product vision and strategy; Own the roadmap and discovery; Align product with company strategy; Lead prioritisation and outcomes; Build the product leadership team",
   "Product strategy; Roadmap; Discovery; Prioritisation; Product leadership",
   "SaaS product; Platform product; Product-led growth", "Jira; Productboard; Amplitude; Figma; Power BI"),
 "J-EXEC-09": ("Owns data strategy, analytics and governance; accountable for turning data into value and for the data organisation.",
   "Set data and analytics strategy; Own data governance and quality; Lead the data platform and analytics; Advise leadership on data-driven decisions; Build the data leadership team",
   "Data strategy; Analytics; Data governance; Data platform; Data leadership",
   "Analytics leadership; Data platform; Data governance", "Snowflake; dbt; Power BI; Python; Looker"),
 "J-EXEC-10": ("Owns IT strategy, enterprise systems and information security; accountable for the technology backbone and information risk of the company.",
   "Set IT and systems strategy; Own enterprise systems and infrastructure; Lead information security and risk; Manage IT budget and vendors; Build the IT leadership team",
   "IT strategy; Enterprise systems; Infrastructure; Information security; Vendor management",
   "Enterprise IT; Information security; Systems integration", "ServiceNow; Microsoft 365; Azure; Okta; Power BI"),
 "J-EXEC-11": ("Owns legal strategy, governance, compliance and risk; accountable for protecting the company legally and for the legal organisation.",
   "Set legal and governance strategy; Own compliance and risk; Lead contracts and corporate legal; Advise the CEO and board on legal risk; Build the legal team",
   "Legal strategy; Governance; Compliance; Risk; Corporate legal",
   "Corporate law; Compliance; Data protection", "iManage; DocuSign; Board portal; OneTrust; Excel"),
}

jobs_ws, jp_ws, rsm_ws, cp_ws = wb["Jobs"], wb["JobProfiles"], wb["RoleSkillMap"], wb["CareerPaths"]

# existing skills per source Head role
rsm_h = hdr(rsm_ws)
skills_by_job = {}
for r in range(2, rsm_ws.max_row + 1):
    jid = rsm_ws.cell(r, rsm_h["JobID"]).value
    if jid is None:
        continue
    skills_by_job.setdefault(str(jid).strip(), []).append(
        (rsm_ws.cell(r, rsm_h["SkillID"]).value,
         rsm_ws.cell(r, rsm_h["RequiredLevel"]).value,
         rsm_ws.cell(r, rsm_h["SkillType"]).value))

for jid, title, esco, src_role in EXECS:
    # Jobs
    append_row(jobs_ws, {"JobID": jid, "StandardTitle": title, "Function": "Executive",
        "Level": "Lead", "Category": "Leadership & Management", "Grade": 13,
        "IscoGroup": "1120", "IscoTitle": "Managing directors and chief executives",
        "EscoLabel": esco, **gov("Job Architecture")})
    # JobProfiles
    desc, resp, req, spec, tools = PROFILES[jid]
    append_row(jp_ws, {"JobID": jid, "Description": desc, "KeyResponsibilities": resp,
        "RequiredSkills": req, "Specialisms": spec,
        "ManagementLevel": f"Executive — C-suite leader of the {title.split('Chief ')[-1].split(' Officer')[0].lower()} function",
        "TypicalTools": tools, **gov("Job Architecture")})
    # RoleSkillMap — inherit the domain Head's curated skills
    for sid, lvl, stype in skills_by_job.get(src_role, []):
        append_row(rsm_ws, {"JobID": jid, "SkillID": sid, "RequiredLevel": lvl,
            "SkillType": stype, **gov("Talent & Capability")})
    # CareerPaths — CxO -> CEO
    append_row(cp_ws, {"JobID": jid, "NextJobID": "J-EXEC-01",
        "NextRole": "Chief Executive Officer", **gov("Job Architecture")})

# ── TitleMapping: remap C-suite synonyms onto the new exec roles ─────────────
REMAP = {
 "CHRO": "J-EXEC-05", "Chief Human Resources Officer": "J-EXEC-05", "Chief People Officer": "J-EXEC-05",
 "CFO": "J-EXEC-04", "Chief Financial Officer": "J-EXEC-04",
 "CTO": "J-EXEC-02", "Chief Technology Officer": "J-EXEC-02",
 "Chief Information Officer": "J-EXEC-10",
 "COO": "J-EXEC-03", "Chief Operating Officer": "J-EXEC-03",
 "CMO": "J-EXEC-06", "Chief Marketing Officer": "J-EXEC-06",
 "CCO": "J-EXEC-07", "Chief Commercial Officer": "J-EXEC-07", "Chief Revenue Officer": "J-EXEC-07",
 "Chief Product Officer": "J-EXEC-08", "Chief Data Officer": "J-EXEC-09", "Chief Legal Officer": "J-EXEC-11",
}
tm_ws = wb["TitleMapping"]; th = hdr(tm_ws)
remapped = 0
existing_titles = set()
for r in range(2, tm_ws.max_row + 1):
    et = tm_ws.cell(r, th["ExistingTitle"]).value
    if et is None:
        continue
    ets = str(et).strip(); existing_titles.add(ets)
    if ets in REMAP:
        tm_ws.cell(r, th["JobID"], REMAP[ets])
        tm_ws.cell(r, th["UpdatedAt"], TODAY)
        remapped += 1

# add CxO abbreviations that weren't already present
NEW_SYNS = {"CIO": "J-EXEC-10", "CDO": "J-EXEC-09", "CPO": "J-EXEC-08", "CRO": "J-EXEC-07",
            "CLO": "J-EXEC-11", "Chief Information Officer (CIO)": "J-EXEC-10"}
added = 0
for et, jid in NEW_SYNS.items():
    if et not in existing_titles:
        append_row(tm_ws, {"ExistingTitle": et, "JobID": jid, **gov("Data Quality")})
        added += 1

wb.save(WB)
print(f"Exec roles added: {len(EXECS)}")
print(f"TitleMapping C-suite synonyms remapped: {remapped}")
print(f"New CxO abbreviations added: {added}")
