"""
Expand the benefits taxonomy from 9 to 19 categories, in place.

New categories reflect real signal from user-supplied CAO/company-benefits
material that didn't fit the original 9: flexible/cafeteria benefits,
older-worker phased-retirement schemes, discounts, bicycle plans, company
cars, life/risk insurance, short-term (year-1) sick pay, childcare, EAP,
and representation allowances. Explicitly excluded: shift allowance,
overtime %, and other pay-side elements already tracked in PayMix/
SalaryBands — these are compensation, not benefits.

Each new category gets:
  1. A BenefitsCatalog definition row (same governance columns as before).
  2. A synthetic self-built baseline distribution in BenefitsObservations
     (same method as tools/add_benefits_benchmarking.py: industry-richness
     multiplier + gaussian jitter, 14 obs per industry), so percentiles are
     computable immediately.
  3. LevelBenefitsFactors rows (4 levels).
  4. Real, cited company observations where genuinely documented (kept to
     what's actually stated in the source material — not fabricated).
"""
import random

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"
HDR_FILL = PatternFill("solid", fgColor="0E7C66")
HDR_FONT = Font(bold=True, color="FFFFFF")

random.seed(20260707)


def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


wb = load_workbook(WB)

# ── 1. BenefitsCatalog: 10 new category definitions ────────────────────────
NEW_CATEGORIES = [
    ("Flexible Benefits / Cafetariaregeling",
     "Gross salary/hours exchangeable for tax-free benefits under the WKR", "% of base",
     "2-10% of base exchangeable", "No", "No (WKR vrije ruimte)",
     "Cafeteria-style scheme letting employees exchange gross salary or hours for "
     "tax-advantaged benefits (bicycle, extra leave, gym) under the werkkostenregeling (WKR)."),
    ("Generation Pact / Vitality Scheme",
     "% of hours retained at (near-)full pay for older employees nearing retirement", "% hours",
     "Often structured as 80% hours / 90% pay / 100% pension accrual", "No (CAO-based)", "N/A",
     "Phased working-time reduction for employees nearing retirement (generatiepact/"
     "vitaliteitsregeling), retaining most pay and full pension accrual."),
    ("Employee Discounts & Perks",
     "Average discount value on gym/sport, retail, and union-reimbursed memberships", "%",
     "5-20% average discount value", "No", "No (small perks, WKR)",
     "Gym/sport membership discounts, retail/brand discounts, and union (FNV/CNV) "
     "membership fee reimbursements."),
    ("Bicycle Plan",
     "Per-km commuting allowance or lease-bike scheme", "EUR/km",
     "EUR 0.15-0.35/km, or lease-bike via WKR", "No", "No (up to fiscal-free km rate)",
     "Bicycle commuting allowance (fietsvergoeding) or lease-bike scheme, often paired "
     "with the general commute allowance."),
    ("Company Car / Lease",
     "Fixed monthly lease budget or company-car eligibility, by grade/scale", "EUR/month",
     "EUR 300-1600/month depending on grade/scale", "No",
     "Yes (bijtelling — added to taxable income based on catalogue value)",
     "Company car or car-lease budget, typically graded by job level/scale; private "
     "use is taxed via bijtelling."),
    ("Life & Risk Insurance",
     "Group life/risk insurance: lump sum insured or salary multiple", "EUR",
     "EUR 50k-300k lump sum, or 1-3x annual salary", "No", "No (death benefit to beneficiary)",
     "Group life insurance / risk insurance (ANW-hiaat or overlijdensrisicoverzekering) "
     "paying a lump sum or salary multiple on death, often including business-travel "
     "accident cover (BTA)."),
    ("Short-Term Disability (Sick Pay Top-up)",
     "% of base paid during year 1 of illness, above the 70% statutory minimum", "%",
     "70% statutory minimum, commonly topped up to 100% in year 1", "Yes (70% minimum, Wet Poortwachter)",
     "Yes",
     "Employer-paid sick leave during the first year of illness, on top of the 70% "
     "statutory minimum (many CAOs top up to 100%)."),
    ("Childcare Support",
     "Employer contribution toward childcare costs, beyond the statutory kinderopvangtoeslag", "EUR/month",
     "EUR 0-200/month; most employers offer none beyond statutory tegemoetkoming",
     "Partly (Wet Kinderopvang: government + parent; employer contribution voluntary)",
     "No (if within WKR)",
     "Employer top-up toward childcare costs beyond the statutory/government childcare "
     "allowance system."),
    ("EAP / Psychosocial Support",
     "Free counselling/coaching sessions per year via the EAP provider", "sessions/year",
     "0-10 free sessions/year via arbodienst or external EAP provider", "No", "No",
     "Confidential counselling/coaching for mental health and psychosocial support, via "
     "the occupational health service (arbodienst) or an external EAP provider."),
    ("Representation Allowance",
     "Fixed monthly allowance for representative/client-facing roles", "EUR/month",
     "EUR 50-400/month, role-dependent", "No", "Yes",
     "Allowance for representation/entertainment costs in client-facing or senior roles."),
]

cat_ws = wb["BenefitsCatalog"]
ch = headers(cat_ws)
existing_ids = [str(cat_ws.cell(r, ch["BenefitID"]).value) for r in range(2, cat_ws.max_row + 1)
                if cat_ws.cell(r, ch["BenefitID"]).value]
next_bid = max((int(x.split("-")[1]) for x in existing_ids if x.startswith("BEN-")), default=0) + 1
existing_cats = {cat_ws.cell(r, ch["Category"]).value for r in range(2, cat_ws.max_row + 1)}

for cat, basis, unit, typ, stat_nl, taxable, desc in NEW_CATEGORIES:
    if cat in existing_cats:
        continue
    cat_ws.append([f"BEN-{next_bid:02d}", cat, basis, unit, typ, stat_nl, taxable, desc,
                   "Seed v2 (benefits, taxonomy expansion)", "Reward", "Active", TODAY, TODAY])
    next_bid += 1

# ── 2. BenefitsObservations: synthetic baseline per new category ───────────
INDUSTRIES = ["IND-TECH", "IND-FIN", "IND-HLTH", "IND-MFG", "IND-RET", "IND-PUB", "IND-PSV", "IND-LOG"]
INDUSTRY_RICHNESS = {
    "IND-TECH": 1.25, "IND-FIN": 1.30, "IND-HLTH": 0.95, "IND-MFG": 0.95,
    "IND-RET": 0.75, "IND-PUB": 0.90, "IND-PSV": 1.15, "IND-LOG": 0.80,
}
NEW_CATEGORY_BASELINE = {
    # category -> (mean, relative jitter, unit)
    "Flexible Benefits / Cafetariaregeling": (5.0, 0.40, "%"),
    "Generation Pact / Vitality Scheme": (80.0, 0.08, "%"),
    "Employee Discounts & Perks": (12.0, 0.35, "%"),
    "Bicycle Plan": (0.23, 0.25, "EUR"),
    "Company Car / Lease": (700.0, 0.40, "EUR"),
    "Life & Risk Insurance": (150000.0, 0.45, "EUR"),
    "Short-Term Disability (Sick Pay Top-up)": (95.0, 0.08, "%"),
    "Childcare Support": (50.0, 0.60, "EUR"),
    "EAP / Psychosocial Support": (5.0, 0.50, "sessions"),
    "Representation Allowance": (150.0, 0.40, "EUR"),
}
N_OBS_PER_GROUP = 14

obs_ws = wb["BenefitsObservations"]
oh = headers(obs_ws)
existing_obs_ids = [str(obs_ws.cell(r, oh["ObsID"]).value) for r in range(2, obs_ws.max_row + 1)
                    if obs_ws.cell(r, oh["ObsID"]).value]
next_obs_id = max((int(x.split("-")[1]) for x in existing_obs_ids if x.startswith("BO-")), default=0) + 1

n_cols = obs_ws.max_column


def append_obs_row(industry_id, category, value, unit, currency, source, company, notes, status="Active"):
    global next_obs_id
    row = [None] * n_cols
    row[oh["ObsID"] - 1] = f"BO-{next_obs_id:05d}"
    row[oh["IndustryID"] - 1] = industry_id
    row[oh["Category"] - 1] = category
    row[oh["Value"] - 1] = value
    row[oh["Unit"] - 1] = unit
    row[oh["Currency"] - 1] = currency
    row[oh["Source"] - 1] = source
    row[oh["Owner"] - 1] = "Reward"
    row[oh["Status"] - 1] = status
    row[oh["EffectiveFrom"] - 1] = TODAY
    row[oh["UpdatedAt"] - 1] = TODAY
    if "CompanyName" in oh:
        row[oh["CompanyName"] - 1] = company
    if "Notes" in oh:
        row[oh["Notes"] - 1] = notes
    obs_ws.append(row)
    next_obs_id += 1


n_synthetic = 0
for industry in INDUSTRIES:
    richness = INDUSTRY_RICHNESS[industry]
    for category, (mean, jitter, unit) in NEW_CATEGORY_BASELINE.items():
        group_mean = mean * richness
        for _ in range(N_OBS_PER_GROUP):
            noise = random.gauss(1.0, jitter)
            value = max(0.0, group_mean * noise)
            currency = "EUR" if unit == "EUR" else ""
            append_obs_row(industry, category, round(value, 2), unit, currency,
                           "Seed v2 (benefits, taxonomy expansion)", "", "", "Active")
            n_synthetic += 1

# ── 3. Real, cited company observations for the new categories ────────────
REAL_NEW_OBSERVATIONS = [
    dict(industry_id="IND-MFG", category="Generation Pact / Vitality Scheme", value=80.0, unit="%",
         source="Unilever CAO (user-supplied benefits comparison, Dec 2025)", company="Unilever Netherlands",
         notes="CAO: 80/90/100 generation-pact scheme (80% hours / 90% pay / 100% pension accrual), "
               "available up to 8 years before pension age; also covers informal caregiving (mantelzorg).",
         status="Draft"),
    dict(industry_id="IND-MFG", category="Generation Pact / Vitality Scheme", value=80.0, unit="%",
         source="AkzoNobel CAO 2024 (user-supplied summary, unverified against primary text)", company="AkzoNobel",
         notes="Reported as '80/90/100 control rules' in a secondary summary tied to disability/older-worker "
               "provisions — mechanism not fully clear from this source. VERIFY against primary AkzoNobel CAO.",
         status="Draft"),
    dict(industry_id="IND-MFG", category="Short-Term Disability (Sick Pay Top-up)", value=100.0, unit="%",
         source="Shell Nederland Raffinaderij CAO 2022-2025", company="Shell Nederland Raffinaderij",
         notes="CAO: '100% loondoorbetaling tot 2 jaar' — 100% pay continuation for the full 2 years "
               "(statutory minimum is 70%). Pension accrual continues (premievrijstelling) during illness.",
         status="Active"),
    dict(industry_id="IND-MFG", category="Short-Term Disability (Sick Pay Top-up)", value=100.0, unit="%",
         source="Zeeland Refinery CAO 2024-2025", company="Zeeland Refinery",
         notes="CAO: '100% loondoorbetaling volgens wettelijke regelgeving' — wording is ambiguous "
               "(statutory floor is 70%, not 100%); likely means 100% in year 1 per common CAO practice. "
               "VERIFY exact year-1/year-2 split against primary CAO text.",
         status="Draft"),
    dict(industry_id="IND-MFG", category="Childcare Support", value=0.0, unit="EUR",
         source="Zeeland Refinery CAO 2024-2025", company="Zeeland Refinery",
         notes="CAO: 'Wettelijke regelingen van toepassing... geen expliciete bedrijfsopvang in CAO' — "
               "explicitly no company-specific childcare support beyond the statutory system. Recorded "
               "as a verified 0, not an absence of data.", status="Active"),
    dict(industry_id="IND-MFG", category="Employee Discounts & Perks", value=15.0, unit="%",
         source="NL chemicals CAO perks/discounts comparison (user-supplied, Dec 2025)",
         company="Shell Nederland Raffinaderij",
         notes="Reported: 10-20% gym/fitness discount, CNV/FNV membership reimbursement EUR 50-100/year, "
               "15% travel/entertainment discount. 15% used as the midpoint of the gym-discount range.",
         status="Draft"),
    dict(industry_id="IND-MFG", category="Employee Discounts & Perks", value=12.5, unit="%",
         source="NL chemicals CAO perks/discounts comparison (user-supplied, Dec 2025)", company="Zeeland Refinery",
         notes="Reported: 10-15% Basic-Fit/Plutosport gym discount, FNV/CNV reimbursement EUR 50-100/year, "
               "~10% energy/brand discounts. 12.5% used as the midpoint of the gym-discount range.",
         status="Draft"),
]

for obs in REAL_NEW_OBSERVATIONS:
    append_obs_row(obs["industry_id"], obs["category"], obs["value"], obs["unit"], "",
                   obs["source"], obs["company"], obs["notes"], obs["status"])

# ── 4. LevelBenefitsFactors: 4 levels x 10 new categories ──────────────────
LEVEL_BASE = {"Junior": 0.85, "Medior": 1.00, "Senior": 1.15, "Lead": 1.35}
FLAT_CATEGORIES = {"Bicycle Plan", "EAP / Psychosocial Support", "Childcare Support",
                    "Short-Term Disability (Sick Pay Top-up)", "Generation Pact / Vitality Scheme"}
LEVELS = ["Junior", "Medior", "Senior", "Lead"]

lf_ws = wb["LevelBenefitsFactors"]
lh = headers(lf_ws)
existing_pairs = {(lf_ws.cell(r, lh["Level"]).value, lf_ws.cell(r, lh["Category"]).value)
                   for r in range(2, lf_ws.max_row + 1)}

n_factors = 0
for category in NEW_CATEGORY_BASELINE:
    for level in LEVELS:
        if (level, category) in existing_pairs:
            continue
        factor = 1.0 if category in FLAT_CATEGORIES else LEVEL_BASE[level]
        lf_ws.append([level, category, factor, "Seed v2 (benefits, taxonomy expansion)",
                     "Reward", "Active", TODAY, TODAY])
        n_factors += 1

wb.save(WB)
print(f"BenefitsCatalog: +{len(NEW_CATEGORIES)} categories")
print(f"BenefitsObservations: +{n_synthetic} synthetic, +{len(REAL_NEW_OBSERVATIONS)} real")
print(f"LevelBenefitsFactors: +{n_factors} rows")
print("Saved", WB)
