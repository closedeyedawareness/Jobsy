"""
Correct the CareerPaths ladder in jobsy_reference_library.xlsx, in place.

The problem
-----------
21 of 80 career steps pointed at Chief Executive Officer. That is not a ladder,
it is a fallback. Every head-of-function and every C-level role terminated in
"become the CEO", which is both wrong for almost all of them and actively
misleading in a tool people use to plan a career.

Why it happened (and why better scoring would not fix it)
---------------------------------------------------------
CEO is defined in RoleSkillMap by eight skills, every one of them generic:
stakeholder management (required by 43 of 81 roles), team leadership (34),
budget and resource management (32), strategic planning (27), board advisory
(24), data-driven decision making (35), change management, organisational
design. It has no distinctive skill at all. In skill space CEO is therefore the
generic-leadership centroid, and any similarity measure will rank it highly for
every senior role. The original ladder looks like exactly that mistake. Fixing
it needs structural rules, not a better distance function.

The rules applied
-----------------
R1  A head-of-function progresses to the C-level of the same function.
    Independently corroborated: running skill-adjacency over the library's own
    RoleSkillMap returns precisely this as the top-ranked candidate for Finance,
    Engineering, Data, Marketing and Legal, with no knowledge of the rule.

R2  COO -> CEO is retained. The operating role is the recognised internal
    springboard to chief executive.

R3  CFO -> CEO is retained as a real but minority route: roughly 7.5% of sitting
    CEOs were promoted from the CFO seat (Crist Kolder Associates), up from 6.5%
    in 2015.

R4  Every other C-level role is recorded as TERMINAL - no onward internal step.
    There is no comparable evidence base for CHRO/CLO/CMO/CIO/CPO/CDO/CTO -> CEO,
    and inventing one is what produced the original problem. The row is kept with
    a blank NextJobID and Status=Terminal so the governance record shows this was
    decided, not overlooked.

Sources
-------
  Spencer Stuart / Crist Kolder route-to-CEO analyses (COO and CFO progression)
  ESCO - European Skills, Competences, Qualifications and Occupations
  O*NET Career Changers Matrix (occupational transition pathways)
  Jobsy's own RoleSkillMap, via skill-adjacency scoring (corroboration for R1)

Usage
    python tools/fix_career_ladder.py --dry-run     # show the diff, change nothing
    python tools/fix_career_ladder.py               # apply in place
"""
from __future__ import annotations

import argparse
import shutil
from datetime import date
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
WORKBOOK = REPO / "jobsy_reference_library.xlsx"
SHEET = "CareerPaths"
TODAY = date.today().isoformat()

FUNCTION_LADDER = (
    "Ladder v2 - function ladder (skill-adjacency corroborated)")
FUNCTION_LADDER_PLAIN = (
    "Ladder v2 - function ladder (organisational structure)")
SPRINGBOARD = (
    "Ladder v2 - COO is the recognised internal springboard to CEO "
    "(Spencer Stuart / Crist Kolder)")
CFO_ROUTE = (
    "Ladder v2 - minority but evidenced route: ~7.5% of sitting CEOs promoted "
    "from CFO (Crist Kolder)")
TERMINAL = (
    "Ladder v2 - terminal: no evidence base for an onward internal step; "
    "recorded rather than defaulting to CEO")

# (from_title, to_title or None for terminal, source)
CHANGES: list[tuple[str, str | None, str]] = [
    # R1 - head of function -> C-level of the same function.
    # The first five were independently returned as the top-ranked candidate by
    # skill-adjacency over RoleSkillMap.
    ("Head of Finance",           "Chief Financial Officer",       FUNCTION_LADDER),
    ("Engineering Lead",          "Chief Technology Officer",      FUNCTION_LADDER),
    ("Head of Data",              "Chief Data Officer",            FUNCTION_LADDER),
    ("Head of Marketing",         "Chief Marketing Officer",       FUNCTION_LADDER),
    ("General Counsel",           "Chief Legal Officer",           FUNCTION_LADDER),
    ("Head of HR",                "Chief Human Resources Officer", FUNCTION_LADDER_PLAIN),
    ("Head of Product",           "Chief Product Officer",         FUNCTION_LADDER_PLAIN),
    ("Head of Sales",             "Chief Commercial Officer",      FUNCTION_LADDER_PLAIN),
    ("Head of Customer Success",  "Chief Commercial Officer",      FUNCTION_LADDER_PLAIN),
    ("Head of Operations",        "Chief Operating Officer",       FUNCTION_LADDER_PLAIN),
    ("Head of Procurement",       "Chief Operating Officer",       FUNCTION_LADDER_PLAIN),

    # R2 / R3 - the two evidenced routes into the chief executive seat.
    ("Chief Operating Officer",   "Chief Executive Officer",       SPRINGBOARD),
    ("Chief Financial Officer",   "Chief Executive Officer",       CFO_ROUTE),

    # R4 - terminal. Previously all of these pointed at CEO.
    ("Chief Commercial Officer",      None, TERMINAL),
    ("Chief Data Officer",            None, TERMINAL),
    ("Chief Human Resources Officer", None, TERMINAL),
    ("Chief Information Officer",     None, TERMINAL),
    ("Chief Legal Officer",           None, TERMINAL),
    ("Chief Marketing Officer",       None, TERMINAL),
    ("Chief Product Officer",         None, TERMINAL),
    ("Chief Technology Officer",      None, TERMINAL),
]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="report the diff, write nothing")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"{SHEET} sheet not found in {WORKBOOK.name}")

    # title <-> job_id from the Jobs sheet (the authority for both)
    jobs = wb["Jobs"]
    jhead = [c.value for c in jobs[1]]
    ji, jt = jhead.index("JobID"), jhead.index("StandardTitle")
    id_for, title_for = {}, {}
    for row in jobs.iter_rows(min_row=2, values_only=True):
        if row[ji]:
            id_for[str(row[jt]).strip()] = str(row[ji]).strip()
            title_for[str(row[ji]).strip()] = str(row[jt]).strip()

    missing = [t for t, to, _ in CHANGES for t in (t, to) if t and t not in id_for]
    if missing:
        raise SystemExit(f"These titles are not in the Jobs sheet: {sorted(set(missing))}")

    ws = wb[SHEET]
    head = [c.value for c in ws[1]]
    col = {name: head.index(name) + 1 for name in head if name}

    by_job = {}
    for r in range(2, ws.max_row + 1):
        jid = ws.cell(r, col["JobID"]).value
        if jid:
            by_job[str(jid).strip()] = r

    applied, unchanged = [], []
    for from_title, to_title, source in CHANGES:
        jid = id_for[from_title]
        row = by_job.get(jid)
        if row is None:
            raise SystemExit(f"No CareerPaths row for {from_title} ({jid})")

        old_id = ws.cell(row, col["NextJobID"]).value
        old_title = title_for.get(str(old_id).strip(), old_id) if old_id else "(none)"
        new_id = id_for[to_title] if to_title else None

        if (str(old_id).strip() if old_id else None) == (new_id or None):
            unchanged.append((from_title, old_title))
            continue

        applied.append((from_title, old_title, to_title or "(terminal)"))
        if not args.dry_run:
            ws.cell(row, col["NextJobID"]).value = new_id
            ws.cell(row, col["NextRole"]).value = to_title
            if "Source" in col:
                ws.cell(row, col["Source"]).value = source
            if "Status" in col:
                ws.cell(row, col["Status"]).value = "Active" if to_title else "Terminal"
            if "UpdatedAt" in col:
                ws.cell(row, col["UpdatedAt"]).value = TODAY

    print(f"{'DRY RUN - ' if args.dry_run else ''}{len(applied)} change(s), "
          f"{len(unchanged)} already correct\n")
    for frm, old, new in applied:
        print(f"  {frm:32} {old:26} ->  {new}")
    if unchanged:
        print("\n  already correct:")
        for frm, old in unchanged:
            print(f"  {frm:32} {old}")

    if args.dry_run:
        print("\nnothing written")
        return

    backup = WORKBOOK.with_name(f"{WORKBOOK.stem}.pre-ladder-v2.xlsx")
    if not backup.exists():
        shutil.copy2(WORKBOOK, backup)
        print(f"\nbackup: {backup.name}")
    wb.save(WORKBOOK)
    print(f"saved:  {WORKBOOK.name}")


if __name__ == "__main__":
    main()
