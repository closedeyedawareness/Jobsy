"""
Enrich jobsy_reference_library.xlsx in place (openpyxl) across 4 tracks:
  1. Complete the gaps (career paths, JobGrades.Authority/Leadership/SpanOfControl, ManagementLevel)
  2. Governance layer (provenance columns + DataDictionary sheet)
  3. Deeper skill model (SkillProficiency rubric sheet)
  4. ESCO/ISCO mapping (Isco/Esco columns on Jobs)
Preserves all existing sheets, formatting, and data. Only adds/fills.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-02"
wb = load_workbook(WB)

HDR_FILL = PatternFill("solid", fgColor="0E7C66")
HDR_FONT = Font(bold=True, color="FFFFFF")

def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}

def col_by_name(ws, *names):
    h = headers(ws)
    for n in names:
        if n in h:
            return h[n]
    return None

def add_column(ws, name, value_fn):
    """Append a new column with header `name`; value_fn(row_idx)->value for data rows."""
    h = headers(ws)
    if name in h:
        col = h[name]
    else:
        col = ws.max_column + 1
        c = ws.cell(1, col, name); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    for r in range(2, ws.max_row + 1):
        ws.cell(r, col, value_fn(r))
    return col

# ---- job title lookup (JobID -> StandardTitle) ---------------------------
jobs_ws = wb["Jobs"]
jh = headers(jobs_ws)
JID, JTITLE = jh["JobID"], jh["StandardTitle"]
title_of = {}
for r in range(2, jobs_ws.max_row + 1):
    jid = jobs_ws.cell(r, JID).value
    if jid:
        title_of[str(jid).strip()] = str(jobs_ws.cell(r, JTITLE).value).strip()

# =========================================================================
# TRACK 1 — COMPLETE THE GAPS
# =========================================================================

# 1a. Career paths for the 21 dead-end roles (CEO stays terminal).
CAREER_ADDS = {
    "J-HR-04": "J-HR-05", "J-HR-05": "J-EXEC-01", "J-HR-07": "J-HR-04",
    "J-FIN-05": "J-FIN-06", "J-FIN-06": "J-EXEC-01",
    "J-ENG-04": "J-EXEC-01", "J-ENG-05": "J-ENG-03", "J-ENG-06": "J-ENG-03", "J-ENG-07": "J-ENG-03",
    "J-DAT-04": "J-DAT-05", "J-DAT-05": "J-EXEC-01",
    "J-PRD-04": "J-EXEC-01",
    "J-OPS-03": "J-OPS-05", "J-OPS-04": "J-OPS-05", "J-OPS-05": "J-EXEC-01",
    "J-SAL-04": "J-SAL-05", "J-SAL-05": "J-EXEC-01",
    "J-MKT-04": "J-MKT-03", "J-MKT-05": "J-EXEC-01",
    "J-CS-04": "J-EXEC-01",
    "J-LEG-03": "J-EXEC-01",
}
cp_ws = wb["CareerPaths"]
ch = headers(cp_ws)
existing_sources = {str(cp_ws.cell(r, ch["JobID"]).value).strip()
                    for r in range(2, cp_ws.max_row + 1) if cp_ws.cell(r, ch["JobID"]).value}
added_cp = 0
for jid, nxt in CAREER_ADDS.items():
    if jid in existing_sources:
        continue
    row = cp_ws.max_row + 1
    cp_ws.cell(row, ch["JobID"], jid)
    cp_ws.cell(row, ch["NextJobID"], nxt)
    cp_ws.cell(row, ch["NextRole"], title_of.get(nxt, ""))
    added_cp += 1

# 1b. JobGrades: fill Authority (all), and missing Leadership / SpanOfControl.
AUTHORITY = {
    1:  "No independent authority; acts within defined instructions and close supervision.",
    2:  "Limited authority; completes defined tasks and escalates exceptions to a supervisor.",
    3:  "Owns routine decisions within established procedures; escalates non-standard cases.",
    4:  "Full authority over own deliverables within professional standards and policy.",
    5:  "Decides approach for own work; recommends improvements to process and quality.",
    6:  "Authority over specialist-domain decisions; sets local standards; advises the team.",
    7:  "Approves work and standards within the team; allocates tasks; signs off routine spend within budget.",
    8:  "Approves within the domain; owns team budget and priorities; input to hiring decisions.",
    9:  "Cross-functional approval authority; owns functional budget lines; makes team hiring decisions.",
    10: "Sets and approves strategy within the function; owns significant budget; senior hiring authority.",
    11: "Full function-wide decision and budget authority; accountable to the board for function performance.",
    12: "Cross-functional decision authority; approves major investment within delegated limits; shapes company strategy.",
    13: "Executive decision rights with board recommendation; approves company-wide policy and major commitments.",
    14: "Full board authority; ultimate accountability for strategic, financial and cultural decisions.",
}
LEADERSHIP_FILL = {1: "None (individual contributor)", 2: "None (individual contributor)", 4: "Informal peer guidance"}
SPAN_FILL = {1: "None", 2: "None", 3: "None"}
jg_ws = wb["JobGrades"]
gh = headers(jg_ws)
G, AUTH, LEAD, SPAN = gh["Grade"], gh["Authority"], gh["Leadership"], gh["SpanOfControl"]
def _empty(v): return v is None or str(v).strip() == "" or str(v).strip().lower() == "nan"
for r in range(2, jg_ws.max_row + 1):
    try:
        g = int(float(jg_ws.cell(r, G).value))
    except (TypeError, ValueError):
        continue
    if g in AUTHORITY:
        jg_ws.cell(r, AUTH, AUTHORITY[g])
    if _empty(jg_ws.cell(r, LEAD).value) and g in LEADERSHIP_FILL:
        jg_ws.cell(r, LEAD, LEADERSHIP_FILL[g])
    if _empty(jg_ws.cell(r, SPAN).value) and g in SPAN_FILL:
        jg_ws.cell(r, SPAN, SPAN_FILL[g])

# 1c. JobProfiles.ManagementLevel: fill blanks (IC vs People Manager from the role title).
jp_ws = wb["JobProfiles"]
ph = headers(jp_ws)
PJID, PML = ph["JobID"], ph["ManagementLevel"]
MGR_WORDS = ("manager", "head of", "lead", "chief", "general counsel", "director")
filled_ml = 0
for r in range(2, jp_ws.max_row + 1):
    if _empty(jp_ws.cell(r, PML).value):
        jid = str(jp_ws.cell(r, PJID).value).strip()
        t = title_of.get(jid, "").lower()
        jp_ws.cell(r, PML, "People Manager" if any(w in t for w in MGR_WORDS) else "Individual Contributor")
        filled_ml += 1

# =========================================================================
# TRACK 4 — ESCO / ISCO MAPPING (add columns to Jobs)  [before governance]
# =========================================================================
ISCO = {  # JobID -> (ISCO-08 unit group code, official title, ESCO preferred label [indicative])
 "J-HR-01": ("4416","Personnel clerks","human resources assistant"),
 "J-HR-02": ("2423","Personnel and careers professionals","human resources officer"),
 "J-HR-03": ("2423","Personnel and careers professionals","human resources business partner"),
 "J-HR-04": ("1212","Human resource managers","human resources manager"),
 "J-HR-05": ("1212","Human resource managers","human resources manager"),
 "J-HR-06": ("2423","Personnel and careers professionals","recruitment consultant"),
 "J-HR-07": ("2423","Personnel and careers professionals","recruitment consultant"),
 "J-FIN-01":("4311","Accounting and bookkeeping clerks","accounting assistant"),
 "J-FIN-02":("2411","Accountants","accountant"),
 "J-FIN-03":("2413","Financial analysts","financial analyst"),
 "J-FIN-04":("1211","Finance managers","financial controller"),
 "J-FIN-05":("1211","Finance managers","financial manager"),
 "J-FIN-06":("1211","Finance managers","financial manager"),
 "J-ENG-01":("2512","Software developers","software developer"),
 "J-ENG-02":("2512","Software developers","software developer"),
 "J-ENG-03":("2512","Software developers","software developer"),
 "J-ENG-04":("1330","Information and communications technology service managers","ICT operations manager"),
 "J-ENG-05":("2512","Software developers","software developer"),
 "J-ENG-06":("2519","Software and applications developers and analysts not elsewhere classified","software tester"),
 "J-ENG-07":("2512","Software developers","software developer"),
 "J-DAT-01":("2529","Database and network professionals not elsewhere classified","data analyst"),
 "J-DAT-02":("2529","Database and network professionals not elsewhere classified","data analyst"),
 "J-DAT-03":("2521","Database designers and administrators","data engineer"),
 "J-DAT-04":("2529","Database and network professionals not elsewhere classified","data scientist"),
 "J-DAT-05":("1330","Information and communications technology service managers","data manager"),
 "J-PRD-01":("2431","Advertising and marketing professionals","product manager"),
 "J-PRD-02":("2431","Advertising and marketing professionals","product manager"),
 "J-PRD-03":("2431","Advertising and marketing professionals","product manager"),
 "J-PRD-04":("1221","Sales and marketing managers","product manager"),
 "J-OPS-01":("3341","Office supervisors","business operations coordinator"),
 "J-OPS-02":("1219","Business services and administration managers not elsewhere classified","project manager"),
 "J-OPS-03":("1219","Business services and administration managers not elsewhere classified","operations manager"),
 "J-OPS-04":("1219","Business services and administration managers not elsewhere classified","programme manager"),
 "J-OPS-05":("1219","Business services and administration managers not elsewhere classified","operations manager"),
 "J-SAL-01":("3322","Commercial sales representatives","sales representative"),
 "J-SAL-02":("3322","Commercial sales representatives","sales account manager"),
 "J-SAL-03":("3322","Commercial sales representatives","sales account manager"),
 "J-SAL-04":("1221","Sales and marketing managers","business development manager"),
 "J-SAL-05":("1221","Sales and marketing managers","sales manager"),
 "J-MKT-01":("2431","Advertising and marketing professionals","marketing assistant"),
 "J-MKT-02":("2431","Advertising and marketing professionals","digital marketing specialist"),
 "J-MKT-03":("1221","Sales and marketing managers","marketing manager"),
 "J-MKT-04":("2431","Advertising and marketing professionals","content manager"),
 "J-MKT-05":("1221","Sales and marketing managers","marketing manager"),
 "J-CS-01": ("4222","Contact centre information clerks","customer service representative"),
 "J-CS-02": ("3322","Commercial sales representatives","client relations manager"),
 "J-CS-03": ("3322","Commercial sales representatives","client relations manager"),
 "J-CS-04": ("1221","Sales and marketing managers","customer service manager"),
 "J-LEG-01":("2611","Lawyers","lawyer"),
 "J-LEG-02":("2619","Legal professionals not elsewhere classified","compliance officer"),
 "J-LEG-03":("2611","Lawyers","lawyer"),
 "J-EXEC-01":("1120","Managing directors and chief executives","chief executive officer"),
}
def jid_at(r): return str(jobs_ws.cell(r, JID).value).strip()
add_column(jobs_ws, "IscoGroup",  lambda r: ISCO.get(jid_at(r), ("","",""))[0])
add_column(jobs_ws, "IscoTitle",  lambda r: ISCO.get(jid_at(r), ("","",""))[1])
add_column(jobs_ws, "EscoLabel",  lambda r: ISCO.get(jid_at(r), ("","",""))[2])

# =========================================================================
# TRACK 3 — DEEPER SKILL MODEL (SkillProficiency rubric: Category x Level 1-5)
# =========================================================================
LEVEL_NAME = {1:"Awareness",2:"Developing",3:"Proficient",4:"Advanced",5:"Expert"}
RUBRIC = {
 "Technical": {
  1:"Recognises core concepts and tooling; follows step-by-step guidance to make small, supervised changes.",
  2:"Completes well-defined tasks with support; can extend existing code/config in familiar areas.",
  3:"Works independently on standard problems; writes maintainable solutions and reviews peers' work.",
  4:"Designs solutions for complex, ambiguous problems; sets patterns and coaches others.",
  5:"Sets technical direction and standards across teams; recognised authority who resolves the hardest problems.",
 },
 "Data & Analytics": {
  1:"Reads dashboards and basic queries; understands key metrics and where data lives.",
  2:"Builds simple queries and reports with guidance; validates results against known figures.",
  3:"Independently models data and produces reliable analysis for standard business questions.",
  4:"Designs analytical approaches for complex questions; ensures statistical rigour and coaches analysts.",
  5:"Defines the analytics strategy and data standards; trusted authority on measurement and insight.",
 },
 "Professional": {
  1:"Understands the fundamentals; applies them to routine tasks under supervision.",
  2:"Handles standard situations with occasional guidance; consistent on familiar work.",
  3:"Operates independently and reliably across most situations in the role.",
  4:"Handles non-standard, high-stakes situations; sets good practice and mentors others.",
  5:"Shapes how the discipline is practised organisation-wide; recognised expert.",
 },
 "Commercial": {
  1:"Understands the offering and basic commercial terms; supports deals with guidance.",
  2:"Manages straightforward opportunities and accounts with support; follows the sales process.",
  3:"Independently runs the full cycle for standard deals; forecasts and hits targets reliably.",
  4:"Wins complex, multi-stakeholder deals; coaches the team and improves commercial approach.",
  5:"Sets commercial strategy and pricing direction; opens markets and builds the commercial engine.",
 },
 "Leadership": {
  1:"Aware of what good leadership looks like; leads self and small tasks.",
  2:"Guides peers informally; takes ownership of small pieces of team work.",
  3:"Leads a small team or workstream; sets clear goals and gives feedback.",
  4:"Leads teams through change and ambiguity; develops managers and builds capability.",
  5:"Sets organisational direction and culture; builds the leadership bench.",
 },
 "Finance & Accounting": {
  1:"Understands core accounting concepts; supports routine processing under review.",
  2:"Completes standard reconciliations and entries with guidance; spots obvious errors.",
  3:"Independently owns financial processes and reporting for a defined area; ensures accuracy and controls.",
  4:"Handles complex accounting, analysis and controls; advises the business and coaches the team.",
  5:"Sets financial policy, controls and strategy; authoritative on the numbers to board level.",
 },
 "Marketing & Digital": {
  1:"Understands channels and basic campaign concepts; supports execution with guidance.",
  2:"Runs defined campaign tasks with support; interprets basic performance metrics.",
  3:"Independently plans and runs campaigns; optimises against clear objectives and budget.",
  4:"Designs multi-channel strategy; drives measurable growth and coaches marketers.",
  5:"Sets brand and growth strategy; recognised authority who shapes the market approach.",
 },
 "People & HR": {
  1:"Understands core HR processes and policy; handles administration under supervision.",
  2:"Manages standard employee-lifecycle tasks with guidance; applies policy consistently.",
  3:"Independently advises on standard people matters; owns processes end-to-end.",
  4:"Handles complex, sensitive cases and change; coaches managers and shapes practice.",
  5:"Sets people strategy, policy and culture direction; trusted advisor to leadership.",
 },
}
if "SkillProficiency" in wb.sheetnames:
    del wb["SkillProficiency"]
sp = wb.create_sheet("SkillProficiency")
sp.append(["Category", "Level", "LevelName", "Anchor"])
for cell in sp[1]:
    cell.font = HDR_FONT; cell.fill = HDR_FILL
skills_ws = wb["Skills"]
skcats = [str(skills_ws.cell(r, headers(skills_ws)["Category"]).value).strip()
          for r in range(2, skills_ws.max_row + 1)]
for cat in sorted(set(skcats)):
    anchors = RUBRIC.get(cat, {lvl: LEVEL_NAME[lvl] for lvl in range(1, 6)})
    for lvl in range(1, 6):
        sp.append([cat, lvl, LEVEL_NAME[lvl], anchors.get(lvl, "")])

# =========================================================================
# TRACK 2b — GOVERNANCE COLUMNS on entity sheets
# =========================================================================
OWNER = {
 "Jobs":"Job Architecture","JobProfiles":"Job Architecture","CareerPaths":"Job Architecture",
 "Categories":"Job Architecture","Industries":"Job Architecture","SeniorityLevels":"Job Architecture",
 "SalaryBands":"Reward","JobGrades":"Reward","IndustrySalaryFactors":"Reward",
 "CareerBands":"Reward","LevelCriteria":"Reward",
 "Skills":"Talent & Capability","RoleSkillMap":"Talent & Capability","IndustrySkills":"Talent & Capability",
 "CompetencyLevels":"Talent & Capability","SkillProficiency":"Talent & Capability",
 "TitleMapping":"Data Quality","Levels":"Job Architecture",
}
GOV_SHEETS = list(OWNER.keys())
for name in GOV_SHEETS:
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    owner = OWNER[name]
    add_column(ws, "Source", lambda r: "Seed v1")
    add_column(ws, "Owner", lambda r, o=owner: o)
    add_column(ws, "Status", lambda r: "Active")
    add_column(ws, "EffectiveFrom", lambda r: TODAY)
    add_column(ws, "UpdatedAt", lambda r: TODAY)

# =========================================================================
# TRACK 2a — DATA DICTIONARY (built by scanning final sheet headers)
# =========================================================================
DESC = {
 # governance (shared)
 "Source":"Origin of the record (e.g. Seed v1, Workday, manual entry).",
 "Owner":"Team accountable for the accuracy of this record.",
 "Status":"Lifecycle state: Active, Draft, or Retired.",
 "EffectiveFrom":"Date from which this record is valid (YYYY-MM-DD).",
 "UpdatedAt":"Date this record was last changed (YYYY-MM-DD).",
 # Jobs
 "JobID":"Unique canonical role identifier (primary key), e.g. J-HR-03.",
 "StandardTitle":"Canonical, standardised job title.",
 "Function":"Business function the role belongs to (HR, Finance, Engineering, ...).",
 "Level":"Seniority level: Junior, Medior, Senior, Lead.",
 "Category":"Job family / grouping the role sits in.",
 "Grade":"Job grade (1-14) linking the role to the grading framework.",
 "IscoGroup":"ISCO-08 unit group code (4-digit) — indicative external mapping, verify with SME.",
 "IscoTitle":"Official ISCO-08 unit group title for the code above.",
 "EscoLabel":"ESCO preferred occupation label — indicative external mapping, verify with SME.",
 # JobProfiles
 "Description":"Short summary of the role's purpose.",
 "KeyResponsibilities":"Primary accountabilities (semicolon-separated).",
 "RequiredSkills":"Headline skills expected (semicolon-separated free text).",
 "Specialisms":"Typical specialisation areas (semicolon-separated).",
 "ManagementLevel":"People-management scope: Individual Contributor or People Manager.",
 "TypicalTools":"Tools and systems commonly used in the role (semicolon-separated).",
 # SalaryBands
 "Min":"Bottom of the pay range for this function/level (EUR).",
 "P25":"25th percentile of the pay range.","P50":"Median of the pay range.",
 "P75":"75th percentile of the pay range.","Max":"Top of the pay range.",
 "Currency":"Currency of the salary band (default EUR).",
 # TitleMapping
 "ExistingTitle":"Raw/existing title used in source systems; maps to a canonical JobID.",
 # CareerPaths
 "NextJobID":"JobID of the typical next role in the career ladder.",
 "NextRole":"Standard title of the next role (denormalised for readability).",
 # Levels / Categories
 "Order":"Sort order of the seniority level.",
 # Skills
 "SkillID":"Unique skill identifier (primary key).","SkillName":"Human-readable skill name.",
 "Definition":"What the skill means in practice.",
 # CompetencyLevels / SkillProficiency
 "Name":"Name of the competency level (Awareness ... Expert).",
 "LevelName":"Name of the proficiency level (1=Awareness ... 5=Expert).",
 "Anchor":"Behavioural description of what this proficiency level looks like for the skill category.",
 # RoleSkillMap
 "RequiredLevel":"Required proficiency (1-5) for the skill in this role.",
 "SkillType":"Skill relevance to the role: Core, Adjacent, or Leadership.",
 # JobGrades
 "GradeLabel":"Display label for the grade.","CareerBand":"Career band grouping the grade.",
 "LevelBand":"Sub-band label within the career band.",
 "HayMin":"Lower bound of the Hay-style point range.","HayMax":"Upper bound of the Hay-style point range.",
 "PayMin":"Grade pay minimum.","PayP25":"Grade pay 25th percentile.","PayP50":"Grade pay median.",
 "PayP75":"Grade pay 75th percentile.","PayMax":"Grade pay maximum.",
 "Scope":"Breadth of the role's remit at this grade.","Complexity":"Complexity of problems handled.",
 "Autonomy":"Degree of independence.","Impact":"Business impact of the role.",
 "Leadership":"People-leadership expectation at this grade.","SpanOfControl":"Typical number of reports.",
 "DecisionRights":"Decisions the grade is empowered to make.",
 "Responsibilities":"Headline responsibilities at this grade.",
 "Authority":"Decision and approval authority granted at this grade.",
 # Industries
 "IndustryID":"Unique industry identifier.","IndustryName":"Industry name.",
 "Scope":"What the industry covers.","Characteristics":"Distinguishing traits of the industry.",
 "Factor":"Salary multiplier applied for this industry/function vs. the baseline band.",
 "DefaultLevel":"Default required proficiency for an industry-specific skill.",
 # SeniorityLevels
 "LCode":"Seniority ladder code (L1-L5).","LName":"Name of the seniority level.",
 "MapsToLevel":"Base level (Junior/Medior/Senior/Lead) this L-code maps to.",
 "GradeRange":"Grade range associated with this level.","Grades":"Explicit grades in range.",
 # CareerBands
 "Band":"Career band name.","Grades ":"Grades in the band.","Focus":"Primary focus of the band.",
 "TypicalTitles":"Representative titles in the band.","ProgressionPath":"How people progress through the band.",
 "KeyDifferentiator":"What separates this band from adjacent ones.",
}
if "DataDictionary" in wb.sheetnames:
    del wb["DataDictionary"]
dd = wb.create_sheet("DataDictionary")
dd.append(["Sheet", "Column", "Description", "Owner"])
for cell in dd[1]:
    cell.font = HDR_FONT; cell.fill = HDR_FILL
# put the dictionary right after Jobs for discoverability
order = [s for s in wb.sheetnames if s not in ("DataDictionary",)]
for sheet_name in order:
    ws = wb[sheet_name]
    if ws.max_row < 1:
        continue
    owner = OWNER.get(sheet_name, "—")
    for c in ws[1]:
        if c.value is None:
            continue
        col = str(c.value)
        dd.append([sheet_name, col, DESC.get(col, f"{col} (see sheet)"), owner])

wb.save(WB)
print(f"CareerPaths rows added: {added_cp}")
print(f"ManagementLevel blanks filled: {filled_ml}")
print(f"SkillProficiency rows: {(sp.max_row-1)}")
print(f"DataDictionary rows: {(dd.max_row-1)}")
print("Sheets now:", wb.sheetnames)
print("Saved", WB)
