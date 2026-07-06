"""
Add an SBI 2008 / NACE Rev. 2 industry-code crosswalk to the Industries sheet
in jobsy_reference_library.xlsx, in place (openpyxl).

SBI 2008 (CBS, the Dutch Standard Industrial Classification) and NACE Rev. 2
(Eurostat, the EU standard) share identical codes up to the 4-digit class
level — SBI is the Dutch-published implementation of NACE Rev. 2 — so each
IND-XXX category gets the same code list under both columns, with each
column's own citation kept distinct (CBS vs. Eurostat) since a Dutch client
looks for the CBS/KVK-familiar SBI framing while EU-wide filtering wants NACE.

Sources:
  CBS  — https://www.cbs.nl/nl-nl/onze-diensten/methoden/classificaties/activiteiten/standaard-bedrijfsindeling--sbi--/de-structuur-van-de-sbi-2008-versie-2018-update-2022
  CBS  — https://www.cbs.nl/nl-nl/onze-diensten/methoden/begrippen/standaard-bedrijfsindeling-2008--sbi-2008--
  Eurostat — https://ec.europa.eu/eurostat/web/nace-rev2
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"
HDR_FILL = PatternFill("solid", fgColor="0E7C66")
HDR_FONT = Font(bold=True, color="FFFFFF")

CBS_URL = "https://www.cbs.nl/nl-nl/onze-diensten/methoden/classificaties/activiteiten/standaard-bedrijfsindeling--sbi--/de-structuur-van-de-sbi-2008-versie-2018-update-2022"
EUROSTAT_URL = "https://ec.europa.eu/eurostat/web/nace-rev2"

# IND-XXX -> (SBI 2008 / NACE Rev.2 codes, human-readable section scope)
CROSSWALK = {
    "IND-TECH": ("J58.2, J62, J63",
                 "Section J (Information & communication): software publishing, "
                 "computer programming/consultancy, information services"),
    "IND-FIN":  ("K64, K65, K66",
                 "Section K (Financial & insurance activities): financial services, "
                 "insurance/reinsurance/pension funding, auxiliary activities"),
    "IND-HLTH": ("Q86, Q87, Q88, C21, C26.6",
                 "Section Q (Human health & social work): hospitals, residential/social "
                 "care; plus Section C: C21 pharma manufacturing, C26.6 medtech/electromedical"),
    "IND-MFG":  ("C10-C12, C20, C24-C25, C28, C29",
                 "Section C (Manufacturing): food/beverage, chemicals, basic/fabricated "
                 "metals, machinery, motor vehicles (CAO Metalektro-adjacent divisions)"),
    "IND-RET":  ("G45, G46, G47, G47.91",
                 "Section G (Wholesale & retail trade): motor vehicle trade, wholesale, "
                 "retail incl. G47.91 internet/mail-order retail"),
    "IND-PUB":  ("O84, P85, Q88, S94",
                 "Section O public administration, Section P education, Q88 social work "
                 "without accommodation, S94 membership organisations (NGOs)"),
    "IND-PSV":  ("M69, M70, M71, M73",
                 "Section M (Professional, scientific & technical): legal/accounting, "
                 "management consultancy, architecture/engineering, advertising/research"),
    "IND-LOG":  ("H49, H50, H51, H52, H53",
                 "Section H (Transportation & storage): land/water/air transport, "
                 "warehousing, postal & courier"),
}

wb = load_workbook(WB)
ws = wb["Industries"]


def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


def add_column(ws, name, value_fn):
    h = headers(ws)
    if name in h:
        col = h[name]
    else:
        col = ws.max_column + 1
        c = ws.cell(1, col, name)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    for r in range(2, ws.max_row + 1):
        ws.cell(r, col, value_fn(r))
    return col


h = headers(ws)
IID = h["IndustryID"]


def iid_at(r):
    return str(ws.cell(r, IID).value).strip()


add_column(ws, "SBI2008", lambda r: CROSSWALK.get(iid_at(r), ("", ""))[0])
add_column(ws, "SBI2008Source", lambda r: CBS_URL if iid_at(r) in CROSSWALK else "")
add_column(ws, "NACERev2", lambda r: CROSSWALK.get(iid_at(r), ("", ""))[0])
add_column(ws, "NACERev2Source", lambda r: EUROSTAT_URL if iid_at(r) in CROSSWALK else "")
add_column(ws, "CodeScope", lambda r: CROSSWALK.get(iid_at(r), ("", ""))[1])
add_column(ws, "CodeCrosswalkUpdatedAt", lambda r: TODAY)

for w, col in zip([12, 18, 55, 12, 18, 60, 18], "BCDEFGH"):
    ws.column_dimensions[col].width = w

wb.save(WB)
print("Industries sheet columns now:", list(headers(ws).keys()))
for r in range(2, ws.max_row + 1):
    print(iid_at(r), "->", ws.cell(r, headers(ws)["SBI2008"]).value)
print("Saved", WB)
