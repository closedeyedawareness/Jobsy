"""
Add Employee Benefits Benchmarking data to jobsy_reference_library.xlsx in place.

Adds 4 sheets, mirroring the SalaryBands / IndustrySalaryFactors / SalarySources
governance pattern used for Pay Benchmarking:

  BenefitsCatalog       — 9 benefit categories (what they are, how they're expressed)
  BenefitsSources       — provenance/citations for the calibration
  BenefitsObservations  — self-built raw survey-style data points (IndustryID x Category),
                          the reference "microdata" that P25/P50/P75/P90 + market median
                          are computed FROM at runtime (services/benefits_service.py),
                          rather than static pre-set columns like SalaryBands.
  LevelBenefitsFactors  — Level x Category multiplier applied on top of the observed
                          industry distribution (mirrors IndustrySalaryFactors' role
                          for pay: a secondary, structural adjustment).

Deterministic (seeded) so re-running reproduces the same numbers.
"""
import random

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"
HDR_FILL = PatternFill("solid", fgColor="0E7C66")
HDR_FONT = Font(bold=True, color="FFFFFF")

random.seed(20260706)

wb = load_workbook(WB)


def new_sheet(name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_header(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")


# ── BenefitsCatalog ─────────────────────────────────────────────────────────
CATEGORIES = [
    # Category, Basis, Unit, TypicalValueDescription, StatutoryNL, Taxable, Description
    ("Pension", "% of pensionable base (employer, above sector-fund minimum)", "% of base",
     "~8-16% employer contribution", "Partly (sector funds)", "No (employer cost)",
     "Aanvullend pension: employer contribution beyond the mandatory sector-fund minimum."),
    ("Health Insurance", "Fixed monthly employer contribution", "EUR/month",
     "EUR 20-70/month", "No", "Yes",
     "Employer contribution toward the employee's zorgverzekering (basic + aanvullend)."),
    ("Extra Leave", "Days above the statutory minimum (20 days FTE)", "days/year",
     "0-12 extra days", "No (statutory min. is 20 days)", "N/A",
     "Additional paid leave days (verlof/ATV) beyond the statutory minimum."),
    ("Wellness & Vitality", "Fixed annual budget", "EUR/year",
     "EUR 0-800/year", "No", "Mixed",
     "Gym membership, mental-health support, vitality programmes."),
    ("Mobility & Commute", "Fixed monthly budget or lease value", "EUR/month",
     "EUR 50-400/month", "Partly (thuiswerkvergoeding is tax-free up to a cap)", "Mixed",
     "Commute allowance, home-working allowance, or lease-car/mobility budget."),
    ("Meal & Lunch", "Fixed monthly allowance or in-kind lunch", "EUR/month",
     "EUR 0-100/month", "No", "Yes",
     "Lunch vouchers or an in-office lunch allowance."),
    ("Disability Top-up", "% of base insured above the WIA cap", "% of base",
     "0-8% equivalent", "No (WIA-hiaatverzekering is voluntary)", "No (employer cost)",
     "WIA-gap insurance topping up income protection above the statutory WIA ceiling."),
    ("Parental Leave Top-up", "% of base paid during extended parental/birth leave", "% of base",
     "0-100% for several weeks", "Partly (statutory paid leave is a UWV percentage)", "Yes",
     "Employer top-up above the statutory UWV percentage during birth/parental leave."),
    ("Education & Development", "Fixed annual budget", "EUR/year",
     "EUR 500-3000/year", "No", "No (if job-related)",
     "Training, certification and study budget beyond mandatory role requirements."),
]

cat_ws = new_sheet("BenefitsCatalog")
write_header(cat_ws, ["BenefitID", "Category", "Basis", "Unit", "TypicalValueDescription",
                      "StatutoryNL", "Taxable", "Description", "Source", "Owner", "Status",
                      "EffectiveFrom", "UpdatedAt"])
for i, (cat, basis, unit, typ, stat_nl, taxable, desc) in enumerate(CATEGORIES, 1):
    cat_ws.append([f"BEN-{i:02d}", cat, basis, unit, typ, stat_nl, taxable, desc,
                   "Seed v1 (benefits)", "Reward", "Active", TODAY, TODAY])
for w, col in zip([10, 22, 40, 12, 24, 30, 22, 60, 18, 10, 10, 14, 12], "ABCDEFGHIJKLM"):
    cat_ws.column_dimensions[col].width = w

# ── BenefitsSources ──────────────────────────────────────────────────────────
SOURCES = [
    ("Netherlands Benefits Survey", "Mercer", "Benefits survey", "https://www.mercer.com/nl",
     "Employer-paid benefits norms: pension top-up, health & wellness richness by sector."),
    ("Employee Benefits Watch", "Aon", "Benefits survey", "https://www.aon.com/netherlands",
     "Benefits prevalence and typical value ranges by industry."),
    ("Pensioenmonitor", "Pensioenfederatie", "Official statistics", "https://www.pensioenfederatie.nl",
     "Sector-fund pension contribution norms (baseline for the Pension category)."),
    ("Verlof en vakantie", "Rijksoverheid", "Official guidance",
     "https://www.rijksoverheid.nl/onderwerpen/verlof-en-vakantie",
     "Statutory leave minimums (baseline for the Extra Leave category)."),
    ("WIA-uitkering en aanvullende verzekeringen", "UWV", "Official guidance",
     "https://www.uwv.nl/particulieren/wia-uitkering", "WIA-gap (disability top-up) insurance norms."),
    ("Arbeidsvoorwaardenonderzoek", "CBS (Statistics NL)", "Official statistics", "https://www.cbs.nl",
     "Employer secondary-benefits prevalence used to sanity-check category spread."),
]
src_ws = new_sheet("BenefitsSources")
write_header(src_ws, ["Source", "Publisher", "Type", "URL", "AccessedOn", "Informs"])
for name, pub, typ, url, informs in SOURCES:
    src_ws.append([name, pub, typ, url, TODAY, informs])
src_ws.append([])
src_ws.append(["METHOD",
               "BenefitsObservations rows are self-built illustrative data points (Seed v1 benefits), "
               "calibrated in aggregate to the sources above — not verbatim survey microdata. "
               "P25/P50/P75/P90 and the market median in the app are COMPUTED from these observations "
               "at load time (see services/benefits_service.py), the same way a real survey vendor's "
               "underlying respondent data would be aggregated.",
               "", "", "", ""])
for w, col in zip([32, 20, 20, 46, 14, 60], "ABCDEF"):
    src_ws.column_dimensions[col].width = w

# ── BenefitsObservations (raw data -> computed percentiles) ────────────────
INDUSTRIES = ["IND-TECH", "IND-FIN", "IND-HLTH", "IND-MFG", "IND-RET", "IND-PUB", "IND-PSV", "IND-LOG"]
INDUSTRY_RICHNESS = {
    "IND-TECH": 1.25, "IND-FIN": 1.30, "IND-HLTH": 0.95, "IND-MFG": 0.95,
    "IND-RET": 0.75, "IND-PUB": 0.90, "IND-PSV": 1.15, "IND-LOG": 0.80,
}
CATEGORY_BASELINE = {
    # category -> (mean, relative jitter, currency/unit)
    "Pension": (11.0, 0.22, "%"),
    "Health Insurance": (40.0, 0.30, "EUR"),
    "Extra Leave": (5.0, 0.45, "days"),
    "Wellness & Vitality": (300.0, 0.55, "EUR"),
    "Mobility & Commute": (150.0, 0.35, "EUR"),
    "Meal & Lunch": (40.0, 0.35, "EUR"),
    "Disability Top-up": (3.0, 0.50, "%"),
    "Parental Leave Top-up": (55.0, 0.40, "%"),
    "Education & Development": (1500.0, 0.35, "EUR"),
}
CATEGORY_UNIT = {c: u for c, (_, _, u) in CATEGORY_BASELINE.items()}
N_OBS_PER_GROUP = 14

obs_ws = new_sheet("BenefitsObservations")
write_header(obs_ws, ["ObsID", "IndustryID", "Category", "Value", "Unit", "Currency",
                      "Source", "Owner", "Status", "EffectiveFrom", "UpdatedAt"])
obs_id = 1
for industry in INDUSTRIES:
    richness = INDUSTRY_RICHNESS[industry]
    for category, (mean, jitter, unit) in CATEGORY_BASELINE.items():
        group_mean = mean * richness
        for _ in range(N_OBS_PER_GROUP):
            noise = random.gauss(1.0, jitter)
            value = max(0.0, group_mean * noise)
            currency = "EUR" if unit == "EUR" else ("" if unit in ("%", "days") else "EUR")
            obs_ws.append([f"BO-{obs_id:05d}", industry, category, round(value, 2), unit, currency,
                           "Seed v1 (benefits)", "Reward", "Active", TODAY, TODAY])
            obs_id += 1
for w, col in zip([10, 12, 24, 10, 8, 10, 18, 10, 10, 14, 12], "ABCDEFGHIJK"):
    obs_ws.column_dimensions[col].width = w

# ── LevelBenefitsFactors ────────────────────────────────────────────────────
LEVEL_BASE = {"Junior": 0.85, "Medior": 1.00, "Senior": 1.15, "Lead": 1.35}
FLAT_CATEGORIES = {"Meal & Lunch"}  # same for everyone regardless of level
LOW_TILT_CATEGORIES = {"Extra Leave": {"Junior": 0.90, "Medior": 1.00, "Senior": 1.05, "Lead": 1.10},
                       "Health Insurance": {"Junior": 0.95, "Medior": 1.00, "Senior": 1.05, "Lead": 1.10}}
LEVELS = ["Junior", "Medior", "Senior", "Lead"]

lf_ws = new_sheet("LevelBenefitsFactors")
write_header(lf_ws, ["Level", "Category", "Factor", "Source", "Owner", "Status",
                     "EffectiveFrom", "UpdatedAt"])
for category in CATEGORY_BASELINE:
    for level in LEVELS:
        if category in FLAT_CATEGORIES:
            factor = 1.0
        elif category in LOW_TILT_CATEGORIES:
            factor = LOW_TILT_CATEGORIES[category][level]
        else:
            factor = LEVEL_BASE[level]
        lf_ws.append([level, category, factor, "Seed v1 (benefits)", "Reward", "Active", TODAY, TODAY])
for w, col in zip([10, 24, 10, 18, 10, 10, 14, 12], "ABCDEFGH"):
    lf_ws.column_dimensions[col].width = w

# ── keep DataDictionary sheet at the end if present, benefits sheets before it ──
if "DataDictionary" in wb.sheetnames:
    order = [s for s in wb.sheetnames if s != "DataDictionary"] + ["DataDictionary"]
    wb._sheets = [wb[s] for s in order]

wb.save(WB)
print("Sheets now:", wb.sheetnames)
print(f"BenefitsCatalog: {cat_ws.max_row - 1} rows")
print(f"BenefitsSources: {src_ws.max_row - 1} rows")
print(f"BenefitsObservations: {obs_ws.max_row - 1} rows")
print(f"LevelBenefitsFactors: {lf_ws.max_row - 1} rows")
print("Saved", WB)
