"""
Add Shift Allowance (Ploegentoeslag) as a benefits/perquisite category.

Genuinely missing from Jobsy until now: it isn't in PayElements/PayMix
(which only cover base/holiday/13th month/variable/pension/LTI) nor in the
original 9 or expanded 19 benefits categories. Real CAO data for it exists
(Shell, Zeeland Refinery both quote 30% for continuous/5-shift rosters),
so it's added here as its own category rather than folded into an
existing one.

Unlike most benefit categories, shift-allowance prevalence varies hugely
by industry — most Tech/Finance/Professional-Services roles never work
shifts at all, while Manufacturing/Healthcare/Logistics commonly do. The
generic INDUSTRY_RICHNESS multiplier (calibrated for benefits
"generosity") would misrepresent this, so a dedicated prevalence
multiplier is used instead for this category's synthetic baseline.

Level factor: flat (1.0) across Junior/Medior/Senior/Lead — the
percentage is tied to the shift SCHEDULE worked, not seniority level.
"""
import random

from openpyxl import load_workbook

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"

random.seed(20260708)

wb = load_workbook(WB)


def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


CATEGORY = "Shift Allowance (Ploegentoeslag)"

# ── 1. BenefitsCatalog ──────────────────────────────────────────────────────
cat_ws = wb["BenefitsCatalog"]
ch = headers(cat_ws)
existing_cats = {cat_ws.cell(r, ch["Category"]).value for r in range(2, cat_ws.max_row + 1)}
existing_ids = [str(cat_ws.cell(r, ch["BenefitID"]).value) for r in range(2, cat_ws.max_row + 1)
                if cat_ws.cell(r, ch["BenefitID"]).value]
next_bid = max((int(x.split("-")[1]) for x in existing_ids if x.startswith("BEN-")), default=0) + 1

if CATEGORY not in existing_cats:
    cat_ws.append([
        f"BEN-{next_bid:02d}", CATEGORY,
        "% of base salary for continuous/multi-shift work schedules", "%",
        "11.5-30% depending on shift pattern (highest for 5-shift/continuous 24/7 rosters)",
        "No (CAO-negotiated)", "Yes",
        "Premium paid for working shift patterns (2-ploegendienst, 5-ploegendienst/volcontinu), "
        "scaled by pattern intensity — typically highest for round-the-clock continuous rosters. "
        "Common in manufacturing/refining/healthcare/logistics; rare in office-based roles.",
        "Seed v2 (benefits, taxonomy expansion)", "Reward", "Active", TODAY, TODAY,
    ])

# ── 2. BenefitsObservations: synthetic baseline, prevalence-adjusted ───────
INDUSTRIES = ["IND-TECH", "IND-FIN", "IND-HLTH", "IND-MFG", "IND-RET", "IND-PUB", "IND-PSV", "IND-LOG"]
# Shift-work PREVALENCE per industry (distinct from the generic "generosity" richness
# multiplier) — most office-based industries have near-zero real shift-allowance exposure.
SHIFT_PREVALENCE = {
    "IND-MFG": 1.00, "IND-HLTH": 0.90, "IND-LOG": 1.00, "IND-RET": 0.50,
    "IND-PUB": 0.30, "IND-TECH": 0.10, "IND-FIN": 0.10, "IND-PSV": 0.10,
}
BASELINE_MEAN = 20.0   # midpoint of the real 11.5-30% range, before prevalence scaling
JITTER = 0.35
N_OBS_PER_GROUP = 14

obs_ws = wb["BenefitsObservations"]
oh = headers(obs_ws)
existing_obs_ids = [str(obs_ws.cell(r, oh["ObsID"]).value) for r in range(2, obs_ws.max_row + 1)
                    if obs_ws.cell(r, oh["ObsID"]).value]
next_obs_id = max((int(x.split("-")[1]) for x in existing_obs_ids if x.startswith("BO-")), default=0) + 1
n_cols = obs_ws.max_column


def append_obs_row(industry_id, value, unit, source, company, notes, status="Active"):
    global next_obs_id
    row = [None] * n_cols
    row[oh["ObsID"] - 1] = f"BO-{next_obs_id:05d}"
    row[oh["IndustryID"] - 1] = industry_id
    row[oh["Category"] - 1] = CATEGORY
    row[oh["Value"] - 1] = value
    row[oh["Unit"] - 1] = unit
    row[oh["Currency"] - 1] = ""
    row[oh["Source"] - 1] = source
    row[oh["Owner"] - 1] = "Reward"
    row[oh["Status"] - 1] = status
    row[oh["EffectiveFrom"] - 1] = TODAY
    row[oh["UpdatedAt"] - 1] = TODAY
    if "CompanyName" in oh:
        row[oh["CompanyName"] - 1] = company
    if "Notes" in oh:
        row[oh["Notes"] - 1] = notes
    obs_ws.append(row)
    next_obs_id += 1


n_synthetic = 0
for industry in INDUSTRIES:
    prevalence = SHIFT_PREVALENCE[industry]
    group_mean = BASELINE_MEAN * prevalence
    for _ in range(N_OBS_PER_GROUP):
        noise = random.gauss(1.0, JITTER)
        value = max(0.0, group_mean * noise)
        append_obs_row(industry, round(value, 2), "%",
                       "Seed v2 (benefits, taxonomy expansion)", "", "", "Active")
        n_synthetic += 1

# ── 3. Real, cited observations (Shell & Zeeland Refinery, 5-shift/continuous) ──
REAL_OBSERVATIONS = [
    dict(industry_id="IND-MFG", value=30.0, source="Shell Nederland Raffinaderij CAO 2022-2025",
         company="Shell Nederland Raffinaderij",
         notes="CAO: vijfploegendienst (5-shift, continuous 24/7) = 30% of base salary/month. "
               "Lower tiers also documented: tweeploegendienst ma-vr = 11.5%, ma-za = 15%. "
               "30% (the highest/most comparable tier) used here.", status="Active"),
    dict(industry_id="IND-MFG", value=30.0, source="Zeeland Refinery CAO 2024-2025",
         company="Zeeland Refinery",
         notes="CAO: ploegentoeslag volcontinudienst (continuous shift) = 30% of base salary/month, "
               "for all types of volcontinudienst. CAO also documents a tenure-based step-down "
               "schedule when moving to day shift (12% at 0-3yr up to 30% at 8yr+).", status="Active"),
]
for obs in REAL_OBSERVATIONS:
    append_obs_row(obs["industry_id"], obs["value"], "%", obs["source"], obs["company"],
                   obs["notes"], obs["status"])

# ── 4. LevelBenefitsFactors: flat across all 4 levels ──────────────────────
lf_ws = wb["LevelBenefitsFactors"]
lh = headers(lf_ws)
existing_pairs = {(lf_ws.cell(r, lh["Level"]).value, lf_ws.cell(r, lh["Category"]).value)
                   for r in range(2, lf_ws.max_row + 1)}
n_factors = 0
for level in ["Junior", "Medior", "Senior", "Lead"]:
    if (level, CATEGORY) in existing_pairs:
        continue
    lf_ws.append([level, CATEGORY, 1.0, "Seed v2 (benefits, taxonomy expansion)",
                 "Reward", "Active", TODAY, TODAY])
    n_factors += 1

wb.save(WB)
print(f"BenefitsCatalog: +1 category ({CATEGORY})")
print(f"BenefitsObservations: +{n_synthetic} synthetic, +{len(REAL_OBSERVATIONS)} real")
print(f"LevelBenefitsFactors: +{n_factors} rows")
print("Saved", WB)
