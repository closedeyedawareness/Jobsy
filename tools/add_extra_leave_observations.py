"""
Blend real Extra Leave data points into BenefitsObservations.

Source: a user-supplied CAO/CSV comparison table (Dec 2025) giving total
annual vacation days (Vakantiedagen) for 12 companies. Extra Leave = total
days - 20 (the NL statutory minimum, matching this category's Basis text).

Only Extra Leave is imported from this table — everything else in the
source (ploegentoeslag %, overwerk %, loonsverhoging, pensioenopbouw %
accrual rate, salary ranges, working hours) is either pay-side data
already tracked elsewhere, or a different metric than this category
tracks (accrual rate vs. employer contribution %), and is deliberately
NOT blended in here to avoid corrupting category definitions.

BP was skipped: its CAO gives no specific day count ("CAO-bepaald").
Pharma companies (MSD, Bristol-Myers Squibb, Pfizer, J&J) are classified
IND-HLTH (pharma manufacturing, SBI/NACE C21) rather than IND-MFG.
"""
from openpyxl import load_workbook

WB = "jobsy_reference_library.xlsx"
TODAY = "2026-07-06"

COMPANIES = [
    # (company_name, total_vacation_days, industry_id)
    ("Shell Nederland Raffinaderij", 33, "IND-MFG"),
    ("Zeeland Refinery", 32, "IND-MFG"),
    ("Yara Sluiskil/Vlaardingen", 30, "IND-MFG"),
    ("Ecolab Production Netherlands", 39, "IND-MFG"),
    ("MSD Netherlands", 32, "IND-HLTH"),
    ("Bristol-Myers Squibb Netherlands", 32.5, "IND-HLTH"),  # source gave a 30-35 range; midpoint used
    ("Linde Netherlands", 32, "IND-MFG"),
    ("LyondellBasell Netherlands", 32, "IND-MFG"),
    ("Pfizer Netherlands", 33, "IND-HLTH"),
    ("Unilever Netherlands", 33, "IND-MFG"),  # FMCG manufacturing — closest existing bucket, imperfect fit
    ("Johnson & Johnson Netherlands", 36, "IND-HLTH"),
    ("Dow Chemical Netherlands", 28, "IND-MFG"),  # base CAO days; CAO also allows buying up to +25 more (not counted here)
]
STATUTORY_MIN = 20
SOURCE = "NL chemicals/pharma CAO comparison (user-supplied CSV, Dec 2025)"

wb = load_workbook(WB)


def headers(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


# ── BenefitsSources citation ────────────────────────────────────────────────
src_ws = wb["BenefitsSources"]
sh = headers(src_ws)
existing = {src_ws.cell(r, sh["Source"]).value for r in range(2, src_ws.max_row + 1)}
if SOURCE not in existing:
    src_ws.append([SOURCE, "Various CAO texts (FNV/CNV) + company HR materials",
                   "Company CAO comparison (user-supplied compilation)", "", TODAY,
                   "Extra Leave (total vacation days) for 12 NL chemical/pharma companies"])

# ── BenefitsObservations rows ────────────────────────────────────────────────
obs_ws = wb["BenefitsObservations"]
oh = headers(obs_ws)
existing_ids = [str(obs_ws.cell(r, oh["ObsID"]).value) for r in range(2, obs_ws.max_row + 1)
                if obs_ws.cell(r, oh["ObsID"]).value]
next_id = max((int(x.split("-")[1]) for x in existing_ids if x.startswith("BO-")), default=0) + 1

added = 0
for company, total_days, industry_id in COMPANIES:
    extra = total_days - STATUTORY_MIN
    row = [None] * obs_ws.max_column
    row[oh["ObsID"] - 1] = f"BO-{next_id:05d}"
    row[oh["IndustryID"] - 1] = industry_id
    row[oh["Category"] - 1] = "Extra Leave"
    row[oh["Value"] - 1] = extra
    row[oh["Unit"] - 1] = "days"
    row[oh["Currency"] - 1] = ""
    row[oh["Source"] - 1] = SOURCE
    row[oh["Owner"] - 1] = "Reward"
    row[oh["Status"] - 1] = "Active"
    row[oh["EffectiveFrom"] - 1] = TODAY
    row[oh["UpdatedAt"] - 1] = TODAY
    row[oh["CompanyName"] - 1] = company
    row[oh["Notes"] - 1] = f"{total_days:g} total vacation days per CAO, minus {STATUTORY_MIN} statutory minimum."
    obs_ws.append(row)
    next_id += 1
    added += 1

wb.save(WB)
print(f"Added {added} Extra Leave observations across {len(set(c[2] for c in COMPANIES))} industries.")
print("Saved", WB)
