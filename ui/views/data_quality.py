"""ui/views/data_quality.py — moved verbatim out of ui/app.py on 2026-09-03."""

from __future__ import annotations

from ui.shared import *  # noqa: F401,F403


# How old a sheet may get before the scorecard stops calling it current. Salary
# and benchmark data is the reason there is a threshold at all: a band nobody
# has revisited in a year is not wrong in any way validation can see, and is
# still the thing most likely to be quietly out of date.
FRESH_DAYS = 90


STALE_DAYS = 365


def _unloaded_sheets(catalog):
    """What the source holds that SHEET_MAP does not load.

    Derived from the source itself, never from a list kept by hand: the DB's own
    table specs when the library came from Postgres, the workbook's sheet names
    when it came from the file. A sheet nobody loads is invisible to every panel
    on this page, which is the one place that should say so.
    """
    from core.catalog import SHEET_MAP
    loaded = set(SHEET_MAP)
    source = getattr(catalog, "active_source", None)
    try:
        if source == "db":
            from services.library_import_service import SPECS
            present = {spec.sheet for spec in SPECS}
        else:
            from openpyxl import load_workbook
            present = set(load_workbook(WORKBOOK_PATH, read_only=True).sheetnames)
    except Exception:
        return []
    return sorted(present - loaded)


def _dq_freshness(catalog):
    """Per-sheet UpdatedAt / Source coverage of the library the app has loaded.

    Every sheet but Employees and DataDictionary carries UpdatedAt and Source
    columns, populated for essentially every row — a provenance trail kept by
    hand and, until the scorecard existed, read by nothing.

    Until 2026-09-03 this read jobsy_reference_library.xlsx directly. That
    stopped being true at the cutover: Postgres became the master and the
    scorecard went on describing a file nothing else in the app reads, which is
    the exact failure — a freshness panel that is itself out of date — this
    panel exists to catch. It now measures catalog.frames, the frames the
    Repository was actually built from, from whichever source they came.
    """
    import pandas as _pd
    from core.catalog import SHEET_MAP
    out = []
    frames = getattr(catalog, "frames", None) or {}
    book = {sheet: frames[key] for sheet, key in SHEET_MAP.items() if key in frames}
    for sheet, df in book.items():
        if df is None or len(df) == 0:
            out.append({"sheet": sheet, "rows": 0, "updated_pct": None,
                        "source_pct": None, "newest": None})
            continue
        def _cov(col):
            return round(df[col].notna().mean() * 100) if col in df.columns else None
        newest = None
        if "UpdatedAt" in df.columns:
            dates = _pd.to_datetime(df["UpdatedAt"], errors="coerce").dropna()
            if len(dates):
                newest = dates.max().date()
        out.append({"sheet": sheet, "rows": len(df),
                    "updated_pct": _cov("UpdatedAt"), "source_pct": _cov("Source"),
                    "newest": newest})
    return out


def data_quality_page(catalog):
    """Live data-quality scorecard for the reference library."""
    repo = catalog.repository
    st.markdown(
        f'<div style="font-family:{FONT_SERIF};font-size:28px;font-weight:600;'
        f'letter-spacing:-0.02em;margin-bottom:4px">Data Quality</div>'
        f'<p style="color:{C["muted"]};font-size:14px;margin-bottom:16px">'
        f'A live scorecard for the reference library — coverage, integrity and freshness. '
        f'Run it after every edit to catch gaps before they reach users.</p>',
        unsafe_allow_html=True,
    )
    # Export the library as a workbook snapshot. Since the cutover the workbook
    # in the repo is not the master and cannot be trusted as a copy of it; this
    # button produces one from whatever the app is actually reading.
    _src_label = getattr(catalog, "active_source", None) or "the library"
    _c_exp, _c_note = st.columns([1, 3])
    with _c_exp:
        try:
            from services.library_export_service import LibraryExportService
            _exporter = LibraryExportService(catalog)
            st.download_button(
                "Export library to Excel",
                data=_exporter.to_bytes(),
                file_name=_exporter.suggested_filename(),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="A snapshot of the library the app is reading right now.",
            )
        except Exception as _exc:
            st.caption(f"Export unavailable: {_exc}")
    with _c_note:
        st.caption(f"Snapshot of the library as loaded from **{_src_label}**, with an ExportInfo "
                   f"sheet recording the source, the time and the row counts. It restores the "
                   f"sheets the app loads — not the database's audit history.")

    jobs = list(repo.jobs.values()); n = max(len(jobs), 1)
    # From the loaded library, not the workbook: same reason as _dq_freshness.
    jraw = catalog.frames.get("jobs")
    tm = catalog.frames.get("titles")

    iso = own = {}
    if jraw is not None:
        iso = {str(r["JobID"]).strip(): str(r.get("IscoGroup", "")).strip() not in ("", "nan")
               for _, r in jraw.iterrows()}
        own = {str(r["JobID"]).strip(): str(r.get("Owner", "")).strip() not in ("", "nan")
               for _, r in jraw.iterrows()}
    syn_ids = {}
    if tm is not None:
        for _, r in tm.iterrows():
            syn_ids[str(r["JobID"]).strip()] = syn_ids.get(str(r["JobID"]).strip(), 0) + 1

    def _profile(j):
        p = repo.profiles.get(j.job_id); return bool(p and (p.description or p.key_responsibilities))
    dims = {
        "Profile":     _profile,
        "Salary band": lambda j: (j.function, j.level) in repo.salary,
        "Skills":      lambda j: len(repo.role_skill_map.get(j.job_id, [])) > 0,
        "Grade":       lambda j: (getattr(j, "grade", 0) or 0) > 0,
        "Career path": lambda j: j.job_id in repo.career_paths or j.standard_title == "Chief Executive Officer",
        "ISCO code":   lambda j: iso.get(j.job_id, False),
        "Synonyms":    lambda j: syn_ids.get(j.job_id, 0) > 0,
        "Owner":       lambda j: own.get(j.job_id, False),
    }
    cov = {name: sum(1 for j in jobs if fn(j)) for name, fn in dims.items()}
    health = round(sum(cov.values()) / (len(dims) * n) * 100)

    # ── headline tiles ──────────────────────────────────────────────────
    hcol = C["teal"] if health >= 90 else (C["amber"] if health >= 70 else C["danger"])
    tiles = [("Health score", f"{health}%", hcol), ("Roles", str(len(jobs)), C["ink"]),
             ("Functions", str(len(repo.jobs_by_function)), C["ink"]),
             ("Title synonyms", str(len(tm)) if tm is not None else "—", C["ink"])]
    trow = "".join(
        f'<div style="flex:1;min-width:120px;background:{C["surface"]};border:1px solid {C["line"]};'
        f'border-radius:12px;padding:14px 16px">'
        f'<div style="font-family:{FONT_SERIF};font-size:30px;font-weight:700;color:{col}">{val}</div>'
        f'<div style="font-family:{FONT_MONO};font-size:10px;letter-spacing:.08em;text-transform:uppercase;'
        f'color:{C["muted"]};margin-top:2px">{lab}</div></div>'
        for lab, val, col in tiles)
    st.markdown(f'<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:18px">{trow}</div>',
                unsafe_allow_html=True)

    # ── coverage bars ───────────────────────────────────────────────────
    st.markdown(f'<div style="font-family:{FONT_MONO};font-size:11px;letter-spacing:.12em;'
                f'text-transform:uppercase;color:{C["muted"]};margin:6px 0 8px">Coverage</div>',
                unsafe_allow_html=True)
    bars = ""
    for name, cnt in cov.items():
        pct = round(cnt / n * 100)
        bc = C["teal"] if pct == 100 else (C["amber"] if pct >= 80 else C["danger"])
        bars += (
            f'<div style="display:flex;align-items:center;gap:12px;margin:5px 0">'
            f'<span style="flex:0 0 130px;font-size:13px;color:{C["ink"]}">{name}</span>'
            f'<span style="flex:1;background:{C["line"]};border-radius:999px;height:10px;position:relative">'
            f'<span style="position:absolute;left:0;top:0;height:10px;width:{pct}%;background:{bc};border-radius:999px"></span></span>'
            f'<span style="flex:0 0 78px;text-align:right;font-family:{FONT_MONO};font-size:12px;color:{bc}">'
            f'{pct}% · {cnt}/{n}</span></div>')
    st.markdown(bars, unsafe_allow_html=True)

    # ── integrity checks ────────────────────────────────────────────────
    ids = [j.job_id for j in jobs]
    dupes = {x for x in ids if ids.count(x) > 1}
    bad_ord = [f"{k[0]}/{k[1]}" for k, b in repo.salary.items() if not (b.min <= b.p50 <= b.max)]
    dang_cp = [jid for jid, cs in repo.career_paths.items()
               if cs.next_job_id and cs.next_job_id not in repo.jobs]
    dang_tm = sorted({str(r["JobID"]).strip() for _, r in tm.iterrows()
                      if str(r["JobID"]).strip() not in repo.jobs}) if tm is not None else []
    checks = [
        ("No duplicate JobIDs", not dupes, ", ".join(sorted(dupes))),
        ("Salary min ≤ P50 ≤ max", not bad_ord, ", ".join(bad_ord)),
        ("Career paths resolve", not dang_cp, ", ".join(dang_cp)),
        ("Synonyms map to real roles", not dang_tm, ", ".join(dang_tm)),
    ]
    st.markdown(f'<div style="font-family:{FONT_MONO};font-size:11px;letter-spacing:.12em;'
                f'text-transform:uppercase;color:{C["muted"]};margin:18px 0 8px">Integrity</div>',
                unsafe_allow_html=True)
    crows = ""
    for label, ok, detail in checks:
        icon = "✓" if ok else "✗"; col = C["teal"] if ok else C["danger"]
        crows += (f'<div style="display:flex;align-items:center;gap:10px;margin:4px 0;font-size:13px">'
                  f'<span style="color:{col};font-weight:700">{icon}</span>'
                  f'<span style="color:{C["ink"]}">{label}</span>'
                  f'<span style="color:{C["muted"]};font-size:12px">{("— "+detail) if detail else ""}</span></div>')
    st.markdown(crows, unsafe_allow_html=True)

    # ── validator report ────────────────────────────────────────────────
    # The checks above are hand-derived and role-shaped. This is the Validator
    # the Repository actually gates the load on — errors never reach here (they
    # raise), so what shows is its warnings, which used to be log-only.
    report = getattr(repo, "validation", None)
    st.markdown(f'<div style="font-family:{FONT_MONO};font-size:11px;letter-spacing:.12em;'
                f'text-transform:uppercase;color:{C["muted"]};margin:18px 0 8px">Validator</div>',
                unsafe_allow_html=True)
    if report is None:
        st.caption("This library was built with validation disabled — no report to show.")
    elif report.warnings:
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:10px;margin:4px 0;font-size:13px">'
            f'<span style="color:{C["amber"]};font-weight:700">!</span>'
            f'<span style="color:{C["ink"]}">Loaded with {len(report.warnings)} warning'
            f'{"s" if len(report.warnings) != 1 else ""}</span></div>', unsafe_allow_html=True)
        with st.expander(f"Validator warnings ({len(report.warnings)})"):
            for w in report.warnings:
                st.markdown(f"- {w}")
    else:
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:10px;margin:4px 0;font-size:13px">'
            f'<span style="color:{C["teal"]};font-weight:700">✓</span>'
            f'<span style="color:{C["ink"]}">Passed with no warnings</span></div>',
            unsafe_allow_html=True)

    # ── freshness ───────────────────────────────────────────────────────
    # Coverage says a field is filled in. It cannot say whether what is in it is
    # still true. UpdatedAt is the only thing in the library that can, so it is
    # worth reading even when — as now — the answer is mostly reassuring.
    import datetime as _dt
    fresh = _dq_freshness(catalog)
    dated = [r for r in fresh if r["newest"] is not None]
    st.markdown(f'<div style="font-family:{FONT_MONO};font-size:11px;letter-spacing:.12em;'
                f'text-transform:uppercase;color:{C["muted"]};margin:18px 0 8px">Freshness</div>',
                unsafe_allow_html=True)
    if not dated:
        st.caption("No UpdatedAt dates could be read from the workbook.")
    else:
        today = _dt.date.today()
        ages = {r["sheet"]: (today - r["newest"]).days for r in dated}
        last_touched = {r["sheet"]: r["newest"] for r in dated}
        oldest_sheet = max(ages, key=ages.get)
        stale = sorted([s for s, a in ages.items() if a > STALE_DAYS], key=lambda s: -ages[s])
        ageing = sorted([s for s, a in ages.items() if FRESH_DAYS < a <= STALE_DAYS],
                        key=lambda s: -ages[s])
        no_date = [r["sheet"] for r in fresh if r["newest"] is None and r["rows"] > 0]

        fcol = C["danger"] if stale else (C["amber"] if ageing else C["teal"])
        ftiles = [("Oldest sheet", f"{ages[oldest_sheet]}d", fcol),
                  (f"Ageing (>{FRESH_DAYS}d)", str(len(ageing)), C["amber"] if ageing else C["ink"]),
                  (f"Stale (>{STALE_DAYS}d)", str(len(stale)), C["danger"] if stale else C["ink"]),
                  ("Undated sheets", str(len(no_date)), C["amber"] if no_date else C["ink"])]
        frow = "".join(
            f'<div style="flex:1;min-width:120px;background:{C["surface"]};border:1px solid {C["line"]};'
            f'border-radius:12px;padding:14px 16px">'
            f'<div style="font-family:{FONT_SERIF};font-size:30px;font-weight:700;color:{col}">{val}</div>'
            f'<div style="font-family:{FONT_MONO};font-size:10px;letter-spacing:.08em;text-transform:uppercase;'
            f'color:{C["muted"]};margin-top:2px">{lab}</div></div>'
            for lab, val, col in ftiles)
        st.markdown(f'<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px">{frow}</div>',
                    unsafe_allow_html=True)
        st.caption(f"Oldest content is **{oldest_sheet}**, last touched {ages[oldest_sheet]} days ago "
                   f"({last_touched[oldest_sheet]}).")
        if no_date:
            st.caption("No UpdatedAt column: " + ", ".join(no_date) +
                       " — these carry no provenance date, so their age cannot be judged.")

        # Who changed what. The trail has been written since the schema landed
        # and read by nothing, which from a user's side is the same as no trail.
        if getattr(catalog, "active_source", None) == "db":
            with st.expander("Recent library changes"):
                try:
                    from core.db_loader import client_and_org
                    from services.library_history_service import recent_changes, summarise
                    _client, _org = client_and_org()
                    _hist = recent_changes(_client, _org, limit=200)
                except Exception as _exc:
                    _hist = None
                    st.caption(f"The change history could not be read: {_exc}")
                if _hist is not None:
                    if _hist.empty:
                        st.caption("No changes recorded yet — the library has not been written "
                                   "to since the audit trail was created.")
                    else:
                        _s = summarise(_hist)
                        st.caption(f"{_s['rows']} most recent changes across {_s['tables']} tables "
                                   f"— {_s['inserts']} inserts, {_s['updates']} updates, "
                                   f"{_s['deletes']} deletes. Latest: {_s['latest']}.")
                        st.dataframe(_hist.drop(columns=["Field count"]),
                                     use_container_width=True, hide_index=True)
                        st.caption("Append-only: migration 0003 revoked update and delete on the "
                                   "trail from every role, the importer's key included.")

        _unloaded = _unloaded_sheets(catalog)
        if _unloaded:
            st.caption("Present in the library but not loaded by the app, so nothing on this "
                       "page can see them: " + ", ".join(_unloaded) + ".")
        st.caption(f"Measured on the {len(fresh)} sheets the app loads, as read from "
                   f"**{getattr(catalog, 'active_source', None) or 'the library'}** — not on the "
                   f"workbook in the repo, which stopped being the master at the cutover.")

        with st.expander("Freshness and provenance by sheet"):
            import pandas as _pd
            rows = []
            for r in sorted(fresh, key=lambda r: (r["newest"] is None, -(ages.get(r["sheet"], 0)))):
                rows.append({
                    "Sheet": r["sheet"],
                    "Rows": r["rows"],
                    "Last updated": str(r["newest"]) if r["newest"] else "—",
                    "Age (days)": ages.get(r["sheet"], "—") if r["newest"] else "—",
                    "UpdatedAt filled": f'{r["updated_pct"]}%' if r["updated_pct"] is not None else "—",
                    "Source filled": f'{r["source_pct"]}%' if r["source_pct"] is not None else "—",
                })
            st.dataframe(_pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # ── attention list ──────────────────────────────────────────────────
    gaps = {name: [j.standard_title for j in jobs if not fn(j)]
            for name, fn in dims.items() if cov[name] < n}
    if gaps:
        with st.expander(f"⚠ Roles needing attention ({sum(len(v) for v in gaps.values())} gaps)"):
            for name, roles in gaps.items():
                st.markdown(f"**Missing {name}** ({len(roles)}): " + ", ".join(roles))
    else:
        st.success("✓ Every role is complete on all coverage dimensions.")

    # ── per-function completeness ───────────────────────────────────────
    with st.expander("Completeness by function"):
        import pandas as _pd
        rows = []
        for fnname, roles in sorted(repo.jobs_by_function.items()):
            m = len(roles) * len(dims)
            got = sum(1 for j in roles for fn in dims.values() if fn(j))
            rows.append({"Function": fnname, "Roles": len(roles),
                         "Completeness": f"{round(got/max(m,1)*100)}%"})
        st.dataframe(_pd.DataFrame(rows), use_container_width=True, hide_index=True)
