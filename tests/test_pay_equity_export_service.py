"""tests/test_pay_equity_export_service.py"""

from io import BytesIO

import pandas as pd
import pytest

from services.pay_equity_service import analyze_gender_pay_gap
from services.pay_equity_export_service import PayEquityExportService


def _grid(gap_factor: float, per_gender: int = 6) -> pd.DataFrame:
    """Same fixture shape as test_pay_equity_service._grid: a leveled grid where,
    at every Function x Level, women earn ``gap_factor`` x the men's salary."""
    funcs = ["B", "P"]
    rows = []
    eid = 1000
    for fi, fn in enumerate(funcs):
        for lv in range(1, 4):
            base = 30000 + 5000 * lv + fi * 3000
            for _ in range(per_gender):
                eid += 1
                rows.append({"EmployeeID": f"E{eid}", "Function": fn, "Level": str(lv),
                             "Gender": "M", "Salary": base})
                eid += 1
                rows.append({"EmployeeID": f"E{eid}", "Function": fn, "Level": str(lv),
                             "Gender": "F", "Salary": round(base * gap_factor)})
    return pd.DataFrame(rows)


def _analyze(df, **kw):
    return analyze_gender_pay_gap(
        df, function_col="Function", level_col="Level",
        gender_col="Gender", salary_col="Salary", **kw,
    )


def _sheets(data: bytes) -> dict:
    return pd.read_excel(BytesIO(data), sheet_name=None)


def test_workbook_has_expected_sheets():
    # "2.0" layout: Notes and Representation are folded onto Summary, not their own tabs.
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    assert set(_sheets(data)) == {"Summary", "Cohorts"}


def test_summary_sheet_carries_headline_numbers():
    # Women earn 0.90x men's pay -> PayGapResult.mean_gap_pct (men-paid-more
    # convention) is positive ~10%; the export flips sign to the wetsvoorstel's
    # (vrouw-man)/man convention, so the exported value should be negative.
    # Metrics live in columns B/C now (A is a blank left margin); pandas still
    # finds them by header name regardless, so this reads the same as before.
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    sm = _sheets(data)["Summary"]
    metrics = dict(zip(sm["Metric"], sm["Value"]))
    assert metrics["Men (M)"] == r.n_m
    assert metrics["Women (F)"] == r.n_f
    assert r.mean_gap_pct > 0  # sanity: source result uses the other sign
    assert metrics["Mean gap % — unadjusted (+ = women paid more, per wetsvoorstel (vrouw-man)/man)"] \
        == pytest.approx(-r.mean_gap_pct)
    assert metrics["Median gap % — unadjusted (+ = women paid more)"] == pytest.approx(-r.median_gap_pct)
    assert metrics["Adjusted gap % (controls for function + level; + = women paid more)"] \
        == pytest.approx(-r.adjusted_gap_pct)


def test_adjusted_ci_is_flipped_and_reordered():
    r = _analyze(_grid(0.90))
    assert r.adjusted_ci is not None and r.adjusted_ci[0] < r.adjusted_ci[1]
    data = PayEquityExportService().to_workbook_bytes(r)
    sm = _sheets(data)["Summary"]
    metrics = dict(zip(sm["Metric"], sm["Value"]))
    lo, hi = metrics["Adjusted 95% CI — low"], metrics["Adjusted 95% CI — high"]
    assert lo < hi
    assert lo == pytest.approx(-r.adjusted_ci[1])
    assert hi == pytest.approx(-r.adjusted_ci[0])


def test_low_n_cohort_count_in_summary():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    sm = _sheets(data)["Summary"]
    metrics = dict(zip(sm["Metric"], sm["Value"]))
    expected = sum(1 for c in r.cohorts if not c.reliable)
    assert metrics["Cohorts below the n>=5-per-gender reliability threshold (low-n, indicative only)"] == expected


def test_cohorts_sheet_matches_result_cohorts():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    tbl = _sheets(data)["Cohorts"]
    assert len(tbl) == len(r.cohorts) == r.n_cohorts_tested
    assert set(tbl["Function"]) == {"B", "P"}
    assert (tbl["Flagged (>= 5%)"] == "Yes").all()  # every cohort has a 10% gap in this fixture
    # every cohort in this fixture has men paid more -> exported (women-paid-more) sign is negative
    assert (tbl["Mean gap % (+ = women paid more)"] < 0).all()
    for c, exported in zip(r.cohorts, tbl["Mean gap % (+ = women paid more)"]):
        assert exported == pytest.approx(-c.mean_gap_pct)


def _find_header_row(ws, col: int, text: str) -> int:
    return next(row for row in range(1, ws.max_row + 1) if ws.cell(row, col).value == text)


def test_representation_tables_are_inline_on_summary():
    # By-level (I:J) and by-function (K:L) now sit side by side on Summary,
    # sharing the dashboard header row with Notes and the reliable-cohorts table.
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data))["Summary"]

    level_header_row = _find_header_row(ws, 9, "Level")
    assert ws.cell(level_header_row, 10).value == "% women"
    by_level = {ws.cell(level_header_row + 1 + i, 9).value: ws.cell(level_header_row + 1 + i, 10).value
                for i in range(len(r.women_by_level))}
    assert by_level == r.women_by_level

    function_header_row = _find_header_row(ws, 11, "Function")
    assert function_header_row == level_header_row  # same shared dashboard header row
    assert ws.cell(function_header_row, 12).value == "% women"
    by_function = {ws.cell(function_header_row + 1 + i, 11).value: ws.cell(function_header_row + 1 + i, 12).value
                   for i in range(len(r.women_by_function))}
    assert by_function == r.women_by_function


def test_notes_are_inline_on_summary():
    # Notes now live in column B, one per row, below the shared dashboard header.
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data))["Summary"]

    header_row = _find_header_row(ws, 2, "Notes")
    expected = PayEquityExportService().notes_list(r)
    actual = [ws.cell(header_row + 1 + i, 2).value for i in range(len(expected))]
    assert actual == expected
    assert "wetsvoorstel" in actual[0] and "vrouw" in actual[0]


def test_export_handles_a_result_with_no_reliable_cohorts():
    # Every cohort here is a single M/F pair -- small-n, but still exportable.
    df = pd.DataFrame([
        {"Function": "B", "Level": "1", "Gender": "M", "Salary": 40000},
        {"Function": "B", "Level": "1", "Gender": "F", "Salary": 30000},
    ])
    r = _analyze(df)
    data = PayEquityExportService().to_workbook_bytes(r)
    sheets = _sheets(data)
    assert "Cohorts" in sheets
    assert len(sheets["Cohorts"]) == 1


def test_summary_sheet_has_a_reliable_cohorts_chart():
    # _grid(0.90) with the default per_gender=6 makes all 6 B/P x 1-3 cohorts reliable.
    r = _analyze(_grid(0.90))
    assert all(c.reliable for c in r.cohorts) and len(r.cohorts) == 6
    data = PayEquityExportService().to_workbook_bytes(r)
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data))
    ws = wb["Summary"]
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    assert len(chart.series) == 2  # Mean M, Mean F


def test_chart_shows_data_labels_bottom_legend_and_is_sized_to_not_overlap():
    # Matches the client's own template: value labels on each bar, legend
    # below the plot, and a size (~18 x 10.3cm) small enough that, anchored
    # near the top of the sheet, it doesn't run down into the dashboard
    # section (Notes / mini-table / representation) that starts a good deal
    # further down.
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data))["Summary"]
    chart = ws._charts[0]
    dl = chart.dataLabels
    assert dl is not None and dl.showVal is True
    # Every other show* flag must be explicitly False, not just left as the
    # class default (None) -- Excel renders an unset flag here as "on", so a
    # None instead of False produces "series name, category, value" all
    # concatenated onto one oversized, overlapping label per bar. This is
    # exactly the bug a previous version of this chart shipped with.
    assert dl.showSerName is False
    assert dl.showCatName is False
    assert dl.showLegendKey is False
    assert dl.showPercent is False
    assert dl.dLblPos == "outEnd"
    assert chart.legend.position == "b"
    # height/width aren't round-tripped onto the Chart object on load (they only
    # drive the anchor's extent at save time) -- read the actual saved size back
    # off the anchor itself, in EMU (914400 per inch, 360000 per cm).
    anchor = chart.anchor
    assert anchor._from.col == 3 and anchor._from.row == 1  # column D, row 2
    assert anchor.ext.cx / 360000 == pytest.approx(18, abs=0.1)
    assert anchor.ext.cy / 360000 == pytest.approx(10.3, abs=0.1)


def test_reliable_cohorts_minitable_matches_source_means():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data))["Summary"]
    header_row = _find_header_row(ws, 5, "Function x Level")
    assert ws.cell(header_row, 6).value == "Mean M"
    assert ws.cell(header_row, 7).value == "Mean F"
    rows = {
        ws.cell(r_, 5).value: (ws.cell(r_, 6).value, ws.cell(r_, 7).value)
        for r_ in range(header_row + 1, header_row + 1 + len(r.cohorts))
    }
    for c in r.cohorts:
        assert rows[f"{c.function}-{c.level}"] == (c.mean_m, c.mean_f)


def test_no_chart_when_no_cohort_is_reliable():
    df = pd.DataFrame([
        {"Function": "B", "Level": "1", "Gender": "M", "Salary": 40000},
        {"Function": "B", "Level": "1", "Gender": "F", "Salary": 30000},
    ])
    r = _analyze(df)
    assert not any(c.reliable for c in r.cohorts)
    data = PayEquityExportService().to_workbook_bytes(r)
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data))
    ws = wb["Summary"]
    assert len(ws._charts) == 0
    assert any("No cohort has a reliable" in str(cell.value)
               for row in ws.iter_rows() for cell in row if cell.value)


def test_brand_colours_are_applied():
    from openpyxl import load_workbook
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)
    wb = load_workbook(BytesIO(data))

    sm = wb["Summary"]
    assert sm["B1"].fill.fgColor.rgb[-6:] == "53037F"  # header = Jobsy deep purple

    # women earn 10% less in this fixture -> exported mean/median gap is negative
    # and >= the 5% threshold in magnitude, so it should render in the "danger" colour.
    mean_gap_row = next(
        row for row in range(2, sm.max_row + 1)
        if str(sm.cell(row, 2).value).startswith("Mean gap")
    )
    assert sm.cell(mean_gap_row, 3).font.color.rgb[-6:] == "FF5A7A"

    coh = wb["Cohorts"]
    gap_col = next(
        c for c in range(1, coh.max_column + 1)
        if str(coh.cell(1, c).value).startswith("Mean gap")
    )
    assert coh.cell(2, gap_col).font.color.rgb[-6:] == "FF5A7A"


def test_write_workbook_to_disk():
    import tempfile
    from pathlib import Path

    r = _analyze(_grid(0.90))
    target = Path(tempfile.gettempdir()) / "jobsy_test_pay_equity_out.xlsx"
    path = PayEquityExportService().write_workbook(r, target)
    assert path.exists() and path.stat().st_size > 0


def _grade_biased_grid(level_shift: float, per_gender: int = 10) -> pd.DataFrame:
    """Same fixture as test_pay_equity_service._grade_biased_grid: pay is fair
    for the level you're at, but women sit `level_shift` levels below an
    equivalent man in the same function -- the grading itself is skewed."""
    funcs = ["B", "P", "M", "S"]
    rows = []
    for fn in funcs:
        for i in range(per_gender):
            base_lv = 3 + (i % 4)
            rows.append({"Function": fn, "Level": str(base_lv), "Gender": "M",
                         "Salary": 30000 + 5000 * base_lv})
            f_lv = base_lv - level_shift
            rows.append({"Function": fn, "Level": str(f_lv), "Gender": "F",
                         "Salary": 30000 + 5000 * f_lv})
    return pd.DataFrame(rows)


def test_summary_sheet_carries_the_grade_assignment_gap():
    r = _analyze(_grade_biased_grid(1.5))
    assert r.grade_gap_significant is True
    data = PayEquityExportService().to_workbook_bytes(r)
    sm = _sheets(data)["Summary"]
    row = sm[sm["Metric"].astype(str).str.startswith("Grade-assignment gap, in LEVELS", na=False)]
    assert len(row) == 1
    # Sign-flipped for the wetsvoorstel convention, like every other gap figure in this export.
    from services.pay_equity_service import flip_gap_sign
    assert row.iloc[0]["Value"] == pytest.approx(flip_gap_sign(r.grade_gap_levels), abs=0.01)

    from openpyxl import load_workbook
    sm_ws = load_workbook(BytesIO(data))["Summary"]
    gg_row = next(
        row for row in range(1, sm_ws.max_row + 1)
        if str(sm_ws.cell(row, 2).value).startswith("Grade-assignment gap, in LEVELS")
    )
    # Significant -> coloured as danger, same red used for a flagged pct gap elsewhere in this sheet.
    assert sm_ws.cell(gg_row, 3).font.color.rgb[-6:] == "FF5A7A"


# ── Dutch-language export ────────────────────────────────────────────────────

def test_dutch_workbook_has_dutch_sheet_names():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r, lang="nl")
    assert set(_sheets(data)) == {"Samenvatting", "Cohorten"}


def test_dutch_summary_uses_dutch_column_and_metric_labels():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService(lang="nl").to_workbook_bytes(r)
    sm = _sheets(data)["Samenvatting"]
    assert "Metriek" in sm.columns and "Waarde" in sm.columns
    labels = [l for l in sm["Metriek"].tolist() if isinstance(l, str)]
    assert any(l.startswith("Gemiddelde loonkloof %") for l in labels)
    assert any(l.startswith("Gecorrigeerde loonkloof %") for l in labels)
    assert any(l == "Medewerkers in scope" for l in labels)
    # No English metric labels should leak into the Dutch report.
    assert not any(l.startswith("Employees in scope") for l in labels)


def test_dutch_cohorts_sheet_uses_dutch_column_headers():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r, lang="nl")
    coh = _sheets(data)["Cohorten"]
    assert "Functie" in coh.columns and "Niveau" in coh.columns
    assert any(c.startswith("Gemarkeerd") for c in coh.columns)
    assert any(c.startswith("Gemiddelde kloof %") for c in coh.columns)


def test_dutch_notes_are_translated_with_numbers_preserved():
    df = _grid(0.90)
    # Force a real exclusion so the dynamic EXCLUSIONS note (with embedded
    # counts) gets generated and must survive translation with the SAME numbers.
    df.loc[df.index[0], "Salary"] = 0
    r = _analyze(df)
    assert any(n.startswith("EXCLUSIONS:") for n in r.notes)
    data = PayEquityExportService().to_workbook_bytes(r, lang="nl")
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data))["Samenvatting"]
    all_text = " ".join(str(ws.cell(row, 2).value) for row in range(1, ws.max_row + 1)
                        if ws.cell(row, 2).value)
    assert "UITSLUITINGEN:" in all_text
    assert "1 van de" in all_text  # the dropped-row count carried through untranslated as a number
    assert "Tekenconventie" in all_text
    assert "EXCLUSIONS:" not in all_text  # English original must not also appear


def test_dutch_gap_values_still_colour_flagged_red():
    # Same grid as the brand-colours English test, just requested in Dutch --
    # the colour logic must recognise the Dutch label wording too.
    r = _analyze(_grid(0.80))  # 20% gap -> flagged
    data = PayEquityExportService().to_workbook_bytes(r, lang="nl")
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data))["Samenvatting"]
    row = next(row for row in range(1, ws.max_row + 1)
              if str(ws.cell(row, 2).value).startswith("Gemiddelde loonkloof %"))
    assert ws.cell(row, 3).font.color.rgb[-6:] == "FF5A7A"


def test_english_report_unaffected_by_nl_translation_additions():
    r = _analyze(_grid(0.90))
    data = PayEquityExportService().to_workbook_bytes(r)  # default lang="en"
    assert set(_sheets(data)) == {"Summary", "Cohorts"}
    sm = _sheets(data)["Summary"]
    assert "Metric" in sm.columns and "Value" in sm.columns


def test_dutch_translation_of_extra_controls_note():
    from services.pay_equity_export_service import _translate_note
    note = "Adjusted gap also controls for tenure and age (in addition to function and level) — not hours, performance or location. A residual gap is a prompt to investigate, not proof of an unjustified gap."
    nl = _translate_note(note)
    assert "dienstjaren" in nl and "leeftijd" in nl
    assert "Adjusted gap" not in nl  # fully translated, not a fallback
