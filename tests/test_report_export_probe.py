"""tests/test_report_export_probe.py

Adversarial probe of the reporting / export / outbound-connector area:

    services/architecture_report_service.py  (the board-ready workbook)
    services/export_service.py               (matching-run export)
    services/workday_connector.py            (external HR system pull)

This file is written fresh against the current code to find real defects, not
to restate what the code does. Every test below either:

  * FAILS today and documents a concrete defect (file/line, input, actual vs.
    expected content), or
  * PASSES and is contributed as a regression guard for behaviour that was
    checked adversarially and held up.

No network call is ever made: WorkdayConnector's `requests.Session` and the
module-level `requests.post` used by `_authenticate` are always replaced with
in-memory fakes.
"""
from __future__ import annotations

import io
import pathlib

import openpyxl
import pandas as pd
import pytest

from core.models import Job, SalaryBand
from services.architecture_report_service import ArchitectureReportService
from services.matching_service import MatchResult, MatchType
from services.workday_connector import WorkdayConnector

ROOT = pathlib.Path(__file__).resolve().parents[1]


# ─────────────────────────────────────────────────────────────────────────
# Minimal fakes for ArchitectureReportService's dependencies.
#
# tests/conftest.py already has a FakeCatalog, but it is built for
# MatchingService and has neither `get_role_skills` / `skill_gap` (needed by
# sheet 6, "Succession Risk") nor `path` (needed by sheets 8-9, which read an
# on-disk reference workbook). Redefined locally per the ground rule that
# forbids editing existing files.
# ─────────────────────────────────────────────────────────────────────────
class _FakeRepo:
    def __init__(self):
        self.jobs = {}
        self.salary = {}
        self.job_grades = {}
        self.career_paths = {}
        self.industries = {}
        self.benefits_catalog = {}


class _FakeCatalog:
    def __init__(self, repo, path="/nonexistent/reference_library.xlsx"):
        self.repository = repo
        self.path = pathlib.Path(path)

    def get_role_skills(self, job_id):
        return []

    def skill_gap(self, current_skills, target_job_id):
        return []


def _one_job_service(currency="€", salary=0, function="Engineering",
                      level="Medior", input_title="Developer", first_name="Jan"):
    """One matched employee against one job with a salary band -- the smallest
    non-degenerate case, with every knob a test needs exposed."""
    repo = _FakeRepo()
    job = Job(job_id="J1", standard_title="Software Engineer", function=function,
              level=level, grade=5)
    repo.jobs["J1"] = job
    repo.salary[(function, level)] = SalaryBand(
        function=function, level=level, min=50000, max=70000,
        p25=55000, p50=60000, p75=65000,
    )
    result = MatchResult(
        input_title=input_title, match_type=MatchType.EXACT, confidence=100,
        requires_review=False, matched=True, job_id="J1",
        standard_title="Software Engineer", function=function, level=level,
    )
    catalog = _FakeCatalog(repo)
    df_emp = pd.DataFrame([{"FirstName": first_name, "LastName": "Kowalski", "Salary": salary}])
    svc = ArchitectureReportService(catalog, [result], df_employees=df_emp,
                                     org_label="Test Co", currency=currency)
    return svc


def _load(data: bytes) -> "openpyxl.Workbook":
    return openpyxl.load_workbook(io.BytesIO(data))


# ─────────────────────────────────────────────────────────────────────────
# 1. Currency conversion: symbol placement
# ─────────────────────────────────────────────────────────────────────────
def test_word_suffix_currencies_are_wrongly_prepended_in_job_architecture_sheet():
    """services/country_service.py's `money()` -- the convention this report
    is supposed to follow -- puts a word-shaped symbol (zl/kr/Kc) AFTER the
    number ("90.000 zl") and only a glyph symbol (euro/pound) BEFORE it. Every
    money cell in architecture_report_service.py, however, does

        f"{self.currency}{val:,}"

    unconditionally (e.g. line 261 in sheet 2, line 330-332 in sheet 3, line
    374-376 in sheet 4, line 641/667-669 in sheets 8-9, line 772-773 in sheet
    10). For a Polish or Swedish client -- who reaches this code through
    ui/app.py's `_cur()`, i.e. `country_service.symbol_for()`, so this is a
    live production path, not a hypothetical one -- every currency figure in
    the board workbook is glued together wrong: "zl50.000" instead of
    "50.000 zl".
    """
    svc = _one_job_service(currency="zł")  # PLN symbol, as symbol_for("PL") returns
    wb = _load(svc.generate())
    ws = wb["2. Job Architecture"]
    min_cell = next(c for row in ws.iter_rows(min_row=2) for c in row
                     if isinstance(c.value, str) and c.value.startswith(("zł", "5")))
    assert min_cell.value is not None, "could not find the Min-salary cell to check"
    value = min_cell.value
    assert not value.startswith("zł"), (
        f"cell {min_cell.coordinate} = {value!r} prepends the PLN symbol; "
        "country_service.money() puts zł AFTER the number with a space, "
        "e.g. '50.000 zł' -- this workbook does not follow its own currency "
        "convention for any suffix currency (PLN/SEK/DKK/NOK/CZK)"
    )


def test_word_suffix_currencies_are_wrongly_prepended_in_org_snapshot_sheet():
    """Same defect, second sheet: sheet 3 "Org Snapshot" is the per-employee
    table -- the one a works council reads line by line. Lines 330-332 build
    Salary Band Min / P50 Market / Salary the same broken way.
    """
    svc = _one_job_service(currency="kr", salary=60000)  # SEK/DKK/NOK symbol
    wb = _load(svc.generate())
    ws = wb["3. Org Snapshot"]
    header = [c.value for c in ws[1]]
    salary_col = header.index("Salary") + 1
    salary_cell = ws.cell(2, salary_col)
    assert salary_cell.value == "60.000 kr", (
        f"expected the Swedish/Danish/Norwegian convention '60.000 kr', got "
        f"{salary_cell.value!r} -- the report prepends 'kr' to every figure "
        "regardless of currency, in every sheet, not just this one"
    )


# ─────────────────────────────────────────────────────────────────────────
# 2. Formula / Excel injection via untrusted uploaded data
# ─────────────────────────────────────────────────────────────────────────
def test_employee_name_starting_with_equals_becomes_a_live_excel_formula():
    """A works-council-bound workbook must never let a person's name or job
    title become executable content. openpyxl auto-detects a leading '=' on
    any string it is given and stores the cell as a FORMULA (data_type 'f'),
    not literal text -- confirmed directly:

        >>> Workbook().active.cell(1, 1, "=1+1").data_type
        'f'

    `_cell()` (architecture_report_service.py line 50-55) does exactly
    `ws.cell(row, col, text)` with no leading-quote escaping and no rejection
    of formula-triggering prefixes, for values that come straight from an
    uploaded employee spreadsheet (FirstName/LastName -> sheet 3 column A,
    and the matched input_title -> sheet 3 column B). Anyone who can get a
    name or an old job title beginning with '=' into the uploaded roster (a
    contractor's self-reported title, a copy-pasted CSV, a disgruntled
    employee who knows this) gets a live formula -- e.g. a HYPERLINK to an
    external URL, or worse -- that runs when a works council member or board
    member opens the file Jobsy generated for them.
    """
    svc = _one_job_service(first_name='=HYPERLINK("http://evil.example","click")',
                            input_title="=2+5")
    wb = _load(svc.generate())
    ws = wb["3. Org Snapshot"]
    name_cell = ws.cell(2, 1)
    title_cell = ws.cell(2, 2)
    assert name_cell.data_type != "f", (
        f"employee name cell {name_cell.coordinate} was stored as a live "
        f"formula ({name_cell.value!r}) instead of literal text -- Excel "
        "will attempt to evaluate this when the file is opened"
    )
    assert title_cell.data_type != "f", (
        f"input-title cell {title_cell.coordinate} was stored as a live "
        f"formula ({title_cell.value!r}) instead of literal text"
    )


# ─────────────────────────────────────────────────────────────────────────
# 3. Zero pay vs. unknown pay
# ─────────────────────────────────────────────────────────────────────────
def test_a_genuine_zero_salary_is_indistinguishable_from_a_missing_one():
    """The flip side of the currency work's own stated principle ("a missing
    salary renders as an em dash rather than '0', because zero and unknown
    are different facts", country_service.py line 143-144): here a REAL,
    recorded salary of exactly 0 (e.g. an unpaid intern/volunteer role,
    which does exist in some client rosters) is thrown away and rendered
    identically to a missing salary.

    `_build_org_snapshot`, line 332:
        (f"{self.currency}{int(salary):,}".replace(",",".") if salary else "-", ...)

    `if salary` treats 0.0 as falsy, so a genuine zero collapses to "-" —
    the same symbol used for "we don't know". A works council reading this
    sheet cannot tell "this person is unpaid" from "we have no salary data
    for this person", which is exactly the ambiguity the currency-display
    module says it exists to avoid.
    """
    svc = _one_job_service(salary=0)
    wb = _load(svc.generate())
    ws = wb["3. Org Snapshot"]
    header = [c.value for c in ws[1]]
    salary_col = header.index("Salary") + 1
    cell_value = ws.cell(2, salary_col).value
    assert cell_value != "—", (
        "a recorded salary of exactly 0 renders as an em dash, identical to "
        "a missing salary -- the report cannot tell 'unpaid' from 'unknown'"
    )


# ─────────────────────────────────────────────────────────────────────────
# 4. Robustness: minimal / empty data must not crash or fabricate numbers
# ─────────────────────────────────────────────────────────────────────────
def test_completely_empty_dataset_does_not_crash():
    """No employees, no matched results, no employee dataframe at all. This
    is the state of a brand-new org before any matching run. The report must
    degrade gracefully (informative placeholder text is fine) rather than
    raising -- verified to hold today; contributed as a regression guard."""
    repo = _FakeRepo()
    catalog = _FakeCatalog(repo)
    svc = ArchitectureReportService(catalog, [], df_employees=None,
                                     org_label="Empty Co", currency="€")
    data = svc.generate()
    wb = _load(data)
    assert len(wb.sheetnames) == 11


@pytest.mark.parametrize("bad_currency", [None, ""])
def test_falsy_currency_falls_back_to_euro_default_rather_than_crashing(bad_currency):
    """__init__ line 90: `self.currency = currency or "\\u20ac"`. Confirms a
    caller that passes None or "" (a plausible bug on the caller's side, e.g.
    an org with no default_country configured yet) gets the deployment
    default rather than a broken/empty-symbol workbook. Regression guard."""
    repo = _FakeRepo()
    catalog = _FakeCatalog(repo)
    svc = ArchitectureReportService(catalog, [], df_employees=None,
                                     org_label="X", currency=bad_currency)
    assert svc.currency == "€"


def test_sheet_names_are_structurally_valid_excel_names():
    """Excel silently corrupts a workbook whose sheet name exceeds 31 chars
    or contains []:*?/\\, or whose name collides with another sheet. All 11
    sheet names are static string literals here, so this is a cheap
    structural regression guard against a future edit accidentally breaking
    one (e.g. lengthening a title past 31 chars)."""
    repo = _FakeRepo()
    catalog = _FakeCatalog(repo)
    svc = ArchitectureReportService(catalog, [], df_employees=None,
                                     org_label="X", currency="€")
    wb = _load(svc.generate())
    forbidden = set('[]:*?/\\')
    seen = set()
    for name in wb.sheetnames:
        assert len(name) <= 31, f"{name!r} is {len(name)} chars, Excel's limit is 31"
        assert not (forbidden & set(name)), f"{name!r} contains a forbidden Excel character"
        assert name not in seen, f"duplicate sheet name {name!r}"
        seen.add(name)


# ─────────────────────────────────────────────────────────────────────────
# 5. workday_connector.py — external system, stubbed transport only
# ─────────────────────────────────────────────────────────────────────────
class _FakeWDResponse:
    def __init__(self, payload=None, status_code=200, text=""):
        self._payload = payload or {}
        self.status_code = status_code
        self.text = text

    def raise_for_status(self):
        if self.status_code >= 400:
            import requests
            raise requests.HTTPError(f"{self.status_code} error for url: (stubbed)")

    def json(self):
        return self._payload


class _FakeWDSession:
    """Stand-in for `requests.Session()` -- no network I/O, ever."""

    def __init__(self, pages_by_offset):
        self.pages = pages_by_offset
        self.calls: list[dict] = []
        self.headers = {}

    def get(self, url, params=None, timeout=None):
        self.calls.append(dict(params or {}))
        offset = (params or {}).get("offset")
        payload = self.pages.get(offset, {"data": []})
        return _FakeWDResponse(payload)


def _connector_with_fake_token():
    conn = WorkdayConnector("acme", "cid", "shh-client-secret", "shh-refresh-token")
    conn._access_token = "already-authenticated"  # skip the real /token POST entirely
    return conn


def test_fetch_workers_silently_truncates_when_total_field_is_absent():
    """services/workday_connector.py, `fetch_workers` (line ~135-150):

        while True:
            ...
            all_workers.extend(batch)
            total = data.get("total", 0)          # <-- defaults to 0
            cur_offset += limit
            if cur_offset >= total: break

    If a page response does not carry a "total" key -- the code has no
    fallback for that beyond silently assuming zero -- `cur_offset` (already
    >= 0) trips the exit condition after the very FIRST page, even though a
    second page of real workers is sitting right there waiting to be
    fetched. No exception, no warning, no partial-fetch indicator: the
    workforce dataset that later feeds the architecture report is just
    quietly smaller than the client's real headcount.

    Two pages of 2 workers each are stubbed (4 total); only the first page's
    HTTP call is ever made.
    """
    conn = _connector_with_fake_token()
    conn.session = _FakeWDSession({
        0: {"data": [{"id": "1"}, {"id": "2"}]},   # no "total" key at all
        2: {"data": [{"id": "3"}, {"id": "4"}]},
    })
    df = conn.fetch_workers(limit=2)
    # Both stubbed pages are FULL (2 of a limit of 2), and the tenant never says
    # how many workers there are. A client that stops after the last full page
    # cannot know it was the last one -- so one further call, which comes back
    # empty, is how the end is discovered. Three calls for two pages is the
    # correct shape here, not an off-by-one; what matters is that no worker is
    # left behind, which the next assertion is for.
    assert len(conn.session.calls) >= 2, (
        f"pagination stopped after {len(conn.session.calls)} HTTP call(s) "
        f"({conn.session.calls}) instead of continuing to the second page "
        "that was waiting to be fetched"
    )
    assert len(df) == 4, (
        f"expected all 4 stubbed workers to be fetched, got {len(df)} -- "
        "the second page of real workers was silently dropped because the "
        "first page's response had no 'total' key"
    )


def test_test_connection_leaks_raw_response_body_into_the_returned_message():
    """`test_connection()` (line 105-115):

        return False, f"HTTP {r.status_code}: {r.text[:200]}"

    passes the raw HTTP response body straight through, unsanitised, as the
    message a caller shows the user. ui/app.py does exactly that:
    `st.error(f"Failed: {msg}")` (ui/app.py line 1242) -- so whatever Workday
    puts in an error body for the `/workers?limit=1` probe call lands
    verbatim in a Streamlit error banner. Client secret / refresh token are
    NOT leaked this way (confirmed: they never appear in the returned
    message), but this is still a raw, unfiltered echo of a third-party
    response body into a user-facing UI surface -- exactly the kind of
    channel that leaks whatever Workday chooses to put in an error payload
    (which, for some tenant configurations, includes echoed request/record
    context).
    """
    conn = _connector_with_fake_token()
    leaky_body = ('{"error":"invalid_filter","instances":[{"employeeID":"E-4471",'
                  '"legalName":"A. Confidential Employee","annualSalary":123456}]}')
    conn.session = _FakeWDSession({})
    conn.session.get = lambda url, params=None, timeout=None: _FakeWDResponse(
        status_code=500, text=leaky_body)
    ok, msg = conn.test_connection()
    assert ok is False
    assert "E-4471" not in msg and "Confidential Employee" not in msg, (
        f"the raw Workday response body was echoed verbatim into the "
        f"connection-test message shown to the user: {msg!r}"
    )
    # Credentials themselves are not part of the leak -- documented as a
    # regression guard alongside the failing assertion above.
    assert "shh-client-secret" not in msg and "shh-refresh-token" not in msg


def test_authenticate_never_puts_credentials_in_the_session_headers_value_alone():
    """Regression guard: the Authorization header must carry the ACCESS
    token (obtained from Workday), never the client secret or refresh token
    directly -- confirms `_authenticate` doesn't accidentally use the wrong
    variable when building the Bearer header."""
    import requests as _requests

    class _FakeTokenResponse:
        status_code = 200

        def raise_for_status(self):
            pass

        def json(self):
            return {"access_token": "tok-xyz"}

    def _fake_post(url, data=None, timeout=None):
        return _FakeTokenResponse()

    conn = WorkdayConnector("acme", "cid", "shh-client-secret", "shh-refresh-token")
    real_post = _requests.post
    _requests.post = _fake_post
    try:
        token = conn._authenticate()
    finally:
        _requests.post = real_post
    assert token == "tok-xyz"
    assert conn.session.headers["Authorization"] == "Bearer tok-xyz"
    assert "shh-client-secret" not in conn.session.headers["Authorization"]
