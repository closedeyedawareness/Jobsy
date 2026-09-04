"""
The cover page, and the one thing it must never do: disagree with its own report.

It is built last and moved to the front precisely so it can quote figures the
other sheets already computed. If it recomputed them it would eventually give a
different answer from the sheet it summarises, and the summary is the page that
gets read.
"""

import io
import random

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.catalog import Catalog
from services.architecture_report_service import ArchitectureReportService
from services.matching_service import MatchingService

WORKBOOK = "jobsy_reference_library.xlsx"

# A COMMITTED fixture, not the local demo file. These tests read
# jobsy_demo_100_people.csv, which .gitignore excludes (*people*.csv), so they
# passed on the machine that wrote them and raised FileNotFoundError everywhere
# else -- including CI, where a missing file in a FIXTURE is reported as an
# ERROR rather than a FAILURE and is easy to read past under a cheerful "444
# passed". main has been red since 2026-09-03 for exactly this.
#
# fixtures/synthetic-100-people.csv is generated from this workbook's own 81
# standard titles, so matching exercises the real path rather than the
# not-found path, and it carries part-timers because the Overview's
# "compared full-time-equivalent" line has nothing to say without them.
# .gitignore already carved out !fixtures/synthetic-*.csv for this purpose.
# No real people are in it.
ROSTER = "fixtures/synthetic-100-people.csv"


@pytest.fixture(scope="module")
def report():
    cat = Catalog(WORKBOOK, source="excel").load()
    svc = MatchingService(cat, index=cat.repository.index)
    emp = pd.read_csv(ROSTER)
    results = [svc.match(str(t)) for t in emp["CurrentTitle"]]
    random.seed(7)
    emp = emp.copy()
    emp["ActualSalary"] = [
        int((cat.repository.salary.get((r.function, r.level)).p50
             if (r.matched and cat.repository.salary.get((r.function, r.level))) else 50000)
            * random.uniform(0.78, 1.12))
        for r in results]
    svc_rpt = ArchitectureReportService(cat, results, df_employees=emp, org_label="Northwind BV")
    data = svc_rpt.generate()
    return svc_rpt, load_workbook(io.BytesIO(data))


def test_the_overview_is_the_first_sheet(report):
    _, wb = report
    assert wb.sheetnames[0] == "1. Overview"


def test_the_sheets_are_numbered_without_a_collision(report):
    _, wb = report
    numbers = [n.split(".")[0] for n in wb.sheetnames]
    assert len(numbers) == len(set(numbers)), wb.sheetnames


def test_the_cover_quotes_the_report_rather_than_recomputing_it(report):
    """Every headline figure has to be one another sheet already found."""
    rpt, wb = report
    ws = wb["1. Overview"]
    row7 = [ws.cell(7, c).value for c in range(2, 7)]
    assert str(rpt._facts["below"]) in row7
    assert str(rpt._facts["avg_compa"]) in row7
    assert str(len(rpt.results)) in row7


def test_the_money_leads_because_a_board_reads_it_first(report):
    rpt, wb = report
    ws = wb["1. Overview"]
    assert "€" in str(ws.cell(7, 2).value)
    assert "band minimum" in str(ws.cell(8, 2).value)


def test_completeness_is_a_confidence_line_not_a_headline(report):
    """A CHRO does not open a board pack with '83% of roles have a profile'."""
    _, wb = report
    ws = wb["1. Overview"]
    basis = str(ws.cell(10, 2).value)
    assert basis.startswith("Basis:")
    assert "absent from every figure above" in basis


def test_the_report_and_the_page_price_part_time_pay_the_same_way(report):
    """The pay sheet had its own compa-ratio -- actual/p50, no FTE -- so after
    the app started pro-rating, the workbook and the screen disagreed about the
    same person. The demo file carries an FTE column; some of it must be used."""
    rpt, _ = report
    assert rpt._facts["pro_rated"] > 0


def test_the_palette_comes_from_the_app_theme(report):
    from services import architecture_report_service as m
    from ui.theme import COLORS
    assert m.INK == COLORS["on_light_ink"].lstrip("#").upper()
    assert m.BLUE == COLORS["fill_accent"].lstrip("#").upper()
