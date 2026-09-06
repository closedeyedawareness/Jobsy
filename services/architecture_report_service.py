"""
architecture_report_service.py — Generate a board-ready Job Architecture Framework report.

Takes matched results + session state and produces a fully formatted Excel workbook
with 7 analytical sections and narrative recommendations driven by pattern detection.
"""
from __future__ import annotations

import io
import math
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                               PatternFill, Side)
from openpyxl.utils import get_column_letter

from services import salary_service as _salary

# ── Design tokens, read from the app's own theme ──────────────────────────
#
# This file used to carry its OWN palette -- teal #0E7C66, slate #17212E -- and
# when Jobsy moved onto the People Harmonics system in aead07e the report was
# left behind. A client got a violet product and a teal workbook. The names
# below are unchanged so the eleven sheets need no edits; only where they point
# has changed, and they now point at ui/theme.py, which is the thing that moves.
#
# The workbook stays LIGHT deliberately. A dark Excel sheet is unreadable when
# printed and wrong for a board pack, so this uses PH's own light palette --
# the on_light_* tokens taken from corporate.css, the same white-and-blue
# system as the investor sheet -- rather than inventing a light theme.
try:
    from ui.theme import COLORS as _PH
except Exception:                                     # pragma: no cover
    _PH = {}


def _hx(key: str, fallback: str) -> str:
    """A theme colour as openpyxl wants it: six hex digits, no leading hash."""
    return str(_PH.get(key) or fallback).lstrip("#").upper()


BLUE   = _hx("fill_accent", "#1B4DD8");      BLUE_L  = _hx("on_light_tint", "#F4F7FE")
TEAL   = _hx("fill_accent_deep", "#0B2A5B"); TEAL_L  = BLUE_L
INK    = _hx("on_light_ink", "#0B1729")
MUTED  = _hx("on_light_muted", "#6B7A8F")
LINE   = _hx("on_light_line", "#E3E8EF")
GREY   = "F7F9FC"; WHITE = "FFFFFF"

# Categorical hues: the PH brand accents, darkened where they must carry white
# text or sit on paper. Kept as literals with the token they descend from, since
# the dark-ground originals fail contrast on white.
VIOLET = "6B2FD6"; VIOLET_L = "F0EAFF"    # <- primary / --ac  #8850EF
AMBER  = "9A6B14"; AMBER_L  = "FBF3E2"    # <- gold           #E6B25E
RED    = "B03A2E"; RED_L    = "FBE9E7"    # <- danger         #FF5A7A
GREEN  = "17916B"; GREEN_L  = "E4F5EF"    # <- success        #6EE7B7

BAND_COLORS = {
    "Support":         AMBER,
    "Professional":    BLUE,
    "Expert / Senior": TEAL,
    "Leadership":      VIOLET,
    "Executive":       INK,
}
BAND_LIGHT = {
    "Support": AMBER_L, "Professional": BLUE_L,
    "Expert / Senior": TEAL_L, "Leadership": VIOLET_L, "Executive": GREY,
}

def _hdr(ws, row, col, text, bg=TEAL, size=10, align="left"):
    c = ws.cell(row, col, text)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c

def _defuse(text):
    """Stop a spreadsheet from executing a client's own data.

    Excel treats a cell whose text begins with "=" as a formula, and openpyxl
    writes it as one. Most of what lands in this workbook is uploaded by the
    client -- employee names, job titles typed by whoever filled in the
    spreadsheet -- so a name of `=HYPERLINK("http://...","click")` becomes a
    live formula in a file that a works council or a board opens. Prefixing an
    apostrophe is Excel's own "treat this as text" marker and does not show in
    the cell. +, - and @ are included because the same string is dangerous again
    the moment anyone re-exports this sheet as CSV.
    """
    if isinstance(text, str) and text[:1] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def _cell(ws, row, col, text, fg=INK, bg=WHITE, bold=False, size=9, align="left", italic=False):
    c = ws.cell(row, col, _defuse(text))
    c.font      = Font(name="Arial", color=fg, bold=bold, size=size, italic=italic)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c

def _border_range(ws, r1, r2, c1, c2, color=LINE):
    s = Side(style="thin", color=color)
    b = Border(left=s, right=s, top=s, bottom=s)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = b

def _row_bg(ri):
    return GREY if ri % 2 == 0 else WHITE


def _grade(obj):
    """Grade of a Job, resilient to stale/old cached objects."""
    return getattr(obj, "grade", 0) or 0


class ArchitectureReportService:
    """Generate a job architecture framework report from Jobsy session data."""

    GRADE_BANDS = {
        1:"Support",2:"Support",3:"Support",
        4:"Professional",5:"Professional",6:"Professional",
        7:"Expert / Senior",8:"Expert / Senior",9:"Expert / Senior",10:"Expert / Senior",
        11:"Leadership",12:"Leadership",13:"Leadership",
        14:"Executive",
    }

    def __init__(self, catalog, results: list, df_employees: Optional[pd.DataFrame] = None,
                 org_label: str = "", currency: str = "\u20ac",
                 country: str = ""):
        # The symbol is passed in rather than read from a session, because a
        # service that builds a workbook must work from a script and a cron job
        # as well as from a browser. The caller knows the client's market; this
        # only has to not assume euro. See services/country_service.py.
        self.currency     = currency or "\u20ac"
        # Which market's benefit observations this report may draw on. Empty
        # means the deployment default; BenefitsService resolves it.
        self.country      = (country or "").strip().upper()
        self.catalog      = catalog
        self.results      = [r for r in results if r.matched]
        self.df_employees = df_employees
        self.org_label    = org_label or "Organisation"
        self.today        = date.today().strftime("%d %B %Y")
        self._wb          = Workbook()
        self._findings: list[dict] = []
        self._recommendations: list[dict] = []
        # What each sheet learned, for the overview to quote. The cover is built
        # LAST and moved to the front, so it reports what the report actually
        # found rather than recomputing it and risking a different answer.
        self._facts: dict = {}

    # ── Public entry point ────────────────────────────────────────────────

    def _money(self, value, blank: str = "\u2014") -> str:
        """One amount, formatted the way this client's market writes money.

        Every money cell used to be `f"{self.currency}{v:,}"` -- the symbol
        always in front. That is right for the euro and the pound and wrong for
        every other currency Jobsy covers: zloty, krona, krone and koruna are
        written after the number. Passing `symbol_for()` into the old code
        produced "zl50.000" and "kr60.000" in a workbook a client forwards to
        their works council. Placement follows country_service.money(), which is
        where the convention is defined; this cannot call it because the report
        must build from a script with no session, so it takes the symbol and
        applies the same rule to it.
        """
        try:
            n = float(value)
        except (TypeError, ValueError):
            return blank
        if n != n or n in (float("inf"), float("-inf")):
            return blank
        formatted = f"{n:,.0f}".replace(",", ".")
        sym = self.currency
        return f"{sym}{formatted}" if sym in ("\u20ac", "\u00a3") else f"{formatted} {sym}"

    def generate(self) -> bytes:
        self._wb.remove(self._wb.active)
        self._build_executive_summary()
        self._build_job_architecture()
        self._build_org_snapshot()
        self._build_grade_distribution()
        self._build_career_paths()
        self._build_succession_risk()
        self._build_recommendations()
        self._build_job_family_leveling()
        self._build_total_pay()
        self._build_pay_equity()
        self._build_benefits_benchmarking()
        self._build_overview()
        self._wb.move_sheet(self._wb["1. Overview"], -(len(self._wb.sheetnames) - 1))
        buf = io.BytesIO()
        self._wb.save(buf)
        return buf.getvalue()

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    # ── Cover: the page a CHRO puts in front of a board ───────────────
    def _build_overview(self):
        """One page, written for the room it is read in.

        A CHRO does not open a board pack with "83% of roles have a profile".
        They open with what it costs, what the risk is, and what needs deciding.
        So completeness moved off the headline and became the CONFIDENCE line
        underneath: it qualifies the numbers rather than being one of them.

        Built last and moved to the front, so every figure here is one another
        sheet already reported. Recomputing them on the cover is how a summary
        starts quietly disagreeing with its own report.
        """
        f = self._facts
        ws = self._wb.create_sheet("1. Overview")
        ws.sheet_view.showGridLines = False
        for ci, w in enumerate([2, 27, 21, 21, 21, 21, 30], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # masthead
        ws.row_dimensions[2].height = 46
        ws.merge_cells("B2:F2")
        c = ws.cell(2, 2, self.org_label)
        c.font = Font(name="Calibri", bold=True, size=24, color=WHITE)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[3].height = 24
        ws.merge_cells("B3:F3")
        c = ws.cell(3, 2, "Job architecture & reward review  \u00b7  " + self.today)
        c.font = Font(name="Calibri", size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        score, score_col = self._compute_health_score()
        highs = [x for x in self._findings if x["severity"] == "high"]

        # the verdict, in a sentence
        ws.row_dimensions[5].height = 34
        ws.merge_cells("B5:F5")
        verdict = ("Architecture health " + str(score) + "/100 \u2014 "
                   + ("no high-severity findings."
                      if not highs else
                      str(len(highs)) + " finding(s) need a decision, below."))
        c = ws.cell(5, 2, verdict)
        c.font = Font(name="Calibri", bold=True, size=13, color=score_col)
        c.fill = PatternFill("solid", fgColor=BLUE_L)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # the numbers a board acts on: money and risk first
        ws.row_dimensions[7].height = 30
        ws.row_dimensions[8].height = 30

        def tile(col, value, label, colour):
            v = ws.cell(7, col, value)
            v.font = Font(name="Calibri", bold=True, size=17, color=colour)
            v.alignment = Alignment(horizontal="left", vertical="center")
            lab = ws.cell(8, col, label)
            lab.font = Font(name="Calibri", size=9, color=MUTED)
            lab.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for row in (7, 8):
                ws.cell(row, col).border = Border(top=Side("thin", color=LINE),
                                                  bottom=Side("thin", color=LINE))

        lift = f.get("lift_to_minimum")
        below = f.get("below")
        # self._money() rather than a euro literal: this tile is the first number
        # a board reads, and for a Polish or Swedish client a euro sign here is
        # not a formatting slip but a different amount. _money() already returns
        # the em dash for None, so the three cases collapse into one call.
        money = self._money(lift)
        tiles = [
            (money, "Annual cost to lift everyone to their band minimum",
             RED if lift else (GREEN if lift == 0 else INK)),
            (str(below) if below is not None else "\u2014", "People paid below their band",
             RED if below else (GREEN if below == 0 else INK)),
            (str(f.get("avg_compa") or "\u2014"), "Average compa-ratio (1.00 = at market)", INK),
            (str(len(self.results)), "Roles matched to the architecture", INK),
            (str(len(highs)), "High-severity findings", RED if highs else GREEN),
        ]
        for i, (val, lab, col) in enumerate(tiles):
            tile(2 + i, val, lab, col)

        # the confidence line: what the numbers rest on, and who is not in them
        ws.merge_cells("B10:F10")
        basis = []
        if f.get("priced") is not None:
            basis.append(str(f["priced"]) + " employees could be priced against a band")
        if f.get("pro_rated"):
            basis.append(str(f["pro_rated"]) + " part-time salaries compared full-time-equivalent")
        basis.append(str(len(self.results)) + " job titles matched a reference role")
        c = ws.cell(10, 2, "Basis: " + "; ".join(basis) + ". Anyone who could not be matched "
                    "or priced is absent from every figure above \u2014 the Pay Equity sheet "
                    "carries the reconciliation.")
        c.font = Font(name="Calibri", size=9, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[10].height = 30

        # what needs a decision
        r = 12
        c = ws.cell(r, 2, "What needs a decision")
        c.font = Font(name="Calibri", bold=True, size=12, color=INK)
        r += 1
        recs = self._recommendations[:3]
        if not recs:
            ws.cell(r, 2, "Nothing outstanding at this level.").font = Font(
                name="Calibri", size=10, color=MUTED)
            r += 1
        for i, rec in enumerate(recs, 1):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            c = ws.cell(r, 2, str(i) + ".  " + str(rec.get("title", "")))
            c.font = Font(name="Calibri", bold=True, size=10, color=INK)
            c.fill = PatternFill("solid", fgColor=GREY)
            c.alignment = Alignment(vertical="center", indent=1)
            r += 1
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            c = ws.cell(r, 2, str(rec.get("impact") or rec.get("action", "")))
            c.font = Font(name="Calibri", size=9, color=MUTED)
            c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.row_dimensions[r].height = 28
            r += 2

        # the shape of the organisation, as a real chart
        c = ws.cell(r, 2, "Where the roles sit")
        c.font = Font(name="Calibri", bold=True, size=12, color=INK)
        levels = {}
        for res in self.results:
            key = res.level or "\u2014"
            levels[key] = levels.get(key, 0) + 1
        order = [lv for lv in ("Junior", "Medior", "Senior", "Lead") if lv in levels]
        order += [lv for lv in levels if lv not in order]

        # the chart's data lives in two hidden columns, out of the reading area
        data_row = 40
        for i, lv in enumerate(order):
            ws.cell(data_row + i, 9, lv)
            ws.cell(data_row + i, 10, levels[lv])
        if order:
            ch = BarChart()
            ch.type = "col"
            ch.style = 2
            ch.legend = None
            ch.height, ch.width = 6.0, 13.5
            ch.add_data(Reference(ws, min_col=10, min_row=data_row,
                                  max_row=data_row + len(order) - 1))
            ch.set_categories(Reference(ws, min_col=9, min_row=data_row,
                                        max_row=data_row + len(order) - 1))
            ws.add_chart(ch, "B" + str(r + 1))
            ws.column_dimensions["I"].hidden = True
            ws.column_dimensions["J"].hidden = True

        # contents, beside the chart rather than under it
        cr = 12
        ws.cell(cr, 7, "Contents").font = Font(name="Calibri", bold=True, size=11, color=INK)
        cr += 1
        for name in self._wb.sheetnames:
            if name == "1. Overview":
                continue
            ws.cell(cr, 7, name).font = Font(name="Calibri", size=9.5, color=BLUE)
            cr += 1

    def _build_executive_summary(self):
        ws = self._wb.create_sheet("2. Executive Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 20

        # Title block
        ws.row_dimensions[1].height = 10
        ws.row_dimensions[2].height = 52
        ws.merge_cells("B2:D2")
        c = ws.cell(2, 2, f"Job Architecture Framework Report")
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=22)
        c.fill      = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[3].height = 26
        ws.merge_cells("B3:D3")
        c = ws.cell(3, 2, f"{self.org_label}  ·  Generated {self.today}")
        c.font      = Font(name="Arial", color=TEAL, size=12)
        c.fill      = PatternFill("solid", fgColor=TEAL_L)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[4].height = 10

        # Compute metrics
        total        = len(self.results)
        functions    = len({r.function for r in self.results})
        levels       = {r.level: 0 for r in self.results}
        for r in self.results:
            levels[r.level] = levels.get(r.level, 0) + 1
        grade_dist   = {}
        for r in self.results:
            job = self.catalog.repository.jobs.get(r.job_id)
            if job:
                g = _grade(job)
                grade_dist[g] = grade_dist.get(g, 0) + 1

        lead_count   = levels.get("Lead", 0)
        senior_count = levels.get("Senior", 0)
        ratio_ls     = round(lead_count / senior_count, 2) if senior_count else 0

        score, score_color = self._compute_health_score()

        # Metric cards
        metrics = [
            ("Employees analysed",    total,          TEAL),
            ("Functions covered",     functions,      BLUE),
            ("Lead-level roles",      lead_count,     VIOLET),
            ("Lead : Senior ratio",   f"1 : {round(senior_count/max(lead_count,1),1)}", AMBER),
            ("Org health score",      f"{score}/100",  score_color),
        ]
        row = 5
        ws.row_dimensions[row].height = 22
        _hdr(ws, row, 2, "METRIC", bg=INK)
        _hdr(ws, row, 3, "VALUE",  bg=INK)
        for i, (label, val, color) in enumerate(metrics):
            r = row+1+i
            ws.row_dimensions[r].height = 26
            _cell(ws, r, 2, label, bold=True, bg=_row_bg(r))
            c = ws.cell(r, 3, str(val))
            c.font      = Font(name="Arial", bold=True, color=color, size=13)
            c.fill      = PatternFill("solid", fgColor=_row_bg(r))
            c.alignment = Alignment(vertical="center")

        row = row + len(metrics) + 2
        # Run pattern detection and populate findings
        self._detect_patterns()

        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"B{row}:D{row}")
        _hdr(ws, row, 2, "KEY FINDINGS", bg=AMBER)
        for i, f in enumerate(self._findings[:6]):
            r = row+1+i
            ws.row_dimensions[r].height = 32
            sev_col = RED if f["severity"]=="high" else (AMBER if f["severity"]=="medium" else MUTED)
            _cell(ws, r, 2, f["title"],   bold=True, fg=sev_col, bg=_row_bg(r))
            _cell(ws, r, 3, f["detail"],  fg=INK,    bg=_row_bg(r))
            ws.cell(r,3).alignment = Alignment(wrap_text=True, vertical="center")

        row = row + len(self._findings[:6]) + 2
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"B{row}:D{row}")
        _hdr(ws, row, 2, "TOP RECOMMENDATIONS", bg=GREEN)
        for i, rec in enumerate(self._recommendations[:5]):
            r = row+1+i
            ws.row_dimensions[r].height = 36
            _cell(ws, r, 2, f"#{i+1}  {rec['title']}", bold=True, fg=GREEN, bg=_row_bg(r))
            _cell(ws, r, 3, rec["action"], fg=INK, bg=_row_bg(r))
            ws.cell(r,3).alignment = Alignment(wrap_text=True, vertical="center")

        _border_range(ws, 5, row+5, 2, 4)

    # ── Sheet 2: Job Architecture ─────────────────────────────────────────
    def _build_job_architecture(self):
        ws = self._wb.create_sheet("3. Job Architecture")
        ws.sheet_view.showGridLines = False
        col_ws = [3,14,22,14,10,12,12,12,12,12,12]
        headers = ["","Function","Standard Role","Grade","Band",f"Min {self.currency}", f"P25 {self.currency}", f"P50 {self.currency}", f"P75 {self.currency}", f"Max {self.currency}","Hay Range"]
        for ci,(w,h) in enumerate(zip(col_ws,headers),1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws,1,ci,h, bg=TEAL)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        repo  = self.catalog.repository
        jobs  = sorted(repo.jobs.values(), key=lambda j: (j.function, _grade(j) or 0, j.standard_title))
        jg    = repo.job_grades

        # Load grade→band map from repo if available
        GRADE_BANDS = self.GRADE_BANDS
        ri = 2
        prev_fn = None
        for job in jobs:
            if job.function != prev_fn:
                ws.row_dimensions[ri].height = 18
                _hdr(ws, ri, 2, job.function, bg=BLUE, size=9)
                ws.merge_cells(f"C{ri}:K{ri}")
                ri += 1
                prev_fn = job.function
            band  = GRADE_BANDS.get(_grade(job) or 0, "Professional")
            band_col = BAND_COLORS.get(band, MUTED)
            band_lt  = BAND_LIGHT.get(band, GREY)
            sal   = repo.salary.get((job.function, job.level))
            grade_data = jg.get(_grade(job) or 0)
            hay_range  = f"{grade_data.pay_min}–{grade_data.pay_max}" if grade_data and hasattr(grade_data,"pay_min") else ""
            # Actually use Hay from JobGrades (stored as pay_min/pay_max for Hay there is in grade criteria)
            # Let's use grade-based Hay ranges from our grade_rows
            HAY = {1:(100,130),2:(135,160),3:(165,200),4:(205,250),5:(255,300),6:(305,365),
                   7:(370,440),8:(450,525),9:(530,615),10:(620,725),11:(730,870),12:(880,1040),
                   13:(1050,1260),14:(1270,1800)}
            hay = HAY.get(_grade(job) or 0)
            hay_str = f"{hay[0]}–{hay[1]}" if hay else ""
            bg = _row_bg(ri)
            _cell(ws, ri, 1, "",                   bg=band_lt)
            _cell(ws, ri, 2, job.function,          bg=bg)
            _cell(ws, ri, 3, job.standard_title,    bg=bg, bold=True)
            _cell(ws, ri, 4, f"G{_grade(job)}" if _grade(job) else "", bg=bg, fg=band_col, bold=True)
            _cell(ws, ri, 5, band,                  bg=bg, fg=band_col)
            for ci_s, attr in enumerate(["min_salary","p25","p50","p75","max_salary"], 6):
                val = int(getattr(sal, attr, 0) or 0) if sal else 0
                _cell(ws, ri, ci_s, self._money(val) if val else "—", bg=bg, fg=TEAL if val else MUTED)
            _cell(ws, ri, 11, hay_str, bg=bg, fg=MUTED)
            ws.row_dimensions[ri].height = 20
            ri += 1

        _border_range(ws, 1, ri-1, 1, 11)

    # ── Sheet 3: Org Snapshot ─────────────────────────────────────────────
    def _build_org_snapshot(self):
        ws = self._wb.create_sheet("4. Org Snapshot")
        ws.sheet_view.showGridLines = False
        headers = ["Name","Input Title","Matched Role","Function","Level","Grade","Band",
                   "Salary Band Min","P50 Market","Salary","Pay Position","Confidence"]
        widths  = [22,26,24,16,10,8,18,14,12,12,14,12]
        for ci,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws,1,ci,h)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        repo = self.catalog.repository
        GRADE_BANDS = self.GRADE_BANDS

        def get_name(idx):
            if self.df_employees is None or idx >= len(self.df_employees): return f"Employee {idx+1}"
            row = self.df_employees.iloc[idx]
            fn = next((str(row[c]) for c in ["FirstName","first_name"] if c in self.df_employees.columns), "")
            ln = next((str(row[c]) for c in ["LastName","last_name"]   if c in self.df_employees.columns), "")
            return (fn+" "+ln).strip() or f"Employee {idx+1}"

        def get_salary(idx):
            if self.df_employees is None or idx >= len(self.df_employees): return None
            row = self.df_employees.iloc[idx]
            for col in ["AnnualSalaryEUR","Salary","salary","annual_salary"]:
                if col in self.df_employees.columns:
                    try: return float(row[col])
                    except: pass
            return None

        all_results = [(i,r) for i,r in enumerate(
            [r for r in self.results], 0)]

        for ri, (idx, r) in enumerate(all_results, 2):
            job  = repo.jobs.get(r.job_id)
            sal  = repo.salary.get((r.function, r.level)) if job else None
            grade = _grade(job) if job else 0
            band  = GRADE_BANDS.get(grade, "Professional")
            band_col = BAND_COLORS.get(band, MUTED)
            salary   = get_salary(idx)
            p50      = sal.p50 if sal else 0
            if salary and p50:
                pay_pct  = salary / p50
                if pay_pct < 0.85:   pay_pos, pay_col = "Below P25",  RED
                elif pay_pct < 0.97: pay_pos, pay_col = "Below P50",  AMBER
                elif pay_pct < 1.10: pay_pos, pay_col = "At market",  TEAL
                elif pay_pct < 1.25: pay_pos, pay_col = "Above P75",  BLUE
                else:                pay_pos, pay_col = "Above band",  VIOLET
            else:
                pay_pos, pay_col = "—", MUTED

            bg = _row_bg(ri)
            vals = [
                (get_name(idx), INK, False),
                (r.input_title, MUTED, False),
                (r.standard_title, INK, True),
                (r.function, INK, False),
                (r.level, INK, False),
                (f"G{grade}" if grade else "—", band_col, True),
                (band, band_col, False),
                (self._money(sal.min_salary) if sal else "—", MUTED, False),
                (self._money(p50) if p50 else "—", TEAL, False),
                (self._money(salary) if salary is not None else "—", INK, False),
                (pay_pos, pay_col, False),
                (f"{r.confidence}%", TEAL if r.confidence>=96 else AMBER, False),
            ]
            for ci,(text,fg,bold) in enumerate(vals,1):
                _cell(ws, ri, ci, text, fg=fg, bg=bg, bold=bold)
            ws.row_dimensions[ri].height = 20

        _border_range(ws, 1, len(all_results)+1, 1, len(headers))

    # ── Sheet 4: Grade Distribution ───────────────────────────────────────
    def _build_grade_distribution(self):
        ws = self._wb.create_sheet("5. Grade Distribution")
        ws.sheet_view.showGridLines = False
        headers = ["Grade","Band","Employees","% of Total",f"Band Min {self.currency}", f"P50 {self.currency}", f"Band Max {self.currency}","Avg Salary","Pay Position vs P50"]
        widths  = [10,20,12,12,12,12,12,12,22]
        for ci,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws,1,ci,h,bg=BLUE)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes="A2"

        repo = self.catalog.repository
        total = max(len(self.results),1)
        grade_data: dict[int,list] = {}
        for r in self.results:
            job = repo.jobs.get(r.job_id)
            g = _grade(job) if job else 0
            grade_data.setdefault(g,[]).append(r)

        for ri, grade in enumerate(sorted(grade_data.keys()),2):
            items = grade_data[grade]
            band  = self.GRADE_BANDS.get(grade,"Professional")
            sal   = next((repo.salary.get((r.function,r.level)) for r in items
                          if repo.salary.get((r.function,r.level))), None)
            count = len(items); pct = round(count/total*100,1)
            bg = BAND_LIGHT.get(band, _row_bg(ri))
            band_col = BAND_COLORS.get(band, MUTED)
            _cell(ws,ri,1,f"G{grade}" if grade else "—",fg=band_col,bg=bg,bold=True)
            _cell(ws,ri,2,band,fg=band_col,bg=bg)
            _cell(ws,ri,3,count,fg=INK,bg=bg)
            _cell(ws,ri,4,f"{pct}%",fg=MUTED,bg=bg)
            _cell(ws,ri,5,self._money(sal.min_salary) if sal else "—",fg=MUTED,bg=bg)
            _cell(ws,ri,6,self._money(sal.p50) if sal else "—",fg=TEAL,bg=bg)
            _cell(ws,ri,7,self._money(sal.max_salary) if sal else "—",fg=MUTED,bg=bg)
            _cell(ws,ri,8,"—",fg=MUTED,bg=bg)
            _cell(ws,ri,9,"—",fg=MUTED,bg=bg)
            ws.row_dimensions[ri].height = 22

        _border_range(ws,1,len(grade_data)+1,1,len(headers))

    # ── Sheet 5: Career Paths ─────────────────────────────────────────────
    def _build_career_paths(self):
        ws = self._wb.create_sheet("6. Career Paths")
        ws.sheet_view.showGridLines = False
        headers = ["Function","From Role","Level","Grade","→ Next Role","Next Level","Next Grade","Steps to Lead"]
        widths  = [18,28,10,8,28,10,10,14]
        for ci,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws,1,ci,h,bg=VIOLET)
        ws.row_dimensions[1].height=22
        ws.freeze_panes="A2"

        repo = self.catalog.repository
        LSORT={"Junior":1,"Medior":2,"Senior":3,"Lead":4}
        LEAD_GRADE={"HR":11,"Finance":14,"Engineering":13,"Data":13,"Product":12,
                    "Operations":12,"Sales":12,"Marketing":11,"Customer":11,"Legal":13,"Executive":14}

        paths = sorted(repo.career_paths.values(), key=lambda cp:
            (repo.jobs[cp.job_id].function if cp.job_id in repo.jobs else "",
             LSORT.get(repo.jobs[cp.job_id].level if cp.job_id in repo.jobs else "Junior",0)))

        for ri, cp in enumerate(paths, 2):
            job  = repo.jobs.get(cp.job_id)
            next_job = repo.jobs.get(cp.next_job_id) if cp.next_job_id else None
            if not job: continue
            # steps to Lead
            cur_grade  = _grade(job) or 5
            lead_grade = LEAD_GRADE.get(job.function, 11)
            steps_to_lead = max(0, lead_grade - cur_grade)
            band_col = BAND_COLORS.get(self.GRADE_BANDS.get(_grade(job) or 0,"Professional"), MUTED)
            bg = _row_bg(ri)
            _cell(ws,ri,1,job.function,fg=INK,bg=bg)
            _cell(ws,ri,2,job.standard_title,fg=INK,bg=bg,bold=True)
            _cell(ws,ri,3,job.level,fg=band_col,bg=bg)
            _cell(ws,ri,4,f"G{_grade(job)}" if _grade(job) else "—",fg=band_col,bg=bg,bold=True)
            _cell(ws,ri,5,next_job.standard_title if next_job else "Top of path",
                  fg=TEAL if next_job else MUTED, bg=bg)
            _cell(ws,ri,6,next_job.level if next_job else "—",fg=MUTED,bg=bg)
            ngrade = _grade(next_job) if next_job else 0
            _cell(ws,ri,7,f"G{ngrade}" if ngrade else "—",fg=MUTED,bg=bg)
            _cell(ws,ri,8,str(steps_to_lead),fg=AMBER if steps_to_lead>3 else TEAL,bg=bg)
            ws.row_dimensions[ri].height=20

        _border_range(ws,1,len(paths)+1,1,len(headers))

    # ── Sheet 6: Succession Risk ──────────────────────────────────────────
    def _build_succession_risk(self):
        ws = self._wb.create_sheet("7. Succession Risk")
        ws.sheet_view.showGridLines = False
        headers=["Role","Function","Grade","Ready Now","6-12 Months","Developing","Total Pipeline","Risk Status","Action Required"]
        widths=[26,16,8,12,14,12,14,14,36]
        for ci,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws,1,ci,h,bg=RED)
        ws.row_dimensions[1].height=22
        ws.freeze_panes="A2"

        repo  = self.catalog.repository
        LSORT = {"Junior":1,"Medior":2,"Senior":3,"Lead":4}
        RELATED={
            "HR":{"HR","Operations","Legal"},"Finance":{"Finance","Operations","Legal"},
            "Engineering":{"Engineering","Data","Product"},"Data":{"Data","Engineering","Product"},
            "Product":{"Product","Engineering","Data"},"Operations":{"Operations","HR","Finance"},
            "Sales":{"Sales","Marketing","Customer"},"Marketing":{"Marketing","Sales","Customer"},
            "Customer":{"Customer","Sales","Operations"},"Legal":{"Legal","Finance","HR"},
        }
        def readiness(gaps):
            if not gaps: return 0
            return round(sum(1 for g in gaps if g["gap"]<=0)/len(gaps)*100)

        role_pool={r.job_id:[] for r in self.results}
        for i,r in enumerate(self.results): role_pool[r.job_id].append(i)

        lead_roles=[j for j in repo.jobs.values() if j.level=="Lead"]
        ri=2
        for job in sorted(lead_roles, key=lambda j:(j.function,j.standard_title)):
            rn=near=dev=0
            for job_id,indices in role_pool.items():
                if job_id==job.job_id: continue
                fj=repo.jobs.get(job_id)
                if not fj: continue
                delta=LSORT.get(job.level,4)-LSORT.get(fj.level,1)
                same=fj.function==job.function
                rel=fj.function in RELATED.get(job.function,{job.function})
                if delta<=0 and not same: continue
                if delta>2 or not rel: continue
                csk={req.skill_id:req.required_level for req,_ in self.catalog.get_role_skills(job_id)}
                try: gs=self.catalog.skill_gap(csk,job.job_id)
                except: gs=[]
                raw=readiness(gs)
                sc=min(100,int(raw*1.15)) if same and delta==1 else (min(100,int(raw*1.05)) if same else max(0,int(raw*0.90)))
                if sc>=80: rn+=len(indices)
                elif sc>=55: near+=len(indices)
                else: dev+=len(indices)
            total=rn+near+dev
            risk="✅ Covered" if rn>0 else ("⚠️ At Risk" if total>0 else "🔴 Critical")
            risk_col=TEAL if rn>0 else (AMBER if total>0 else RED)
            action=("Maintain pipeline — consider stretch assignments for ready candidates." if rn>0
                    else ("Accelerate development for near-ready candidates. Set succession target date." if total>0
                          else "No pipeline. Prioritise external search and internal development programme immediately."))
            bg=_row_bg(ri)
            grade=_grade(job) or 0
            _cell(ws,ri,1,job.standard_title,fg=INK,bg=bg,bold=True)
            _cell(ws,ri,2,job.function,fg=MUTED,bg=bg)
            _cell(ws,ri,3,f"G{grade}" if grade else "—",fg=VIOLET,bg=bg,bold=True)
            _cell(ws,ri,4,str(rn),fg=TEAL if rn>0 else MUTED,bg=bg,bold=rn>0)
            _cell(ws,ri,5,str(near),fg=AMBER if near>0 else MUTED,bg=bg)
            _cell(ws,ri,6,str(dev),fg=MUTED,bg=bg)
            _cell(ws,ri,7,str(total),fg=INK,bg=bg,bold=True)
            _cell(ws,ri,8,risk,fg=risk_col,bg=bg,bold=True)
            _cell(ws,ri,9,action,fg=INK,bg=bg)
            ws.cell(ri,9).alignment=Alignment(wrap_text=True,vertical="center")
            ws.row_dimensions[ri].height=36
            ri+=1
        _border_range(ws,1,ri-1,1,len(headers))

    # ── Sheet 7: Recommendations ──────────────────────────────────────────
    def _build_recommendations(self):
        ws = self._wb.create_sheet("8. Recommendations")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 52
        ws.column_dimensions["D"].width = 20

        ws.row_dimensions[1].height=10
        ws.row_dimensions[2].height=40
        ws.merge_cells("B2:D2")
        c=ws.cell(2,2,"Recommendations — Job Architecture & People Strategy")
        c.font=Font(name="Arial",bold=True,color=WHITE,size=16)
        c.fill=PatternFill("solid",fgColor=GREEN)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1)

        categories={"Architecture":[],"Compensation":[],"Succession":[],"Development":[]}
        for rec in self._recommendations:
            categories.setdefault(rec.get("category","Architecture"),[]).append(rec)

        row=4
        for cat, recs in categories.items():
            if not recs: continue
            ws.row_dimensions[row].height=22
            ws.merge_cells(f"B{row}:D{row}")
            bg={"Architecture":BLUE,"Compensation":TEAL,"Succession":VIOLET,"Development":AMBER}.get(cat,BLUE)
            _hdr(ws,row,2,cat.upper(),bg=bg)
            row+=1
            for ri_i,rec in enumerate(recs):
                ws.row_dimensions[row].height=52
                ws.merge_cells(f"C{row}:D{row}")
                bgr=_row_bg(row)
                _cell(ws,row,2,f"#{ri_i+1}  {rec['title']}",bold=True,fg=bg,bg=bgr)
                _cell(ws,row,3,f"{rec['action']}\n\nImpact: {rec.get('impact','')}",fg=INK,bg=bgr)
                ws.cell(row,3).alignment=Alignment(wrap_text=True,vertical="center")
                row+=1
            row+=1
        _border_range(ws,2,row,2,4)

    # ── Sheet 8: Job Family & Pay (leveling grid) ─────────────────────────
    def _build_job_family_leveling(self):
        ws = self._wb.create_sheet("9. Job Family & Pay")
        ws.sheet_view.showGridLines = False
        headers = ["Role", "Level", "Grade", "Salary Min", "Median (P50)", "Salary Max",
                   "Knowledge / Scope", "Problem Solving", "Accountability", "Top Skills"]
        widths  = [30, 10, 8, 13, 13, 13, 34, 30, 34, 34]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws, 1, ci, h, bg=TEAL)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # From the loaded library, for the same reason as the pay sheet below.
        frames = getattr(self.catalog, "frames", None) or {}
        try:
            jobs = frames["jobs"].copy()
            bands = frames["salary"]
            grades = frames["jobgrades"]
        except KeyError as exc:
            _cell(ws, 2, 1, f"Reference data not available ({exc}).", fg=MUTED)
            return
        jobs["Grade"] = pd.to_numeric(jobs.get("Grade"), errors="coerce")
        bmap = {(r["Function"], r["Level"]): r for _, r in bands.iterrows()}
        gmap = {}
        for _, r in grades.iterrows():
            try: gmap[int(float(r["Grade"]))] = r
            except (TypeError, ValueError): pass

        def _e(v):
            try: return self._money(float(v))
            except Exception: return "—"
        def _t(v, n=180):
            s = "" if v is None else str(v)
            if not s or s.lower() == "nan": return "—"
            s = s.replace(";", " · ")
            return s if len(s) <= n else s[:n].rsplit(" ", 1)[0] + "…"
        def _skills(jid):
            try:
                names = [sk.skill_name for _, sk in self.catalog.get_role_skills(jid)[:3]]
                return " · ".join(names) if names else "—"
            except Exception:
                return "—"

        ri = 2
        for fn in sorted(jobs["Function"].dropna().unique()):
            fam = jobs[jobs["Function"] == fn].dropna(subset=["Grade"]).sort_values("Grade")
            if fam.empty:
                continue
            _hdr(ws, ri, 1, f"{fn} — Job Family", bg=BLUE)
            for ci in range(2, len(headers) + 1):
                _cell(ws, ri, ci, "", bg=BLUE_L)
            ws.row_dimensions[ri].height = 20
            ri += 1
            for role in fam.itertuples(index=False):
                jid = getattr(role, "JobID"); lvl = getattr(role, "Level")
                grade = int(getattr(role, "Grade"))
                b = bmap.get((getattr(role, "Function"), lvl)); g = gmap.get(grade)
                bg = _row_bg(ri)
                _cell(ws, ri, 1, _t(getattr(role, "StandardTitle"), 60), fg=INK, bg=bg, bold=True)
                _cell(ws, ri, 2, _t(lvl, 20), fg=MUTED, bg=bg)
                _cell(ws, ri, 3, f"G{grade}", fg=MUTED, bg=bg)
                _cell(ws, ri, 4, _e(b["Min"]) if b is not None else "—", fg=MUTED, bg=bg)
                _cell(ws, ri, 5, _e(b["P50"]) if b is not None else "—", fg=TEAL, bg=bg, bold=True)
                _cell(ws, ri, 6, _e(b["Max"]) if b is not None else "—", fg=MUTED, bg=bg)
                _cell(ws, ri, 7, _t(g["Scope"]) if g is not None else "—", fg=INK, bg=bg)
                _cell(ws, ri, 8, _t(g["Complexity"]) if g is not None else "—", fg=INK, bg=bg)
                _cell(ws, ri, 9, _t(g["DecisionRights"]) if g is not None else "—", fg=INK, bg=bg)
                _cell(ws, ri, 10, _skills(jid), fg=INK, bg=bg)
                ws.row_dimensions[ri].height = 30
                ri += 1
            ri += 1  # spacer between families

        _border_range(ws, 1, ri - 1, 1, len(headers))

    # ── Sheet 9: Total Pay & Reward ───────────────────────────────────────
    def _build_total_pay(self):
        ws = self._wb.create_sheet("10. Total Pay & Reward")
        ws.sheet_view.showGridLines = False
        headers = ["Role", "Function", "Level", "Base median (P50)", "Holiday allowance",
                   "13th month", "Variable (on-target)", "Total target cash",
                   "Employer pension", "Other benefits", "Total reward", "LTI"]
        widths = [30, 14, 10, 16, 13, 16, 18, 18, 14, 13, 16, 8]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws, 1, ci, h, bg=TEAL)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        # Read the library the app has loaded, not the workbook on disk: after
        # the cutover that file is a snapshot, and a board report priced off a
        # stale copy of the bands is the worst place for that difference to
        # land. The pay composition comes from the same service the pages use.
        from services import pay_components_service as _pay
        frames = getattr(self.catalog, "frames", None) or {}
        repo = self.catalog.repository
        try:
            jobs = frames["jobs"].copy()
            bands = frames["salary"].copy()
        except KeyError as exc:
            _cell(ws, 2, 1, f"Pay data not available ({exc}).", fg=MUTED)
            return
        jobs["Grade"] = pd.to_numeric(jobs.get("Grade"), errors="coerce")
        for _c in ("Min", "P50", "Max"):
            if _c in bands: bands[_c] = pd.to_numeric(bands[_c], errors="coerce")
        bmap = {(r["Function"], r["Level"]): r for _, r in bands.iterrows()}

        def _e(v):
            try: return self._money(float(v))
            except Exception: return "—"

        ri = 2
        for fn in sorted(jobs["Function"].dropna().unique()):
            fam = jobs[jobs["Function"] == fn].dropna(subset=["Grade"]).sort_values("Grade")
            for role in fam.itertuples(index=False):
                lvl = getattr(role, "Level"); b = bmap.get((fn, lvl))
                if b is None or pd.isna(b.get("P50")):
                    continue
                base = float(b.get("P50"))
                comp = _pay.compose(base, fn, lvl, repo)

                def _part(key, comp=comp):
                    """One component as text: the amount and the rate behind it."""
                    c = next((c for c in comp.components if c.key == key), None)
                    if c is not None and c.amount is not None:
                        return f"{_e(c.amount)} ({c.pct:g}%)"
                    if c is not None and c.low_amount is not None:
                        return (f"{_e(c.low_amount)} – {_e(c.high_amount)} "
                                f"({c.low_pct:g}–{c.high_pct:g}%)")
                    return "not stated"

                lti = ("—" if comp.lti_eligible is None
                       else ("Yes" if comp.lti_eligible else "No"))
                bg = _row_bg(ri)
                _cell(ws, ri, 1, str(getattr(role, "StandardTitle")), fg=INK, bg=bg, bold=True)
                _cell(ws, ri, 2, fn, fg=MUTED, bg=bg)
                _cell(ws, ri, 3, str(lvl), fg=MUTED, bg=bg)
                _cell(ws, ri, 4, _e(base), fg=INK, bg=bg)
                _cell(ws, ri, 5, _part("holiday"), fg=MUTED, bg=bg)
                _cell(ws, ri, 6, _part("thirteenth"), fg=MUTED, bg=bg)
                _cell(ws, ri, 7, _part("variable"), fg=MUTED, bg=bg)
                _cell(ws, ri, 8, _e(comp.total_target_cash), fg=TEAL, bg=bg, bold=True)
                _cell(ws, ri, 9, _part("pension"), fg=MUTED, bg=bg)
                _cell(ws, ri, 10, _part("benefits"), fg=MUTED, bg=bg)
                _cell(ws, ri, 11, (f"{_e(comp.total_reward_low)} – {_e(comp.total_reward_high)}"
                                   if comp.is_range else _e(comp.total_target_cash)),
                      fg=VIOLET, bg=bg, bold=True)
                _cell(ws, ri, 12, str(lti), fg=MUTED, bg=bg)
                ws.row_dimensions[ri].height = 20
                ri += 1
        _cell(ws, ri + 1, 1,
              "Every rate is taken from the reference library: holiday allowance and employer "
              "pension from PayElements, 13th month and on-target variable from PayMix for that "
              "Function x Level. Pension is stated as a range, so total reward is a range. "
              "A component the library does not state as a rate reads 'not stated' and is left "
              "out of the totals rather than estimated.",
              fg=MUTED, italic=True)
        _border_range(ws, 1, ri - 1, 1, len(headers))

    # ── Sheet 10: Pay Equity (compa-ratio) ────────────────────────────────
    def _build_pay_equity(self):
        import re as _re
        ws = self._wb.create_sheet("11. Pay Equity")
        ws.sheet_view.showGridLines = False
        for ci, w in enumerate([30, 26, 10, 12, 12, 12, 14], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        df = self.df_employees

        def _find(colset, keys):
            low = {str(c).strip().lower(): c for c in colset}
            for k in keys:
                for lc, orig in low.items():
                    if k in lc:
                        return orig
            return None

        _hdr(ws, 1, 1, "Pay Equity — compa-ratio vs role bands", bg=TEAL)
        for ci in range(2, 8):
            _cell(ws, 1, ci, "", bg=TEAL)
        if df is None or len(df) == 0:
            _cell(ws, 2, 1, "Upload workforce data with an actual-salary column to see compa-ratios.", fg=MUTED)
            return
        sal_col = _find(df.columns, ["actualsalary", "salaris", "salary", "loon", "bruto", "pay"])
        title_col = _find(df.columns, ["jobtitle", "functie", "title", "role", "current"])
        name_col = _find(df.columns, ["name", "naam"])
        gender_col = _find(df.columns, ["gender", "geslacht", "sex"])
        fte_col = _find(df.columns, ["fte", "parttime", "deeltijd"])
        if not sal_col:
            _cell(ws, 2, 1, "No actual-salary column in the uploaded data — add one (e.g. ActualSalary) "
                            "to compute compa-ratios and outliers.", fg=MUTED)
            return

        repo = self.catalog.repository
        rmap = {}
        for r in self.results:
            if r.input_title:
                rmap.setdefault(str(r.input_title).strip().lower(), r)

        rows = []
        for _, er in df.iterrows():
            s = _re.sub(r"[^\d]", "", str(er.get(sal_col, "")))
            if not s:
                continue
            actual = int(s)
            title = str(er.get(title_col, "")).strip() if title_col else ""
            res = rmap.get(title.lower())
            band = repo.salary.get((res.function, res.level)) if (res and res.matched) else None
            if band is None:
                continue
            # The SAME placement the Pay Equity page uses. This sheet had its own
            # copy -- actual / p50, no FTE -- so after the app started pro-rating
            # part-time pay the report and the screen disagreed about the same
            # person. One calculation, one place.
            _fte = None
            if fte_col:
                try:
                    _fte = float(str(er.get(fte_col, "")).replace(",", "."))
                except (TypeError, ValueError):
                    _fte = None
            pos = _salary.position(actual, band, _fte)
            rows.append({"name": (str(er.get(name_col)) if name_col else title), "role": res.standard_title,
                         "level": res.level, "actual": actual, "p50": int(pos.band_p50 or 0),
                         "min": int(pos.band_min or 0), "compared": pos.compared_pay or actual,
                         "compa": pos.compa_ratio or 0, "status": pos.status, "pro_rated": pos.pro_rated,
                         "gender": (str(er.get(gender_col, "")).strip().upper()[:1] if gender_col else "")})
        if not rows:
            _cell(ws, 2, 1, "No salaries could be matched to role bands.", fg=MUTED)
            return

        compas = [r["compa"] for r in rows if r["compa"]]
        avg = round(sum(compas) / len(compas), 2)
        below = sum(1 for r in rows if r["status"] == "Below range")
        above = sum(1 for r in rows if r["status"] == "Above range")
        # The board question is not "how many", it is "what does it cost to
        # stop". Annual base cost to lift everyone below their band up to its
        # minimum -- the smallest defensible remediation, not the nicest one.
        self._facts.update({
            "priced": len(rows), "avg_compa": avg, "below": below, "above": above,
            "pro_rated": sum(1 for r in rows if r.get("pro_rated")),
            # Measured on the pay that was COMPARED (full-time-equivalent where an
            # FTE was supplied), because that is the basis the "below range"
            # verdict was reached on. Mixing the two would quote a cost for a
            # finding made another way.
            "lift_to_minimum": sum(max(0, r["min"] - r["compared"])
                                   for r in rows if r.get("min") and r["status"] == "Below range"),
        })
        SC = {"Below range": RED, "Below market": AMBER, "At market": GREEN,
              "Above market": BLUE, "Above range": VIOLET}
        # summary block
        summ = [("Employees priced", len(rows), INK), ("Average compa-ratio", avg, TEAL),
                ("Below range (underpaid)", below, RED), ("Above range", above, VIOLET)]
        for i, (lab, val, col) in enumerate(summ):
            rr = 2 + i
            _cell(ws, rr, 1, lab, bold=True, bg=_row_bg(rr))
            _cell(ws, rr, 2, val, fg=col, bold=True, bg=_row_bg(rr))
        r0 = 2 + len(summ) + 1
        if gender_col:
            gm = [r["actual"] for r in rows if r["gender"] == "M"]
            gf = [r["actual"] for r in rows if r["gender"] == "F"]
            if gm and gf:
                raw = round((sum(gm) / len(gm) - sum(gf) / len(gf)) / (sum(gm) / len(gm)) * 100, 1)
                _cell(ws, r0, 1, "Gender pay gap (raw M vs F)", bold=True)
                _cell(ws, r0, 2, f"{raw:+.1f}%", fg=(RED if abs(raw) >= 5 else GREEN), bold=True)
                r0 += 2
        # per-employee table (outliers first: sort by |compa-1| desc)
        rows.sort(key=lambda r: abs(r["compa"] - 1), reverse=True)
        hdr_row = r0
        for ci, h in enumerate(["Name / title", "Matched role", "Level", "Actual", "Band P50", "Compa-ratio", "Status"], 1):
            _hdr(ws, hdr_row, ci, h, bg=INK)
        for i, r in enumerate(rows):
            rr = hdr_row + 1 + i
            bg = _row_bg(rr)
            _cell(ws, rr, 1, r["name"], bg=bg)
            _cell(ws, rr, 2, r["role"], bg=bg)
            _cell(ws, rr, 3, str(r["level"]), fg=MUTED, bg=bg)
            _cell(ws, rr, 4, self._money(r['actual']), bg=bg)
            _cell(ws, rr, 5, self._money(r['p50']), fg=MUTED, bg=bg)
            _cell(ws, rr, 6, r["compa"], bg=bg)
            _cell(ws, rr, 7, r["status"], fg=SC.get(r["status"], MUTED), bold=True, bg=bg)
        _border_range(ws, hdr_row, hdr_row + len(rows), 1, 7)

    # ── Sheet 11: Benefits Benchmarking (market reference) ────────────────
    def _build_benefits_benchmarking(self):
        ws = self._wb.create_sheet("12. Benefits Benchmarking")
        ws.sheet_view.showGridLines = False
        repo = self.catalog.repository
        if not getattr(repo, "benefits_catalog", None):
            _cell(ws, 1, 1, "No BenefitsCatalog data in the reference workbook.", fg=MUTED)
            return

        from services.benefits_service import BenefitsService
        svc = BenefitsService(self.catalog, country=self.country or None)
        industries = sorted(repo.industries.keys()) if repo.industries else []

        headers = ["Category", "Unit", "Market P25", "Market Median", "Market P75", "Market P90", "n"]
        headers += [f"{iid} median" for iid in industries]
        widths = [24, 10, 12, 13, 12, 12, 8] + [14] * len(industries)
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            _hdr(ws, 1, ci, h, bg=VIOLET)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        def _fmt(v, unit):
            if v is None:
                return "—"
            return self._money(v) if unit == "EUR" else f"{v:g} {unit}".strip()

        ri = 2
        for category in svc.categories():
            item = svc.catalog_item(category)
            unit = item.unit if item else ""
            pooled = svc.get_band(category, None, None)
            bg = _row_bg(ri)
            _cell(ws, ri, 1, category, fg=INK, bg=bg, bold=True)
            _cell(ws, ri, 2, unit, fg=MUTED, bg=bg)
            if pooled:
                _cell(ws, ri, 3, _fmt(pooled.p25, unit), fg=MUTED, bg=bg)
                _cell(ws, ri, 4, _fmt(pooled.p50, unit), fg=TEAL, bg=bg, bold=True)
                _cell(ws, ri, 5, _fmt(pooled.p75, unit), fg=MUTED, bg=bg)
                _cell(ws, ri, 6, _fmt(pooled.p90, unit), fg=MUTED, bg=bg)
                _cell(ws, ri, 7, pooled.n_observations, fg=MUTED, bg=bg)
            else:
                for ci in range(3, 8):
                    _cell(ws, ri, ci, "—", fg=MUTED, bg=bg)
            for ci, iid in enumerate(industries, 8):
                band = svc.get_band(category, iid, None)
                _cell(ws, ri, ci, _fmt(band.p50, unit) if band else "—", fg=INK, bg=bg)
            ws.row_dimensions[ri].height = 20
            ri += 1

        _border_range(ws, 1, ri - 1, 1, len(headers))
        note_row = ri + 1
        _cell(ws, note_row, 1,
              "Percentiles and medians are computed at run time from the self-built BenefitsObservations "
              "reference data (not static columns) — see services/benefits_service.py. For a company-specific "
              "comparison, percentile rank, status and advice, use the live Benefits Benchmarking page.",
              fg=MUTED, italic=True)

    # ── Pattern detection & recommendations ───────────────────────────────
    def _detect_patterns(self):
        repo = self.catalog.repository
        LSORT={"Junior":1,"Medior":2,"Senior":3,"Lead":4}
        findings=[]; recs=[]

        # Grade distribution
        grade_dist:dict[int,int]={}
        for r in self.results:
            job=repo.jobs.get(r.job_id)
            g=_grade(job) if job else 0
            grade_dist[g]=grade_dist.get(g,0)+1
        total=max(len(self.results),1)

        # Pattern 1: Level inflation (too many Lead relative to total)
        lead_pct=sum(1 for r in self.results if r.level=="Lead")/total
        if lead_pct>0.3:
            findings.append({"severity":"high","title":"Level inflation detected",
                "detail":f"{lead_pct*100:.0f}% of matched roles are Lead-level (benchmark: 10-20%). Review whether all Lead titles reflect genuine leadership scope or whether re-grading is needed."})
            recs.append({"category":"Architecture","title":"Review and right-size Lead-level designations",
                "action":"Audit all Lead-level roles against the Grade 11–14 criteria (function ownership, P&L accountability, board proximity). Roles that do not meet these criteria should be re-classified to Senior/Expert band.",
                "impact":"Improved grade integrity, clearer career pathways, and more credible succession pipeline."})

        # Pattern 2: Grade compression at single grade
        if grade_dist:
            top_g=max(grade_dist,key=grade_dist.get)
            top_pct=grade_dist[top_g]/total
            if top_pct>0.4:
                findings.append({"severity":"medium","title":f"Grade compression at G{top_g}",
                    "detail":f"{top_pct*100:.0f}% of employees are at G{top_g}. This suggests either accurate job evaluation with a large homogeneous workforce, or insufficient role differentiation within the band."})
                recs.append({"category":"Architecture","title":f"Differentiate roles within G{top_g} band",
                    "action":f"Review roles mapped to G{top_g}. Consider introducing sub-grades (e.g. G{top_g}a, G{top_g}b) or distinct role profiles to recognise performance and scope differences within the same grade.",
                    "impact":"Better retention of high performers, more meaningful progression, reduced flight risk."})

        # Pattern 3: Function with no Lead role
        # An unmatched title carries no function, and None is not a function
        # with no Lead role -- it is a row this pattern cannot speak about. It
        # used to be counted anyway, and then joined into a sentence, which
        # raised TypeError and took the whole report down: every generate with
        # one unmatched title failed, which is most real client files.
        fn_levels:dict[str,set]={};
        for r in self.results:
            if not r.function:
                continue
            fn_levels.setdefault(str(r.function),set()).add(r.level)
        no_lead_fns=[fn for fn,lvls in fn_levels.items() if "Lead" not in lvls and fn!="Executive"]
        if no_lead_fns:
            findings.append({"severity":"high","title":"Functions without leadership representation",
                "detail":f"No Lead-level roles found in: {', '.join(no_lead_fns)}. These functions lack clear leadership ownership."})
            recs.append({"category":"Architecture","title":"Appoint function leadership for uncovered areas",
                "action":f"Define and fill Lead-level roles in {', '.join(no_lead_fns[:3])}. Clarify whether these functions roll up to a broader head, or whether dedicated leadership is needed for scale.",
                "impact":"Clear accountability, better function strategy execution, succession pipeline development."})

        # Pattern 4: Succession gaps
        lead_roles=[j for j in repo.jobs.values() if j.level=="Lead"]
        RELATED={"HR":{"HR","Operations","Legal"},"Finance":{"Finance","Operations","Legal"},
                 "Engineering":{"Engineering","Data","Product"},"Data":{"Data","Engineering","Product"},
                 "Product":{"Product","Engineering","Data"},"Operations":{"Operations","HR","Finance"},
                 "Sales":{"Sales","Marketing","Customer"},"Marketing":{"Marketing","Sales","Customer"},
                 "Customer":{"Customer","Sales","Operations"},"Legal":{"Legal","Finance","HR"}}
        role_pool={r.job_id:1 for r in self.results}
        critical_roles=[]
        for job in lead_roles:
            has_pipeline=any(1 for jid in role_pool
                if jid!=job.job_id
                and repo.jobs.get(jid)
                and repo.jobs[jid].function in RELATED.get(job.function,{job.function})
                and LSORT.get(job.level,4)-LSORT.get(repo.jobs[jid].level,"Junior",)<=2)
            if not has_pipeline:
                critical_roles.append(job.standard_title)
        if critical_roles:
            findings.append({"severity":"high","title":f"{len(critical_roles)} critical role(s) with no succession pipeline",
                "detail":f"Roles with no eligible successor in the talent pool: {', '.join(critical_roles[:4])}{'...' if len(critical_roles)>4 else ''}."})
            recs.append({"category":"Succession","title":"Address critical succession voids immediately",
                "action":f"For each critical role ({', '.join(critical_roles[:3])}), define a 12-month succession action: identify internal stretch candidates, begin targeted development, and consider parallel external search for highest-risk roles.",
                "impact":"Reduced key-person dependency, business continuity protection, board confidence."})

        # Pattern 5: Career path gaps (functions with Junior/Medior but no Senior)
        fn_has={}
        for r in self.results:
            fn_has.setdefault(r.function,set()).add(r.level)
        skipped=[fn for fn,lvls in fn_has.items()
                 if ("Junior" in lvls or "Medior" in lvls) and "Senior" not in lvls and "Lead" not in lvls]
        if skipped:
            findings.append({"severity":"medium","title":"Missing Senior career layer in some functions",
                "detail":f"Functions where Junior/Medior roles exist but no Senior layer: {', '.join(skipped)}. This creates a career ceiling and flight risk for developing talent."})
            recs.append({"category":"Development","title":"Build Senior career layer in underrepresented functions",
                "action":f"In {', '.join(skipped[:3])}, define Senior-level role profiles and create structured pathways from Medior to Senior. Consider internal promotion targets and competency-based progression criteria.",
                "impact":"Improved retention of top Medior talent, richer internal succession pool, lower external hiring costs."})

        # General recommendation: skills assessment
        findings.append({"severity":"low","title":"Role-assumed skill levels in use",
            "detail":"Skill gap analysis currently uses role-required levels as employee baselines. Uploading actual skill assessments will produce individual-level gap analysis and significantly sharper succession readiness scores."})
        recs.append({"category":"Development","title":"Collect individual skill assessments to sharpen succession scoring",
            "action":"Use the Skills Assessment template to collect manager-assessed or self-assessed skill levels for all employees. This transforms succession readiness from a role-proxy to an individual measure.",
            "impact":"Higher accuracy succession planning, personalised development plans, credible board reporting."})

        self._findings      = findings
        self._recommendations = recs

    def _compute_health_score(self) -> tuple[int,str]:
        """Compute a 0-100 org health score based on detectable patterns."""
        self._detect_patterns()
        score = 100
        for f in self._findings:
            score -= {"high":15,"medium":8,"low":3}.get(f["severity"],5)
        score = max(0,min(100,score))
        color = TEAL if score>=75 else (AMBER if score>=50 else RED)
        return score, color
