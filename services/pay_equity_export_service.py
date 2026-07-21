"""
jobsy/services/pay_equity_export_service.py

Turns a PayGapResult (services.pay_equity_service.analyze_gender_pay_gap) into a
formatted Excel workbook so a structural gender pay-gap run can leave the app as
a shareable report — mirrors services.export_service.ExportService, but for the
band-free Function x Level gap instead of matched titles.

Two entry points:

    to_workbook_bytes(result) -> bytes   (for Streamlit download_button)
    write_workbook(result, path) -> Path (to disk)

Workbook layout ("2.0" -- consolidated onto one page rather than four tabs,
per client feedback on the earlier per-sheet version):
    Summary   headline metrics, notes, a reliable-cohorts mean-pay mini-table
              feeding a Male-vs-Female bar chart, and both representation
              tables (by level, by function) -- all on one sheet, side by side
    Cohorts   every Function x Level cohort with both men and women

The export is a static snapshot of one analysis run, so it holds values rather
than formulas.

Sign convention: PayGapResult's own fields are "positive = men paid more".
This export instead reports "positive = women paid more" -- (vrouw - man) /
man -- to match the NL wetsvoorstel's own definition of "loonkloof", since
that's the wording this report gets checked against. The live Jobsy screen
applies the same flip (see ui/app.py::_render_leveled_gap) so the two never
disagree; see services.pay_equity_service.flip_gap_sign / flip_gap_ci.

Styling is on-brand: header fill is Jobsy's deep purple, and gap values are
colour-coded with the same danger/teal threshold logic as _render_leveled_gap's
on-screen _col(). Colours are literal hex here (not imported from ui/theme.py)
so this module -- pure pandas/openpyxl -- never needs streamlit importable;
keep the two in sync by hand if the Jobsy palette changes.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd

import logging
logger = logging.getLogger('jobsy')
from services.pay_equity_service import PayGapResult, DIRECTIVE_THRESHOLD_PCT, flip_gap_sign, flip_gap_ci

__all__ = ["PayEquityExportService"]

_GAP_PCT_FMT = '+0.0"%";-0.0"%"'  # values are already 0-100 scale, not 0-1 -- no native '%' format

# Jobsy brand palette -- hex without the '#', as openpyxl wants it.
_PURPLE = "53037F"    # header fill (deep purple; darker than ui/theme.py's on-screen primary,
                       # chosen for print/spreadsheet contrast)
_PINK = "FF73D0"      # accent -- Female chart series
_TEAL = "34B5FF"      # secondary -- on-screen "fine" gap colour
_DANGER = "FF5A7A"    # clay -- on-screen "flagged" gap colour
_DANGER_SOFT = "FDE3E9"
_TEAL_SOFT = "E2F2FF"

_FONT = "Arial"


class PayEquityExportService:
    """Renders a PayGapResult to a styled workbook."""

    # ------------------------------------------------------------- dataframes
    def summary_to_dataframe(self, r: PayGapResult) -> pd.DataFrame:
        ci_lo, ci_hi = flip_gap_ci(r.adjusted_ci) or (None, None)
        rows = [
            ("Rows received (before any exclusion)", r.n_input or r.n),
            ("  dropped — missing/zero salary or blank function/level", r.n_dropped_invalid),
            ("Employees in scope", r.n),
            ("Men (M)", r.n_m),
            ("Women (F)", r.n_f),
            ("Excluded from binary gap (non-binary / unknown gender)", r.n_excluded),
            ("Reconciliation: received − dropped = in scope",
             "OK" if (r.n_input or r.n) - r.n_dropped_invalid == r.n else "MISMATCH — investigate"),
            ("FTE-normalised", "Yes" if r.fte_normalised else "No"),
            (None, None),
            ("Mean gap % — unadjusted (+ = women paid more, per wetsvoorstel (vrouw-man)/man)",
             flip_gap_sign(r.mean_gap_pct)),
            ("Median gap % — unadjusted (+ = women paid more)", flip_gap_sign(r.median_gap_pct)),
            (None, None),
            ("Adjusted gap % (controls for function + level; + = women paid more)",
             flip_gap_sign(r.adjusted_gap_pct)),
            ("Adjusted 95% CI — low", ci_lo),
            ("Adjusted 95% CI — high", ci_hi),
            ("Adjusted gap statistically significant",
             None if r.adjusted_significant is None else ("Yes" if r.adjusted_significant else "No")),
            (None, None),
            ("Grade-assignment gap, in LEVELS not % (does gender predict the level itself, "
             "not just pay within it; + = women sit at a higher level)",
             flip_gap_sign(r.grade_gap_levels)),
            ("Grade-assignment 95% CI — low", (flip_gap_ci(r.grade_gap_ci) or (None, None))[0]),
            ("Grade-assignment 95% CI — high", (flip_gap_ci(r.grade_gap_ci) or (None, None))[1]),
            ("Grade-assignment gap statistically significant",
             None if r.grade_gap_significant is None else ("Yes" if r.grade_gap_significant else "No")),
            (None, None),
            ("Cohorts tested (Function x Level, both genders)", r.n_cohorts_tested),
            (f"Cohorts flagged (|gap| >= {DIRECTIVE_THRESHOLD_PCT:.0f}%)", r.n_cohorts_flagged),
            ("  ...of which reliable (n >= 5 each gender)", r.n_cohorts_flagged_reliable),
            ("Cohorts below the n>=5-per-gender reliability threshold (low-n, indicative only)",
             sum(1 for c in r.cohorts if not c.reliable)),
            ("Levels 100% one gender (no computable gap — absent from Cohorts; "
             "a segregation signal in its own right)",
             ", ".join(f"{lvl} ({g}, n={n})" for lvl, (g, n) in sorted(
                 r.single_gender_levels.items(), key=lambda kv: str(kv[0])))
             if r.single_gender_levels else "None"),
            (None, None),
            ("% women overall", r.pct_women_overall),
        ]
        return pd.DataFrame(rows, columns=["Metric", "Value"])

    def cohorts_to_dataframe(self, r: PayGapResult) -> pd.DataFrame:
        rows = [{
            "Function": c.function, "Level": c.level,
            "n Male": c.n_m, "n Female": c.n_f,
            "Mean M": c.mean_m, "Mean F": c.mean_f,
            "Median M": c.median_m, "Median F": c.median_f,
            "Mean gap % (+ = women paid more)": flip_gap_sign(c.mean_gap_pct),
            "Median gap % (+ = women paid more)": flip_gap_sign(c.median_gap_pct),
            f"Flagged (>= {DIRECTIVE_THRESHOLD_PCT:.0f}%)": "Yes" if c.flagged else "No",
            "Reliable (n>=5 each)": "Yes" if c.reliable else "No",
        } for c in r.cohorts]
        return pd.DataFrame(rows)

    def representation_to_dataframe(self, r: PayGapResult, by: str) -> pd.DataFrame:
        data = r.women_by_level if by == "level" else r.women_by_function
        label = "Level" if by == "level" else "Function"
        rows = [{label: k, "% women": v} for k, v in data.items()]
        return pd.DataFrame(rows, columns=[label, "% women"])

    def notes_list(self, r: PayGapResult) -> list[str]:
        sign_note = ("Sign convention: figures in this report are (vrouw - man) / man, i.e. "
                     "positive = women paid more -- matching the NL wetsvoorstel's definition of "
                     "'loonkloof'. The live Jobsy screen uses the same convention.")
        return [sign_note, *r.notes]

    # --------------------------------------------------------------- exporters
    def to_workbook_bytes(self, result: PayGapResult) -> bytes:
        summary = self.summary_to_dataframe(result)
        cohorts = self.cohorts_to_dataframe(result)

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # startcol=1 leaves column A as a narrow left margin, matching the rest of the sheet
            summary.to_excel(writer, sheet_name="Summary", index=False, startcol=1)
            if not cohorts.empty:
                cohorts.to_excel(writer, sheet_name="Cohorts", index=False)
            try:
                ws = writer.sheets["Summary"]
                self._format_summary_metrics(ws, len(summary), result)
                self._add_dashboard_section(ws, result, dashboard_row=len(summary) + 4)
                if not cohorts.empty:
                    self._format_data_sheet(writer.sheets["Cohorts"], cohorts)
            except Exception as exc:  # formatting must never break the export
                logger.warning("Pay-equity workbook formatting skipped: %s", exc)

        logger.info("Exported pay-equity report: n=%d (M=%d, F=%d), %d cohorts.",
                    result.n, result.n_m, result.n_f, result.n_cohorts_tested)
        return buffer.getvalue()

    def write_workbook(self, result: PayGapResult, path: str | Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(self.to_workbook_bytes(result))
        return path

    # ---------------------------------------------------------------- styling
    def _gap_color(self, value) -> str:
        """Same threshold logic as the on-screen _col() in ui/app.py::_render_leveled_gap.
        Guards on type, not just `is not None` -- pandas writes a blank Value cell
        (e.g. Adjusted gap % when the regression has too few rows) as '' once it
        shares a mixed-type column with strings like FTE-normalised's 'Yes'/'No',
        not as a true empty cell, so `value` can be '' here rather than None."""
        if not isinstance(value, (int, float)):
            return _TEAL
        return _DANGER if abs(value) >= DIRECTIVE_THRESHOLD_PCT else _TEAL

    def _style_header_cells(self, cells) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        font = Font(name=_FONT, bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor=_PURPLE)
        center = Alignment(horizontal="center", vertical="center")
        for cell in cells:
            cell.font = font
            cell.fill = fill
            cell.alignment = center

    def _header_row_style(self, ws, row: int = 1):
        self._style_header_cells(ws[row])

    def _format_summary_metrics(self, ws, n_metric_rows: int, r: PayGapResult | None = None) -> None:
        """Metrics live in columns B (Metric) / C (Value); column A is a left margin."""
        from openpyxl.styles import Font
        self._style_header_cells((ws.cell(1, 2), ws.cell(1, 3)))
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 4.5
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 16
        # Grade-assignment gap is in LEVELS, not %, and its own threshold is
        # statistical significance (not the 5% pct-gap trigger _gap_color uses),
        # so it gets a distinct colour decision rather than reusing that logic.
        grade_gap_color = _DANGER if (r is not None and r.grade_gap_significant) else _TEAL
        for row in range(2, n_metric_rows + 2):
            label_cell = ws.cell(row=row, column=2)
            label = label_cell.value
            if label is None:
                continue
            value_cell = ws.cell(row=row, column=3)
            if "gap %" in str(label).lower() or "% women" in str(label).lower():
                value_cell.number_format = _GAP_PCT_FMT
            if str(label).startswith(("Mean gap", "Median gap", "Adjusted gap %")):
                value_cell.font = Font(name=_FONT, bold=True, color=self._gap_color(value_cell.value))
            elif str(label).startswith("Grade-assignment gap, in LEVELS"):
                value_cell.number_format = '+0.00;-0.00'
                value_cell.font = Font(name=_FONT, bold=True, color=grade_gap_color)

    def _add_dashboard_section(self, ws, r: PayGapResult, dashboard_row: int) -> None:
        """Below the metrics: a title row, then one shared header row for four
        side-by-side blocks -- Notes (B), reliable-cohorts mean-pay mini-table
        feeding a bar chart (E:G), representation by level (I:J), representation
        by function (K:L) -- each sized to its own content."""
        from openpyxl.styles import Font

        title_cell = ws.cell(dashboard_row, 2, "Reliable cohorts — mean pay by gender (n>=5 each)")
        title_cell.font = Font(name=_FONT, bold=True)

        header_row = dashboard_row + 1
        data_row = header_row + 1

        headers = {2: "Notes", 5: "Function x Level", 6: "Mean M", 7: "Mean F",
                   9: "Level", 10: "% women", 11: "Function", 12: "% women"}
        for col, label in headers.items():
            ws.cell(header_row, col, label)
        self._style_header_cells([ws.cell(header_row, col) for col in headers])

        notes = self.notes_list(r)
        for i, note in enumerate(notes):
            ws.cell(data_row + i, 2, note)

        reliable = [c for c in r.cohorts if c.reliable]
        if reliable:
            for i, c in enumerate(reliable):
                row = data_row + i
                ws.cell(row, 5, f"{c.function}-{c.level}")
                ws.cell(row, 6, c.mean_m).number_format = "#,##0"
                ws.cell(row, 7, c.mean_f).number_format = "#,##0"
            self._add_reliable_chart(ws, header_row=header_row, n_reliable=len(reliable))
        else:
            note_cell = ws.cell(data_row, 5, "No cohort has a reliable (n>=5 each gender) sample -- no chart to show.")
            note_cell.font = Font(name=_FONT, italic=True, color="6B7684")

        by_level = self.representation_to_dataframe(r, "level")
        for i, row in enumerate(by_level.itertuples(index=False)):
            ws.cell(data_row + i, 9, row[0])
            ws.cell(data_row + i, 10, row[1])

        by_function = self.representation_to_dataframe(r, "function")
        for i, row in enumerate(by_function.itertuples(index=False)):
            ws.cell(data_row + i, 11, row[0])
            ws.cell(data_row + i, 12, row[1])

        ws.column_dimensions["D"].width = 3.5
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 3.5
        ws.column_dimensions["I"].width = 9
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 10

    def _add_reliable_chart(self, ws, *, header_row: int, n_reliable: int) -> None:
        """Clustered Male-vs-Female bar chart, sourced from the mini-table this
        method assumes is already written at columns E:G starting at header_row,
        anchored near the top of the sheet beside the headline metrics. Sized and
        positioned to match the client's own template (~18 x 10.3cm at D2) so it
        doesn't run down into the dashboard section below."""
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList

        last_row = header_row + n_reliable
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Mean pay by cohort, Male vs Female (reliable cohorts only)"
        chart.y_axis.title = "Salary"
        chart.x_axis.title = "Function x Level"
        chart.height, chart.width = 10.3, 18
        chart.legend.position = "b"
        cats = Reference(ws, min_col=5, min_row=header_row + 1, max_row=last_row)
        data = Reference(ws, min_col=6, max_col=7, min_row=header_row, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = _PURPLE  # Mean M
        chart.series[1].graphicalProperties.solidFill = _PINK    # Mean F
        # Every show* flag left as None (the class default) renders in Excel as
        # "series name, category, value" all concatenated on one label -- huge,
        # overlapping text across the whole plot. Every flag must be explicit.
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showSerName = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showLegendKey = False
        chart.dataLabels.showPercent = False
        chart.dataLabels.showBubbleSize = False
        chart.dataLabels.dLblPos = "outEnd"
        chart.dataLabels.numFmt = "#,##0"
        ws.add_chart(chart, "D2")

    def _format_data_sheet(self, ws, df: pd.DataFrame) -> None:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        self._header_row_style(ws)
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        status_fill = {
            "Yes": PatternFill("solid", fgColor=_DANGER_SOFT),
            "No": PatternFill("solid", fgColor=_TEAL_SOFT),
        }
        for idx, header in enumerate(df.columns, start=1):
            lengths = df[header].fillna("").astype(str).map(len)
            longest = int(lengths.max()) if len(lengths) else 0
            width = min(60, max(len(str(header)) + 2, longest + 2))
            ws.column_dimensions[get_column_letter(idx)].width = width
            header_l = str(header).lower()
            if "gap %" in header_l:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=idx)
                    cell.number_format = _GAP_PCT_FMT
                    cell.font = Font(name=_FONT, color=self._gap_color(cell.value))
            elif header_l.startswith("flagged"):
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=idx)
                    cell.fill = status_fill.get(cell.value, status_fill["No"])
