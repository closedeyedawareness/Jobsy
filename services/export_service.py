"""
jobsy/services/export_service.py

Turns a batch of MatchResult objects into a formatted Excel workbook so matched
results can leave the app. Two entry points:

    to_workbook_bytes(results, summary)  -> bytes   (for Streamlit download_button)
    write_workbook(results, path, summary) -> Path   (to disk)

Workbook layout:
    Matches       every input title with its resolved role, confidence, salary
    Needs Review  only the rows flagged requires_review (low confidence / no match)
    Summary       run totals + match-type breakdown

The export is a static snapshot of one matching run, so it holds values rather
than formulas.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Optional, Sequence

import pandas as pd

from jobsy.core.logger import logger
from jobsy.services.matching_service import MatchingSummary, MatchResult

__all__ = ["ExportService"]

# internal key -> workbook header
_COLUMNS = [
    ("input_title", "Input Title"),
    ("standard_title", "Matched Role"),
    ("function", "Function"),
    ("level", "Level"),
    ("match_type", "Match Type"),
    ("confidence", "Confidence"),
    ("requires_review", "Needs Review"),
    ("salary_min", "Salary Min"),
    ("salary_max", "Salary Max"),
    ("currency", "Currency"),
    ("job_id", "Job ID"),
    ("description", "Description"),
]
_HEADERS = {k: v for k, v in _COLUMNS}
_SALARY_HEADERS = {"Salary Min", "Salary Max"}
_REVIEW_HEADER = "Needs Review"


class ExportService:
    """Renders MatchResult batches to a styled workbook."""

    # ------------------------------------------------------------- dataframes
    def results_to_dataframe(self, results: Sequence[MatchResult]) -> pd.DataFrame:
        rows = [
            {
                "input_title": r.input_title,
                "standard_title": r.standard_title,
                "function": r.function,
                "level": r.level,
                "match_type": r.match_type.value,
                "confidence": r.confidence,
                "requires_review": bool(r.requires_review),
                "salary_min": r.salary_min,
                "salary_max": r.salary_max,
                "currency": r.currency if r.matched else None,
                "job_id": r.job_id,
                "description": r.description,
            }
            for r in results
        ]
        df = pd.DataFrame(rows, columns=[k for k, _ in _COLUMNS])
        return df.rename(columns=_HEADERS)

    def _summary_dataframe(self, summary: MatchingSummary) -> pd.DataFrame:
        rows = [
            ("Total titles", summary.total),
            ("Matched", summary.matched),
            ("Needs review", summary.review),
            ("Unmatched", summary.unmatched),
            ("Average confidence", summary.avg_confidence),
        ]
        rows += [(f"  via {k}", v) for k, v in sorted(summary.by_type.items())]
        return pd.DataFrame(rows, columns=["Metric", "Value"])

    # --------------------------------------------------------------- exporters
    def to_workbook_bytes(
        self,
        results: Sequence[MatchResult],
        summary: Optional[MatchingSummary] = None,
    ) -> bytes:
        matches = self.results_to_dataframe(results)
        review = matches[matches[_REVIEW_HEADER]].reset_index(drop=True)

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            matches.to_excel(writer, sheet_name="Matches", index=False)
            review.to_excel(writer, sheet_name="Needs Review", index=False)
            if summary is not None:
                self._summary_dataframe(summary).to_excel(
                    writer, sheet_name="Summary", index=False
                )
            try:
                self._format_sheet(writer.sheets["Matches"], matches)
                self._format_sheet(writer.sheets["Needs Review"], review)
                if summary is not None:
                    self._format_sheet(writer.sheets["Summary"], None, simple=True)
            except Exception as exc:  # formatting must never break the export
                logger.warning("Workbook formatting skipped: %s", exc)

        logger.info("Exported %d results (%d flagged for review).", len(matches), len(review))
        return buffer.getvalue()

    def write_workbook(
        self,
        results: Sequence[MatchResult],
        path: str | Path,
        summary: Optional[MatchingSummary] = None,
    ) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(self.to_workbook_bytes(results, summary))
        return path

    # ---------------------------------------------------------------- styling
    def _format_sheet(self, ws, df: Optional[pd.DataFrame], *, simple: bool = False) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0E7C66")
        review_fill = PatternFill("solid", fgColor="F7EEDD")
        center = Alignment(horizontal="center", vertical="center")

        # header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        if simple or df is None:
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 24
            return

        # column widths from content (capped) + salary number format
        headers = list(df.columns)
        review_idx = headers.index(_REVIEW_HEADER) + 1 if _REVIEW_HEADER in headers else None
        for idx, header in enumerate(headers, start=1):
            lengths = df[header].fillna("").astype(str).map(len)
            longest = int(lengths.max()) if len(lengths) else 0
            width = min(60, max(len(header) + 2, longest + 2))
            ws.column_dimensions[get_column_letter(idx)].width = width
            if header in _SALARY_HEADERS:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).number_format = '#,##0'

        # highlight rows that need review
        if review_idx is not None:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=review_idx).value:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = review_fill
